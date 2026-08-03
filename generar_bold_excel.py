import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.odoo_utils import get_odoo_models, ODOO_DB, ODOO_PASSWORD

uid, models, err = get_odoo_models()

claves = ['JE537','4E013','BF149','LD648','LD664','ND728','FA271','HA433','HF427']

partners_odoo = models.execute_kw(ODOO_DB, uid, ODOO_PASSWORD, 'res.partner', 'search_read',
    [[]], {'fields': ['id', 'name', 'ref', 'parent_id']})

clave_to_ids = {c: [] for c in claves}
id_to_clave  = {}
id_to_nombre = {}

for p in partners_odoo:
    ref = str(p.get('ref') or '').strip().upper()
    for c in claves:
        if ref == c:
            clave_to_ids[c].append(p['id'])
            id_to_clave[p['id']]  = c
            id_to_nombre[p['id']] = p['name']
            break

for p in partners_odoo:
    if p['id'] not in id_to_clave and p.get('parent_id'):
        pid = p['parent_id'][0] if isinstance(p['parent_id'], (list, tuple)) else p['parent_id']
        if pid in id_to_clave:
            clave_to_ids[id_to_clave[pid]].append(p['id'])
            id_to_clave[p['id']]  = id_to_clave[pid]
            id_to_nombre[p['id']] = p['name']

all_ids = list(id_to_clave.keys())

lines = models.execute_kw(ODOO_DB, uid, ODOO_PASSWORD, 'account.move.line', 'search_read',
    [[
        ('move_id.move_type', '=', 'out_invoice'),
        ('move_id.state', '=', 'posted'),
        ('move_id.invoice_date', '>=', '2025-05-01'),
        ('move_id.invoice_date', '<=', '2026-06-03'),
        ('display_type', '=', 'product'),
        ('partner_id', 'in', all_ids),
        '|',
            ('product_id.product_tmpl_id.name', 'ilike', 'BOLD'),
            ('name', 'ilike', 'BOLD'),
    ]],
    {'fields': ['move_id','partner_id','name','product_id','quantity',
                'price_unit','discount','price_subtotal','price_total','date']}
)

# Categorias
product_ids_set = set(l['product_id'][0] for l in lines if l.get('product_id'))
products_info = {}
if product_ids_set:
    prods = models.execute_kw(ODOO_DB, uid, ODOO_PASSWORD, 'product.product', 'search_read',
        [[('id', 'in', list(product_ids_set))]],
        {'fields': ['id', 'product_tmpl_id']})
    tmpl_ids = [p['product_tmpl_id'][0] for p in prods if p.get('product_tmpl_id')]
    tmpls = models.execute_kw(ODOO_DB, uid, ODOO_PASSWORD, 'product.template', 'search_read',
        [[('id', 'in', tmpl_ids)]],
        {'fields': ['id', 'categ_id']})
    tmpl_cat = {t['id']: (t['categ_id'][1] if t.get('categ_id') else 'Sin categoria') for t in tmpls}
    prod_to_tmpl = {p['id']: p['product_tmpl_id'][0] for p in prods if p.get('product_tmpl_id')}
    for pid in product_ids_set:
        tmpl = prod_to_tmpl.get(pid)
        products_info[pid] = tmpl_cat.get(tmpl, 'Sin categoria') if tmpl else 'Sin categoria'

# Info facturas
move_ids = list({l['move_id'][0] for l in lines if l.get('move_id')})
moves_info = {}
if move_ids:
    mvs = models.execute_kw(ODOO_DB, uid, ODOO_PASSWORD, 'account.move', 'search_read',
        [[('id', 'in', move_ids)]],
        {'fields': ['id', 'name', 'invoice_origin', 'partner_id', 'ref']})
    for m in mvs:
        moves_info[m['id']] = {
            'factura': m['name'],
            'orden':   m.get('invoice_origin') or '',
            'cliente_factura': m['partner_id'][1] if m.get('partner_id') else '',
            'referencia': m.get('ref') or ''
        }

# ---- EXCEL ----
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOLD - Compras por Cliente"

COLOR_HEADER    = "1F2937"
COLOR_BICI      = "D1FAE5"
COLOR_ACCESORIO = "FEF3C7"
COLOR_ALT       = "F9FAFB"

hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
body_font = Font(name="Calibri", size=10)
thin  = Side(border_style="thin", color="D1D5DB")
borde = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = [
    "CLAVE", "CLIENTE (REF ODOO)", "FACTURA", "ORDEN VENTA",
    "FECHA", "SKU / PRODUCTO", "CATEGORIA", "CANTIDAD",
    "PRECIO UNITARIO", "DESCUENTO %", "SUBTOTAL SIN IVA", "TOTAL CON IVA",
    "ES BICICLETA"
]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = hdr_font
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = borde

ws.row_dimensions[1].height = 30

row = 2
for l in sorted(lines, key=lambda x: (
    id_to_clave.get(x['partner_id'][0] if x['partner_id'] else 0, ''),
    str(x['date'])
)):
    pid       = l['partner_id'][0] if l['partner_id'] else 0
    clave     = id_to_clave.get(pid, '?')
    nombre    = id_to_nombre.get(pid, '')
    mv        = moves_info.get(l['move_id'][0] if l['move_id'] else 0, {})
    prod_id   = l['product_id'][0] if l.get('product_id') else None
    categoria = products_info.get(prod_id, 'Sin categoria')
    nombre_linea = l.get('name', '') or ''
    es_bici   = 'BICICLETA' in nombre_linea.upper() or 'BICICLETA' in categoria.upper()

    valores = [
        clave,
        nombre,
        mv.get('factura',''),
        mv.get('orden',''),
        str(l['date'] or ''),
        nombre_linea,
        categoria,
        l['quantity'],
        l['price_unit'],
        l.get('discount', 0),
        l['price_subtotal'],
        l['price_total'],
        'SI' if es_bici else 'NO'
    ]

    color_fila = COLOR_BICI if es_bici else (COLOR_ACCESORIO if row % 2 == 0 else "FFFFFF")

    for col, val in enumerate(valores, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = body_font
        cell.fill = PatternFill("solid", fgColor=color_fila)
        cell.border = borde
        cell.alignment = Alignment(vertical="center", wrap_text=(col == 6))
        if col in [9, 11, 12]:
            cell.number_format = '#,##0.00'
        if col == 10:
            cell.number_format = '0.0"%"'
    row += 1

anchos = [9, 35, 22, 16, 12, 55, 35, 9, 16, 12, 16, 16, 14]
for i, ancho in enumerate(anchos, 1):
    ws.column_dimensions[get_column_letter(i)].width = ancho

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

ruta = r"C:\Users\Aux Sistemas\Desktop\BOLD_Compras_Clientes.xlsx"
wb.save(ruta)
print(f"Excel guardado: {ruta}")
print(f"Total registros: {row - 1}")
