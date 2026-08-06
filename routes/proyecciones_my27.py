"""
Módulo: Proyecciones MY27
Vista consolidada de los 92 artículos MY27 con cantidades por mes
sumando todas las proyecciones de todos los distribuidores.
"""

import io
import logging
import re
import time
from datetime import datetime
from utils.tiempo import ahora_mx, ahora_str

from flask import Blueprint, jsonify, request, send_file

from db_conexion import obtener_conexion
from routes.forecast import (SKU_CATALOG, FORECAST_SKU_WHITELIST,
                             _SCOTT_CORRECT_NAMES, _SCOTT_COLORS, _SCOTT_TALLAS,
                             _ensure_scott_names, _redis_get, _redis_set)
from utils.odoo_utils import get_odoo_models, ODOO_DB, ODOO_PASSWORD

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

proyecciones_my27_bp = Blueprint('proyecciones_my27', __name__, url_prefix='/proyecciones-my27')

MESES       = ['mayo', 'junio', 'julio', 'agosto', 'septiembre',
               'octubre', 'noviembre', 'diciembre', 'enero', 'febrero', 'marzo', 'abril']
MESES_LABEL = ['May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic', 'Ene', 'Feb', 'Mar', 'Abr']
MESES_ORDEN = MESES  # alias — mismo orden cronológico MY27

_MY27_R_TTL  = 600  # 10 min Redis — monitor completo

def _rkey_my27(periodo: str) -> str:
    return f'proy_my27:{periodo}'

_COSTOS_CACHE: dict = {'data': {}, 'ts': 0.0}
_COSTOS_TTL = 300  # 5 minutos

_MEGAMO_PRECIOS_CACHE: dict = {'data': {}, 'ts': 0.0}
_MEGAMO_PRECIOS_TTL  = 300

_ORDENES_CACHE: dict = {'data': {}, 'periodo': '', 'ts': 0.0}
_ORDENES_TTL = 180  # 3 minutos

# Mapeo nivel → pricelist ID en Odoo
_NIVEL_TO_PL_ID: dict = {
    'Distribuidor':        13,
    'Partner':             11,
    'Partner Elite':        9,
    'Partner Elite Plus!': 37,
}
# Factor sobre precio público cuando no hay pricelist específico
_NIVEL_FACTORS: dict = {
    'Distribuidor':        0.760,
    'Partner':             0.740,
    'Partner Elite':       0.715,
    'Partner Elite Plus!': 0.695,
}


def _get_megamo_precios_por_nivel() -> dict:
    """
    Retorna {sku: {nivel: precio_sin_iva}} para todos los SKUs Megamo.
    Prioridad: pricelist Odoo → precio_publico × factor. Cacheado 5 min.
    """
    global _MEGAMO_PRECIOS_CACHE
    now = time.time()
    if _MEGAMO_PRECIOS_CACHE['data'] and (now - _MEGAMO_PRECIOS_CACHE['ts']) < _MEGAMO_PRECIOS_TTL:
        return _MEGAMO_PRECIOS_CACHE['data']

    try:
        uid, models, err = get_odoo_models()
        if not uid:
            logging.warning('[megamo_precios] no Odoo: %s', err)
            return _MEGAMO_PRECIOS_CACHE['data']

        megamo_skus = [s for s in FORECAST_SKU_WHITELIST if s.startswith('MH')]

        prods = models.execute_kw(ODOO_DB, uid, ODOO_PASSWORD,
            'product.product', 'search_read',
            [[['default_code', 'in', megamo_skus]]],
            {'fields': ['id', 'default_code'], 'limit': 0})
        prod_id_map = {p['default_code']: p['id'] for p in prods}
        id_to_sku   = {v: k for k, v in prod_id_map.items()}
        prod_ids    = list(prod_id_map.values())

        # Precios públicos (pricelist 4) como base de fallback
        pl4 = models.execute_kw(ODOO_DB, uid, ODOO_PASSWORD,
            'product.pricelist.item', 'search_read',
            [[['pricelist_id', '=', 4], ['product_id', 'in', prod_ids]]],
            {'fields': ['product_id', 'fixed_price'], 'limit': 0})
        pub_map = {id_to_sku[i['product_id'][0]]: float(i['fixed_price'])
                   for i in pl4 if i['product_id'][0] in id_to_sku}

        # Precios por pricelist (Distribuidor=13, Partner=11, Partner Elite=37)
        pl_sku_prices: dict = {}
        for nivel, pl_id in _NIVEL_TO_PL_ID.items():
            if pl_id is None:
                continue
            items = models.execute_kw(ODOO_DB, uid, ODOO_PASSWORD,
                'product.pricelist.item', 'search_read',
                [[['pricelist_id', '=', pl_id], ['product_id', 'in', prod_ids]]],
                {'fields': ['product_id', 'fixed_price'], 'limit': 0})
            for item in items:
                sku = id_to_sku.get(item['product_id'][0])
                if sku:
                    if pl_id not in pl_sku_prices:
                        pl_sku_prices[pl_id] = {}
                    pl_sku_prices[pl_id][sku] = float(item['fixed_price'])

        # Construir resultado: {sku: {nivel: precio_sin_iva}}
        result: dict = {}
        for sku in megamo_skus:
            result[sku] = {}
            pub = pub_map.get(sku, 0.0)
            for nivel, factor in _NIVEL_FACTORS.items():
                pl_id = _NIVEL_TO_PL_ID.get(nivel)
                if pl_id and pl_id in pl_sku_prices and sku in pl_sku_prices[pl_id]:
                    result[sku][nivel] = pl_sku_prices[pl_id][sku]
                elif pub > 0:
                    result[sku][nivel] = round(pub * factor, 2)
                else:
                    result[sku][nivel] = 0.0

        _MEGAMO_PRECIOS_CACHE = {'data': result, 'ts': now}
        logging.info('[megamo_precios] %d SKUs con precios por nivel', len(result))
        return result

    except Exception as e:
        logging.warning('[megamo_precios] error: %s', e)
        return _MEGAMO_PRECIOS_CACHE['data']


def _get_costos_odoo() -> dict:
    """Devuelve {sku: standard_price} para todos los SKUs MY27. Cacheado 5 min."""
    global _COSTOS_CACHE
    now = time.time()
    if _COSTOS_CACHE['data'] and (now - _COSTOS_CACHE['ts']) < _COSTOS_TTL:
        return _COSTOS_CACHE['data']

    try:
        uid, models, err = get_odoo_models()
        if not uid:
            logging.warning('[costos_odoo] no se pudo conectar: %s', err)
            return _COSTOS_CACHE['data']

        skus = list(FORECAST_SKU_WHITELIST)

        # Búsqueda en product.product (variantes) — standard_price con list_price como fallback
        productos_pp = models.execute_kw(ODOO_DB, uid, ODOO_PASSWORD,
            'product.product', 'search_read',
            [[['default_code', 'in', skus]]],
            {'fields': ['default_code', 'standard_price', 'list_price'], 'limit': 0})

        # Normalizar: strip + mapear código → costo.
        # Umbral mínimo de $100 para descartar placeholders (ej. $1.00 en Odoo).
        # Prioridad: standard_price (si ≥ 100) → list_price → 0
        _MIN_PRECIO = 100.0
        costos: dict = {}
        for p in productos_pp:
            code = (p.get('default_code') or '').strip()
            cost = float(p.get('standard_price') or 0)
            lst  = float(p.get('list_price') or 0)
            precio = cost if cost >= _MIN_PRECIO else lst
            if code and precio >= _MIN_PRECIO:
                costos[code] = precio

        # Para los SKUs sin precio, intentar product.template como fallback
        skus_sin_costo = [s for s in skus if s not in costos]
        if skus_sin_costo:
            productos_pt = models.execute_kw(ODOO_DB, uid, ODOO_PASSWORD,
                'product.template', 'search_read',
                [[['default_code', 'in', skus_sin_costo]]],
                {'fields': ['default_code', 'standard_price', 'list_price'], 'limit': 0})
            for p in productos_pt:
                code = (p.get('default_code') or '').strip()
                cost = float(p.get('standard_price') or 0)
                lst  = float(p.get('list_price') or 0)
                precio = cost if cost >= _MIN_PRECIO else lst
                if code and precio >= _MIN_PRECIO and code not in costos:
                    costos[code] = precio

        # Fallback para SKUs Megamo sin precio: usar pricelist 13 (DISTRIBUIDOR) sin IVA
        megamo_sin_precio = [s for s in skus if s.startswith('MH') and s not in costos]
        if megamo_sin_precio:
            prods_mh = models.execute_kw(ODOO_DB, uid, ODOO_PASSWORD,
                'product.product', 'search_read',
                [[['default_code', 'in', megamo_sin_precio]]],
                {'fields': ['id', 'default_code'], 'limit': 0})
            mh_id_map = {p['id']: p['default_code'] for p in prods_mh}
            if mh_id_map:
                pl13_items = models.execute_kw(ODOO_DB, uid, ODOO_PASSWORD,
                    'product.pricelist.item', 'search_read',
                    [[['pricelist_id', '=', 13], ['product_id', 'in', list(mh_id_map.keys())]]],
                    {'fields': ['product_id', 'fixed_price'], 'limit': 0})
                for item in pl13_items:
                    code  = mh_id_map.get(item['product_id'][0])
                    price = float(item.get('fixed_price') or 0)
                    if code and price > 0:
                        costos[code] = price

        _COSTOS_CACHE = {'data': costos, 'ts': now}
        logging.info('[costos_odoo] %d/%d SKUs con costo (pp=%d, pt=%d, megamo_pl13=%d)',
                     len(costos), len(skus),
                     sum(1 for p in productos_pp if float(p.get('standard_price') or 0) > 0),
                     len([s for s in skus_sin_costo if s in costos]),
                     len(megamo_sin_precio) - len([s for s in megamo_sin_precio if s not in costos]))
        return costos

    except Exception as e:
        logging.warning('[costos_odoo] error: %s', e)
        return _COSTOS_CACHE['data']


# ─────────────────────────────────────────────────────────────────────────────
# Helper: órdenes confirmadas en Odoo por distribuidor
# ─────────────────────────────────────────────────────────────────────────────

def _norm_sku(s: str) -> str:
    return re.sub(r'[\-\s]', '', str(s or '')).upper()


def _get_ordenes_my27(periodo: str) -> dict:
    """
    Retorna {clave_cliente: {sku_norm: total_pedido}} con las unidades ya
    confirmadas en Odoo (sale.order state=sale/done) para el periodo MY27.
    Caché de 3 minutos para no saturar Odoo en cada recarga.
    """
    global _ORDENES_CACHE
    now = time.time()
    if (now - _ORDENES_CACHE['ts'] < _ORDENES_TTL and
            _ORDENES_CACHE['periodo'] == periodo):
        return _ORDENES_CACHE['data']

    try:
        m = re.match(r'^(\d{4})-(\d{4})$', periodo)
        if not m:
            return {}
        year1, year2 = int(m.group(1)), int(m.group(2))
        fecha_inicio = f'{year1 - 1}-07-01'
        fecha_fin    = f'{year2}-04-30'

        uid, models, err = get_odoo_models()
        if err or not uid:
            return _ORDENES_CACHE['data']

        # 1. Todas las órdenes confirmadas del periodo
        orders = models.execute_kw(
            ODOO_DB, uid, ODOO_PASSWORD,
            'sale.order', 'search_read',
            [[['state', 'in', ['sale', 'done']],
              ['date_order', '>=', fecha_inicio],
              ['date_order', '<=', fecha_fin + ' 23:59:59']]],
            {'fields': ['id', 'partner_id'], 'limit': 0}
        )
        if not orders:
            _ORDENES_CACHE = {'data': {}, 'periodo': periodo, 'ts': now}
            return {}

        # 2. Refs de los partners
        partner_ids = list({o['partner_id'][0] for o in orders if o.get('partner_id')})
        partners = models.execute_kw(
            ODOO_DB, uid, ODOO_PASSWORD,
            'res.partner', 'search_read',
            [[['id', 'in', partner_ids], ['ref', '!=', False]]],
            {'fields': ['id', 'ref'], 'limit': 0}
        )
        ref_map = {p['id']: (p.get('ref') or '').strip() for p in partners}

        # 3. order_id → clave_cliente
        order_clave = {
            o['id']: ref_map[o['partner_id'][0]]
            for o in orders
            if o.get('partner_id') and ref_map.get(o['partner_id'][0])
        }

        # 4. Líneas de venta
        sol = models.execute_kw(
            ODOO_DB, uid, ODOO_PASSWORD,
            'sale.order.line', 'search_read',
            [[['order_id', 'in', list(order_clave.keys())],
              ['state', 'not in', ['cancel']]]],
            {'fields': ['order_id', 'product_id', 'product_uom_qty'], 'limit': 0}
        )

        # 5. Códigos de producto
        prod_ids = list({l['product_id'][0] for l in sol if l.get('product_id')})
        prods = models.execute_kw(
            ODOO_DB, uid, ODOO_PASSWORD,
            'product.product', 'search_read',
            [[['id', 'in', prod_ids]]],
            {'fields': ['id', 'default_code'], 'limit': 0}
        )
        prod_code = {p['id']: (p.get('default_code') or '').strip() for p in prods}

        # 6. Agregar por clave + sku_norm
        result: dict = {}
        for l in sol:
            oid = l['order_id'][0] if isinstance(l.get('order_id'), list) else l.get('order_id')
            clave = order_clave.get(oid, '')
            if not clave:
                continue
            pid = l['product_id'][0] if l.get('product_id') else None
            if not pid:
                continue
            sn = _norm_sku(prod_code.get(pid, ''))
            if not sn:
                continue
            qty = int(l.get('product_uom_qty') or 0)
            if clave not in result:
                result[clave] = {}
            result[clave][sn] = result[clave].get(sn, 0) + qty

        _ORDENES_CACHE = {'data': result, 'periodo': periodo, 'ts': now}
        logging.info('[ordenes_my27] %d clientes con pedidos en %s', len(result), periodo)
        return result

    except Exception as e:
        logging.exception('[ordenes_my27] error: %s', e)
        return _ORDENES_CACHE['data']


# ─────────────────────────────────────────────────────────────────────────────
# Helper: obtener datos consolidados
# ─────────────────────────────────────────────────────────────────────────────

def _get_datos_consolidados(periodo: str = '') -> dict:
    """
    Consulta la BD y devuelve los 92 artículos con sumas por mes
    de todos los distribuidores y el desglose individual.
    """
    conn = obtener_conexion()
    cur  = conn.cursor(dictionary=True)

    where  = "WHERE periodo = %s" if periodo else ""
    params = (periodo,) if periodo else ()

    try:
        # Totales por SKU
        cur.execute(f"""
            SELECT
                sku,
                MAX(producto)   AS producto,
                MAX(marca)      AS marca,
                MAX(modelo)     AS modelo,
                MAX(color)      AS color,
                MAX(talla)      AS talla,
                SUM(mayo)       AS mayo,
                SUM(junio)      AS junio,
                SUM(julio)      AS julio,
                SUM(agosto)     AS agosto,
                SUM(septiembre) AS septiembre,
                SUM(octubre)    AS octubre,
                SUM(noviembre)  AS noviembre,
                SUM(diciembre)  AS diciembre,
                SUM(enero)      AS enero,
                SUM(febrero)    AS febrero,
                SUM(marzo)      AS marzo,
                SUM(abril)      AS abril,
                SUM(mayo+junio+julio+agosto+septiembre+
                    octubre+noviembre+diciembre+
                    enero+febrero+marzo+abril) AS total_anual,
                COUNT(DISTINCT CASE WHEN (mayo+junio+julio+agosto+septiembre+octubre+noviembre+diciembre+enero+febrero+marzo+abril) > 0 THEN clave_cliente END) AS num_distribuidores
            FROM forecast_proyecciones
            {where}
            GROUP BY sku
        """, params)
        totales_map = {r['sku']: r for r in cur.fetchall()}

        # Desglose por SKU + distribuidor — solo los que tienen al menos 1 unidad
        _total_expr = ("(fp.mayo+fp.junio+fp.julio+fp.agosto+fp.septiembre+"
                       "fp.octubre+fp.noviembre+fp.diciembre+"
                       "fp.enero+fp.febrero+fp.marzo+fp.abril) > 0")
        where_fp = f"WHERE fp.periodo = %s AND {_total_expr}" if periodo else f"WHERE {_total_expr}"
        cur.execute(f"""
            SELECT
                fp.sku, fp.clave_cliente,
                COALESCE(c.nombre_cliente, fp.clave_cliente) AS nombre_cliente,
                fp.mayo, fp.junio, fp.julio, fp.agosto, fp.septiembre,
                fp.octubre, fp.noviembre, fp.diciembre,
                fp.enero, fp.febrero, fp.marzo, fp.abril,
                (fp.mayo+fp.junio+fp.julio+fp.agosto+fp.septiembre+
                 fp.octubre+fp.noviembre+fp.diciembre+
                 fp.enero+fp.febrero+fp.marzo+fp.abril) AS total_dist
            FROM forecast_proyecciones fp
            LEFT JOIN clientes c ON c.clave = fp.clave_cliente
            {where_fp}
            ORDER BY fp.sku, fp.clave_cliente
        """, params)
        desglose_map: dict = {}
        for fd in cur.fetchall():
            s = fd['sku']
            if s not in desglose_map:
                desglose_map[s] = []
            desglose_map[s].append({
                'clave_cliente':  fd['clave_cliente'],
                'nombre_cliente': fd['nombre_cliente'],
                'total':          int(fd['total_dist'] or 0),
                'meses':          {mes: int(fd[mes] or 0) for mes in MESES},
            })

        # Descontar unidades ya pedidas en Odoo (FIFO por mes) de cada desglose
        ordenes_dist = _get_ordenes_my27(periodo) if periodo else {}
        if ordenes_dist:
            for sku, desgloses in desglose_map.items():
                sn = _norm_sku(sku)
                for d in desgloses:
                    total_pedido = ordenes_dist.get(d['clave_cliente'], {}).get(sn, 0)
                    if total_pedido <= 0:
                        continue
                    restante = total_pedido
                    for mes in MESES:
                        qty = d['meses'].get(mes, 0)
                        deducido = min(restante, qty)
                        d['meses'][mes] = qty - deducido
                        restante -= deducido
                        if restante <= 0:
                            break
                    d['total'] = sum(d['meses'][m] for m in MESES)
                # Eliminar distribuidores cuyo total quedó en 0 tras descontar
                desglose_map[sku] = [d for d in desgloses if d['total'] > 0]

        # Nivel de cada distribuidor que tiene Megamo con unidades
        megamo_claves = {
            d['clave_cliente']
            for sku, desgloses in desglose_map.items()
            if sku.startswith('MH')
            for d in desgloses
        }
        nivel_map: dict = {}
        if megamo_claves:
            ph = ','.join(['%s'] * len(megamo_claves))
            cur.execute(f"SELECT clave, nivel FROM clientes WHERE clave IN ({ph})",
                        list(megamo_claves))
            nivel_map = {r['clave']: r['nivel'] for r in cur.fetchall()}

        # Costos desde Odoo (cacheados)
        costos_map     = _get_costos_odoo()
        megamo_precios = _get_megamo_precios_por_nivel()

        # Cache de nombres Scott con es_MX (color/talla incluidos)
        _ensure_scott_names()

        # Info de odoo_catalogo para SKUs sin proyecciones registradas
        skus_sin_totales = [s for s in FORECAST_SKU_WHITELIST if s not in totales_map]
        oc_fallback: dict = {}
        if skus_sin_totales:
            ph_st = ','.join(['%s'] * len(skus_sin_totales))
            cur.execute(
                f"SELECT referencia_interna, nombre_producto, marca, color, talla "
                f"FROM odoo_catalogo WHERE referencia_interna IN ({ph_st})",
                skus_sin_totales
            )
            for row in cur.fetchall():
                oc_fallback[row['referencia_interna']] = row

        # Construir los 92 artículos
        articulos = []
        for sku in FORECAST_SKU_WHITELIST:
            cat_info  = SKU_CATALOG.get(sku, {})
            avail_map = cat_info.get('avail', {})
            precios   = cat_info.get('prices', {})
            t         = totales_map.get(sku)

            # Si hay datos de órdenes, recalcular cantidades desde el desglose deducido
            desgloses_sku = desglose_map.get(sku, [])
            if ordenes_dist:
                meses_data = {
                    mes: {
                        'cantidad':   sum(d['meses'].get(mes, 0) for d in desgloses_sku),
                        'disponible': avail_map.get(mes, True),
                    }
                    for mes in MESES
                }
                total_anual = sum(meses_data[mes]['cantidad'] for mes in MESES)
            else:
                meses_data = {
                    mes: {
                        'cantidad':   int(t[mes] or 0) if t else 0,
                        'disponible': avail_map.get(mes, True),
                    }
                    for mes in MESES
                }
                total_anual = int(t['total_anual'] or 0) if t else 0

            if sku.startswith('MH'):
                # Megamo: costo por nivel de cada distribuidor con unidades
                dists_con_unidades = desglose_map.get(sku, [])  # ya filtrado a total > 0
                num_dists = len(dists_con_unidades)
                precios_sku = megamo_precios.get(sku, {})
                suma_precios = 0.0
                for d in dists_con_unidades:
                    nivel      = nivel_map.get(d['clave_cliente'], 'Distribuidor')
                    precio_niv = precios_sku.get(nivel, precios_sku.get('Distribuidor', 0.0))
                    suma_precios += precio_niv
                costo_total    = round(suma_precios, 2)
                costo_unitario = round(suma_precios / num_dists, 2) if num_dists > 0 else 0.0
            else:
                # Scott/otros: standard_price × unidades
                costo_unitario = costos_map.get(sku, 0.0)
                costo_total    = round(total_anual * costo_unitario, 2)

            costos_mes = {mes: round(meses_data[mes]['cantidad'] * costo_unitario, 2) for mes in MESES}

            # Datos de producto: forecast_proyecciones → cache Scott → odoo_catalogo
            oc = oc_fallback.get(sku, {})
            _prod  = (t['producto'] if t else None) or _SCOTT_CORRECT_NAMES.get(sku) or oc.get('nombre_producto', '')
            _marca = (t['marca']    if t else None) or oc.get('marca', '')
            _modelo= (t['modelo']   if t else None) or ''
            _color = (t['color']    if t else None) or _SCOTT_COLORS.get(sku) or oc.get('color', '')
            _talla = (t['talla']    if t else None) or _SCOTT_TALLAS.get(sku) or oc.get('talla', '')
            # Marca: si el nombre contiene SCOTT o MEGAMO, usarla aunque falte en odoo_catalogo
            if not _marca:
                nm_up = (_prod or '').upper()
                if 'SCOTT' in nm_up:
                    _marca = 'SCOTT'
                elif 'MEGAMO' in nm_up:
                    _marca = 'MEGAMO'

            articulos.append({
                'sku':               sku,
                'producto':          _prod,
                'marca':             _marca,
                'modelo':            _modelo,
                'color':             _color,
                'talla':             _talla,
                'precio_dist':       float(precios.get('Distribuidor', 0)),
                'costo_unitario':    round(costo_unitario, 2),
                'costo_total':       costo_total,
                'costos_mes':        costos_mes,
                'num_distribuidores': int(t['num_distribuidores'] or 0) if t else 0,
                'total_anual':       total_anual,
                'meses':             meses_data,
                'desglose':          desglose_map.get(sku, []),
            })

        totales_mes   = {mes: sum(a['meses'][mes]['cantidad'] for a in articulos) for mes in MESES}
        total_general = sum(totales_mes.values())
        total_costo_mes   = {mes: round(sum(a['costos_mes'][mes] for a in articulos), 2) for mes in MESES}
        total_costo_general = round(sum(a['costo_total'] for a in articulos), 2)
        skus_con_costo      = sum(1 for a in articulos if a['costo_unitario'] > 0)
        inversion_promedio  = round(total_costo_general / total_general, 2) if total_general > 0 else 0.0

        distribuidores_activos = {
            d['clave_cliente']
            for a in articulos
            for d in a['desglose']
            if d['total'] > 0
        }

        return {
            'articulos':          articulos,
            'totales_mes':        totales_mes,
            'total_general':      total_general,
            'total_costo_mes':    total_costo_mes,
            'total_costo_general': total_costo_general,
            'kpis': {
                'total_articulos':        len(articulos),
                'articulos_con_pedido':   sum(1 for a in articulos if a['total_anual'] > 0),
                'articulos_sin_pedido':   sum(1 for a in articulos if a['total_anual'] == 0),
                'total_unidades':         total_general,
                'distribuidores_activos': len(distribuidores_activos),
                'skus_con_costo':         skus_con_costo,
                'inversion_total':        total_costo_general,
                'inversion_promedio':     inversion_promedio,
            },
            'meses':        MESES,
            'meses_labels': MESES_LABEL,
            'periodo':      periodo or '2026-2027',
            'generado_en':  ahora_str('%d/%m/%Y %H:%M'),
        }

    finally:
        cur.close()
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# GET /proyecciones-my27  — datos JSON
# ─────────────────────────────────────────────────────────────────────────────

@proyecciones_my27_bp.route('', methods=['GET'])
def listar():
    """
    Retorna los 92 artículos MY27 con cantidades consolidadas por mes.
    Query param: ?periodo=2026-2027  &refresh=1 para forzar recarga de costos
    """
    try:
        periodo = request.args.get('periodo', '2026-2027').strip()
        force   = bool(request.args.get('refresh'))

        if force:
            _COSTOS_CACHE['ts'] = 0.0
            _MEGAMO_PRECIOS_CACHE['ts'] = 0.0
            _ORDENES_CACHE['ts'] = 0.0
        else:
            cached = _redis_get(_rkey_my27(periodo))
            if cached is not None:
                return jsonify(cached), 200

        data = _get_datos_consolidados(periodo)
        _redis_set(_rkey_my27(periodo), data, _MY27_R_TTL)
        return jsonify(data), 200
    except Exception as e:
        logging.exception('[proyecciones_my27] listar error: %s', e)
        return jsonify({'error': str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
# GET /proyecciones-my27/exportar  — descarga Excel
# ─────────────────────────────────────────────────────────────────────────────

@proyecciones_my27_bp.route('/exportar', methods=['GET'])
def exportar_excel():
    """
    Exporta el resumen de proyecciones MY27 a Excel con formato profesional.
    Query param: ?periodo=2026-2027
    """
    if not OPENPYXL_OK:
        return jsonify({'error': 'openpyxl no disponible'}), 500

    try:
        periodo = request.args.get('periodo', '2026-2027').strip()
        marca   = request.args.get('marca', '').strip().upper()  # ej. MEGAMO, SCOTT
        data    = _get_datos_consolidados(periodo)

        if marca:
            arts = [a for a in data['articulos']
                    if a.get('marca', '').upper() == marca and a['total_anual'] > 0]
            # Recalcular totales con el subconjunto filtrado
            data['articulos'] = arts
            data['totales_mes'] = {
                mes: sum(a['meses'][mes]['cantidad'] for a in arts) for mes in MESES}
            data['total_general'] = sum(data['totales_mes'].values())
            data['total_costo_mes'] = {
                mes: round(sum(a['costos_mes'][mes] for a in arts), 2) for mes in MESES}
            data['total_costo_general'] = round(sum(a['costo_total'] for a in arts), 2)
            dists_act = {d['clave_cliente'] for a in arts for d in a['desglose'] if d['total'] > 0}
            data['kpis'].update({
                'total_articulos':        len(arts),
                'articulos_con_pedido':   sum(1 for a in arts if a['total_anual'] > 0),
                'articulos_sin_pedido':   sum(1 for a in arts if a['total_anual'] == 0),
                'total_unidades':         data['total_general'],
                'distribuidores_activos': len(dists_act),
                'skus_con_costo':         sum(1 for a in arts if a['costo_unitario'] > 0),
                'inversion_total':        data['total_costo_general'],
                'inversion_promedio':     round(data['total_costo_general'] / data['total_general'], 2)
                                          if data['total_general'] > 0 else 0.0,
            })

        excel  = _generar_excel(data)
        sufijo = f"_{marca}" if marca else ""
        nombre = f"ProyeccionesMY27{sufijo}_{periodo}_{ahora_str('%Y%m%d')}.xlsx"
        buf    = io.BytesIO(excel)
        buf.seek(0)

        return send_file(
            buf,
            as_attachment=True,
            download_name=nombre,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        logging.exception('[proyecciones_my27] exportar error: %s', e)
        return jsonify({'error': str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
# Helper: generar Excel
# ─────────────────────────────────────────────────────────────────────────────

def _generar_excel(data: dict) -> bytes:
    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen consolidado ──────────────────────────────────────────
    ws = wb.active
    ws.title = 'Resumen MY27'

    # Colores
    AZUL_OSCURO = 'FF1A3C5E'
    AZUL_MED    = 'FF2E6DA4'
    GRIS_CLARO  = 'FFF2F2F2'
    NARANJA     = 'FFFF6B00'
    VERDE       = 'FF1E8449'
    ROJO        = 'FFC0392B'
    BLANCO      = 'FFFFFFFF'

    def cell_style(cell, bold=False, bg=None, fg='FF000000',
                   size=10, center=False, wrap=False, border=False):
        cell.font      = Font(bold=bold, color=fg, size=size, name='Calibri')
        cell.alignment = Alignment(
            horizontal='center' if center else 'left',
            vertical='center',
            wrap_text=wrap,
        )
        if bg:
            cell.fill = PatternFill('solid', fgColor=bg)
        if border:
            thin = Side(style='thin', color='FFD0D0D0')
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    thin_side = Side(style='thin', color='FFD0D0D0')
    thin_bdr  = Border(left=thin_side, right=thin_side,
                       top=thin_side,  bottom=thin_side)

    # Columnas: SKU(1) Prod(2) Marca(3) Modelo(4) Color(5) Talla(6)
    #           May-Abr (7-18)  Total(19)  CostoUnit(20)  CostoTotal(21)
    NUM_COLS = 21

    # ── Título principal ─────────────────────────────────────────────────────
    ws.merge_cells(f'A1:{get_column_letter(NUM_COLS)}1')
    titulo = ws['A1']
    titulo.value = f"PROYECCIONES MY27  |  Periodo: {data['periodo']}  |  Generado: {data['generado_en']}"
    cell_style(titulo, bold=True, bg=AZUL_OSCURO, fg=BLANCO, size=13, center=True)
    ws.row_dimensions[1].height = 30

    # ── KPIs fila 2 ─────────────────────────────────────────────────────────
    kpis = data['kpis']
    kpi_datos = [
        ('Total artículos',       kpis['total_articulos']),
        ('Con proyección',        kpis['articulos_con_pedido']),
        ('Sin proyección',        kpis['articulos_sin_pedido']),
        ('Total unidades',        kpis['total_unidades']),
        ('Distribuidores activos', kpis['distribuidores_activos']),
    ]
    col_kpi = 1
    for label, valor in kpi_datos:
        ws.merge_cells(
            start_row=2, start_column=col_kpi,
            end_row=2,   end_column=col_kpi + 2
        )
        c = ws.cell(row=2, column=col_kpi, value=f"{label}: {valor}")
        cell_style(c, bold=True, bg=AZUL_MED, fg=BLANCO, size=10, center=True)
        col_kpi += 3
    ws.row_dimensions[2].height = 22

    # ── Encabezados tabla (fila 4) ───────────────────────────────────────────
    encabezados = ['SKU', 'Producto', 'Marca', 'Modelo', 'Color', 'Talla'] + \
                  MESES_LABEL + ['TOTAL AÑO', 'COSTO UNIT.', 'COSTO TOTAL']

    for ci, enc in enumerate(encabezados, start=1):
        c = ws.cell(row=4, column=ci, value=enc)
        bg_enc = AZUL_OSCURO if enc not in ('COSTO UNIT.', 'COSTO TOTAL') else 'FF6B4C11'
        cell_style(c, bold=True, bg=bg_enc, fg=BLANCO, size=10, center=True)
        c.border = thin_bdr
    ws.row_dimensions[4].height = 24

    # ── Fila de totales generales (fila 3) ───────────────────────────────────
    ws.cell(row=3, column=1, value='TOTALES').font = Font(bold=True, color=AZUL_OSCURO, size=10)
    for ci, mes in enumerate(MESES, start=7):
        c = ws.cell(row=3, column=ci, value=data['totales_mes'].get(mes, 0))
        cell_style(c, bold=True, bg=NARANJA, fg=BLANCO, size=10, center=True)
        c.border = thin_bdr
    c_tot = ws.cell(row=3, column=19, value=data['total_general'])
    cell_style(c_tot, bold=True, bg=NARANJA, fg=BLANCO, size=10, center=True)
    c_tot.border = thin_bdr
    # Totales de costo en fila 3
    ws.cell(row=3, column=20, value='')  # costo unitario no aplica en totales
    c_ctot = ws.cell(row=3, column=21, value=data.get('total_costo_general', 0))
    cell_style(c_ctot, bold=True, bg='FF6B4C11', fg=BLANCO, size=10, center=True)
    c_ctot.number_format = '"$"#,##0.00'
    c_ctot.border = thin_bdr
    ws.row_dimensions[3].height = 20

    # ── Datos: solo artículos con al menos 1 unidad pedida ──────────────────
    arts_con_pedido = [a for a in data['articulos'] if a['total_anual'] > 0]
    for ri, art in enumerate(arts_con_pedido, start=5):
        fila_par = (ri % 2 == 0)
        bg_fila  = GRIS_CLARO if fila_par else BLANCO

        valores = [
            art['sku'], art['producto'], art['marca'],
            art['modelo'], art['color'], art['talla'],
        ]
        for ci, val in enumerate(valores, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            cell_style(c, bg=bg_fila, size=9, border=True)
            c.border = thin_bdr

        total_fila = 0
        for ci, mes in enumerate(MESES, start=7):
            cant = art['meses'][mes]['cantidad']
            disp = art['meses'][mes]['disponible']
            c    = ws.cell(row=ri, column=ci, value=cant if cant > 0 else '')
            total_fila += cant

            if not disp:
                cell_style(c, bg='FF2C2C2C', fg='FF2C2C2C', center=True, size=9)
            elif cant > 0:
                cell_style(c, bold=True, bg=VERDE, fg=BLANCO, center=True, size=9)
            else:
                cell_style(c, bg=bg_fila, center=True, size=9)
            c.border = thin_bdr

        # Total anual
        c_tot = ws.cell(row=ri, column=19, value=total_fila if total_fila > 0 else '')
        if total_fila > 0:
            cell_style(c_tot, bold=True, bg=AZUL_MED, fg=BLANCO, center=True, size=10)
        else:
            cell_style(c_tot, bg=bg_fila, center=True, size=9)
        c_tot.border = thin_bdr

        # Costo unitario
        costo_u = art.get('costo_unitario', 0)
        c_cu = ws.cell(row=ri, column=20, value=costo_u if costo_u > 0 else '')
        cell_style(c_cu, bg=bg_fila, center=True, size=9)
        if costo_u > 0:
            c_cu.number_format = '"$"#,##0.00'
        c_cu.border = thin_bdr

        # Costo total
        costo_t = art.get('costo_total', 0)
        c_ct = ws.cell(row=ri, column=21, value=costo_t if costo_t > 0 else '')
        if costo_t > 0:
            cell_style(c_ct, bold=True, bg='FFFEF3E2', center=True, size=9)
            c_ct.number_format = '"$"#,##0.00'
        else:
            cell_style(c_ct, bg=bg_fila, center=True, size=9)
        c_ct.border = thin_bdr

        ws.row_dimensions[ri].height = 16

    # Anchos de columna
    anchos = [18, 35, 12, 20, 15, 8] + [6] * 12 + [10, 14, 15]
    for ci, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = ancho

    ws.freeze_panes = 'G5'

    # ── Hoja 2: Desglose por distribuidor ────────────────────────────────────
    ws2 = wb.create_sheet('Desglose Distribuidores')

    # Col 7: Clave  |  Col 8: Nombre  |  Col 9-20: meses  |  Col 21: Total
    enc2 = ['SKU', 'Producto', 'Marca', 'Modelo', 'Color', 'Talla',
            'Clave Dist.', 'Nombre Distribuidor'] + MESES_LABEL + ['TOTAL']
    for ci, enc in enumerate(enc2, start=1):
        c = ws2.cell(row=1, column=ci, value=enc)
        cell_style(c, bold=True, bg=AZUL_OSCURO, fg=BLANCO, size=10, center=True)
        c.border = thin_bdr
    ws2.row_dimensions[1].height = 22

    ri2 = 2
    for art in data['articulos']:
        for dist in art['desglose']:
            if dist['total'] == 0:
                continue
            fila_par = (ri2 % 2 == 0)
            bg       = GRIS_CLARO if fila_par else BLANCO

            vals = [art['sku'], art['producto'], art['marca'],
                    art['modelo'], art['color'], art['talla'],
                    dist['clave_cliente'], dist.get('nombre_cliente', dist['clave_cliente'])]
            for ci, v in enumerate(vals, start=1):
                c = ws2.cell(row=ri2, column=ci, value=v)
                cell_style(c, bg=bg, size=9)
                c.border = thin_bdr

            for ci, mes in enumerate(MESES, start=9):
                cant = dist['meses'].get(mes, 0)
                c    = ws2.cell(row=ri2, column=ci, value=cant if cant > 0 else '')
                if cant > 0:
                    cell_style(c, bold=True, bg=VERDE, fg=BLANCO, center=True, size=9)
                else:
                    cell_style(c, bg=bg, center=True, size=9)
                c.border = thin_bdr

            c_tot = ws2.cell(row=ri2, column=21, value=dist['total'] or '')
            cell_style(c_tot, bold=True, bg=AZUL_MED, fg=BLANCO, center=True, size=10)
            c_tot.border = thin_bdr
            ws2.row_dimensions[ri2].height = 16
            ri2 += 1

    anchos2 = [18, 35, 12, 20, 15, 8, 14, 28] + [6] * 12 + [10]
    for ci, ancho in enumerate(anchos2, start=1):
        ws2.column_dimensions[get_column_letter(ci)].width = ancho
    ws2.freeze_panes = 'I2'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# POST /proyecciones-my27/inventario-megamo — subir Excel de stock entrante
# ─────────────────────────────────────────────────────────────────────────────

@proyecciones_my27_bp.route('/inventario-megamo', methods=['POST'])
def subir_inventario_megamo():
    """
    POST /proyecciones-my27/inventario-megamo
    Form fields:
      - file: archivo Excel (.xlsx/.xls)
      - periodo: string, ej. '2026-2027' (default '2026-2027')
    """
    import openpyxl

    periodo = request.form.get('periodo', '2026-2027').strip()
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'Se requiere el campo "file"'}), 400

    ext = (f.filename or '').lower()
    if not (ext.endswith('.xlsx') or ext.endswith('.xls')):
        return jsonify({'error': 'Solo se aceptan archivos .xlsx o .xls'}), 400

    content = f.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        return jsonify({'error': f'No se pudo leer el archivo: {e}'}), 400

    ws = wb.active

    # Detectar fila de headers
    header_row = None
    col_sku = 1
    col_cant = 2
    col_desc = None
    for ri in range(1, 6):
        row_vals = [ws.cell(ri, ci).value for ci in range(1, ws.max_column + 1)]
        row_upper = [str(v).strip().upper() if v else '' for v in row_vals]
        if 'SKU' in row_upper:
            header_row = ri
            col_sku  = row_upper.index('SKU') + 1
            # Buscar CANTIDAD / CANT / QTY / UNIDADES
            for kw in ('CANTIDAD', 'CANT', 'QTY', 'UNIDADES', 'PIEZAS'):
                if kw in row_upper:
                    col_cant = row_upper.index(kw) + 1
                    break
            for kw in ('DESCRIPCION', 'DESCRIPCIÓN', 'NOMBRE', 'PRODUCTO'):
                if kw in row_upper:
                    col_desc = row_upper.index(kw) + 1
                    break
            break

    data_start = (header_row + 1) if header_row else 1

    registros = []
    errores   = []
    for ri in range(data_start, ws.max_row + 1):
        sku_val  = ws.cell(ri, col_sku).value
        cant_val = ws.cell(ri, col_cant).value
        desc_val = ws.cell(ri, col_desc).value if col_desc else None

        if not sku_val:
            continue
        sku = str(sku_val).strip().upper()
        if not sku:
            continue
        try:
            cantidad = int(float(str(cant_val).replace(',', '')))
        except (TypeError, ValueError):
            errores.append(f'Fila {ri}: cantidad inválida para SKU {sku} ({cant_val})')
            continue
        if cantidad <= 0:
            continue

        registros.append({
            'sku': sku,
            'cantidad': cantidad,
            'descripcion': str(desc_val).strip()[:500] if desc_val else None,
        })

    if not registros:
        return jsonify({'error': 'No se encontraron filas válidas en el archivo', 'errores': errores}), 400

    conn = obtener_conexion()
    cur  = conn.cursor()
    insertados = 0
    actualizados = 0
    for rec in registros:
        cur.execute("""
            INSERT INTO forecast_inventario_megamo (periodo, sku, cantidad, descripcion)
            VALUES (%s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                cantidad    = VALUES(cantidad),
                descripcion = VALUES(descripcion),
                subido_en   = NOW()
        """, (periodo, rec['sku'], rec['cantidad'], rec['descripcion']))
        if cur.rowcount == 1:
            insertados += 1
        else:
            actualizados += 1

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({
        'ok': True,
        'periodo': periodo,
        'insertados': insertados,
        'actualizados': actualizados,
        'total': insertados + actualizados,
        'errores': errores,
    }), 200


# ─────────────────────────────────────────────────────────────────────────────
# GET /proyecciones-my27/inventario-megamo — devuelve inventario actual
# ─────────────────────────────────────────────────────────────────────────────

@proyecciones_my27_bp.route('/inventario-megamo', methods=['GET'])
def get_inventario_megamo():
    """GET /proyecciones-my27/inventario-megamo?periodo=2026-2027"""
    periodo = request.args.get('periodo', '2026-2027').strip()
    conn = obtener_conexion()
    cur  = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT sku, cantidad, descripcion, subido_en
        FROM forecast_inventario_megamo
        WHERE periodo = %s
        ORDER BY sku
    """, (periodo,))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    for r in rows:
        if r.get('subido_en'):
            r['subido_en'] = r['subido_en'].isoformat()
    return jsonify({'periodo': periodo, 'inventario': rows}), 200


# ─────────────────────────────────────────────────────────────────────────────
# GET /proyecciones-my27/cobertura-megamo — análisis FIFO de cobertura
# ─────────────────────────────────────────────────────────────────────────────

def _compute_cobertura(periodo: str) -> list:
    """
    Calcula el análisis FIFO de cobertura para un periodo dado.
    Devuelve la lista de dicts de cobertura (o lista vacía si no hay inventario).
    """
    conn = obtener_conexion()
    cur  = conn.cursor(dictionary=True)

    # 1. Inventario entrante
    cur.execute("""
        SELECT sku, cantidad, descripcion
        FROM forecast_inventario_megamo
        WHERE periodo = %s
    """, (periodo,))
    inventario_rows = cur.fetchall()
    if not inventario_rows:
        cur.close()
        conn.close()
        return []

    inventario = {r['sku']: r for r in inventario_rows}
    skus_inv   = list(inventario.keys())

    # 2. Proyecciones globales por SKU
    proy_by_sku = {}
    monitor_cached = _redis_get(_rkey_my27(periodo))

    if monitor_cached:
        for art in monitor_cached.get('articulos', []):
            sku = art['sku']
            if sku not in inventario:
                continue
            proy_by_sku[sku] = {mes: art['meses'][mes]['cantidad'] for mes in MESES_ORDEN}
            proy_by_sku[sku]['sku']      = sku
            proy_by_sku[sku]['producto'] = art.get('producto', sku)
        cur.close()
        conn.close()
    else:
        cols_mes = ', '.join(f'COALESCE(SUM({m}), 0) AS {m}' for m in MESES_ORDEN)
        cur.execute(f"""
            SELECT fp.sku, {cols_mes}, MAX(fp.producto) AS producto
            FROM forecast_proyecciones fp
            WHERE fp.periodo = %s AND fp.sku IN ({','.join(['%s']*len(skus_inv))})
            GROUP BY fp.sku
        """, (periodo, *skus_inv))
        proyecciones_rows = cur.fetchall()

        if proyecciones_rows:
            skus_proy = [r['sku'] for r in proyecciones_rows]
            cur.execute(f"""
                SELECT referencia_interna, nombre_producto
                FROM odoo_catalogo
                WHERE referencia_interna IN ({','.join(['%s']*len(skus_proy))})
            """, skus_proy)
            nombres_odoo = {r['referencia_interna']: r['nombre_producto'] for r in cur.fetchall()}
            for r in proyecciones_rows:
                r['producto'] = nombres_odoo.get(r['sku']) or r.get('producto') or r['sku']
        cur.close()
        conn.close()
        proy_by_sku = {r['sku']: r for r in proyecciones_rows}

    # 3. Calcular cobertura FIFO por SKU
    resultado = []
    for sku in skus_inv:
        inv_rec  = inventario[sku]
        proy_rec = proy_by_sku.get(sku)
        cantidad_entrante = inv_rec['cantidad']

        if not proy_rec:
            resultado.append({
                'sku': sku,
                'producto': inv_rec.get('descripcion') or sku,
                'cantidad_entrante': cantidad_entrante,
                'total_proyectado': 0,
                'total_cubierto': 0,
                'total_deficit': 0,
                'sobrante': cantidad_entrante,
                'cobertura': [{
                    'mes': m, 'proyectado': 0, 'cubierto': 0,
                    'deficit': 0, 'estado': 'sin_demanda'
                } for m in MESES_ORDEN],
            })
            continue

        disponible = cantidad_entrante
        cobertura_meses = []
        total_proyectado = 0
        total_cubierto   = 0

        for mes in MESES_ORDEN:
            proy_mes = int(proy_rec.get(mes) or 0)
            total_proyectado += proy_mes

            if proy_mes == 0:
                cobertura_meses.append({
                    'mes': mes, 'proyectado': 0, 'cubierto': 0,
                    'deficit': 0, 'estado': 'sin_demanda'
                })
                continue

            if disponible >= proy_mes:
                cubierto = proy_mes
                disponible -= proy_mes
                estado = 'completo'
            elif disponible > 0:
                cubierto   = disponible
                disponible = 0
                estado     = 'parcial'
            else:
                cubierto = 0
                estado   = 'sin_cobertura'

            deficit = proy_mes - cubierto
            total_cubierto += cubierto
            cobertura_meses.append({
                'mes': mes, 'proyectado': proy_mes, 'cubierto': cubierto,
                'deficit': deficit, 'estado': estado
            })

        resultado.append({
            'sku': sku,
            'producto': proy_rec.get('producto') or inv_rec.get('descripcion') or sku,
            'cantidad_entrante': cantidad_entrante,
            'total_proyectado': total_proyectado,
            'total_cubierto': total_cubierto,
            'total_deficit': total_proyectado - total_cubierto,
            'sobrante': max(0, disponible),
            'cobertura': cobertura_meses,
        })

    resultado.sort(key=lambda x: (x['total_proyectado'] == 0, x['sku']))
    return resultado


@proyecciones_my27_bp.route('/cobertura-megamo', methods=['GET'])
def get_cobertura_megamo():
    """
    GET /proyecciones-my27/cobertura-megamo?periodo=2026-2027
    Devuelve análisis FIFO de cobertura: stock entrante vs proyecciones globales por SKU.
    """
    periodo  = request.args.get('periodo', '2026-2027').strip()
    resultado = _compute_cobertura(periodo)
    if not resultado:
        return jsonify({'periodo': periodo, 'cobertura': [],
                        'mensaje': 'No hay inventario cargado para este periodo'}), 200
    return jsonify({'periodo': periodo, 'cobertura': resultado}), 200


# ─────────────────────────────────────────────────────────────────────────────
# GET /proyecciones-my27/exportar-cobertura  — descarga Excel de inventario
# ─────────────────────────────────────────────────────────────────────────────

@proyecciones_my27_bp.route('/exportar-cobertura', methods=['GET'])
def exportar_cobertura():
    """
    Exporta el análisis de cobertura FIFO (Inventario Entrante) a Excel.
    Query param: ?periodo=2026-2027
    """
    if not OPENPYXL_OK:
        return jsonify({'error': 'openpyxl no disponible'}), 500

    try:
        periodo = request.args.get('periodo', '2026-2027').strip()

        cobertura = _compute_cobertura(periodo)
        if not cobertura:
            return jsonify({'error': 'No hay datos de inventario para exportar'}), 404

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Inventario Entrante'

        # Estilos
        C_HEADER   = PatternFill('solid', fgColor='1A1A1A')
        C_GREEN    = PatternFill('solid', fgColor='1B4A2A')
        C_RED      = PatternFill('solid', fgColor='4A1A1A')
        C_AMBER    = PatternFill('solid', fgColor='3A2E0A')
        C_NEUTRAL  = PatternFill('solid', fgColor='1E1E1E')
        C_MES_HDR  = PatternFill('solid', fgColor='252525')
        F_WHITE    = Font(color='FFFCF2', bold=True, size=10)
        F_NORMAL   = Font(color='CCCCCC', size=9)
        F_GREEN    = Font(color='4CAF50', bold=True, size=9)
        F_RED      = Font(color='E53935', bold=True, size=9)
        F_AMBER    = Font(color='EB5E28', bold=True, size=9)
        F_MUTED    = Font(color='888888', size=9)
        AL_C       = Alignment(horizontal='center', vertical='center')
        AL_L       = Alignment(horizontal='left',   vertical='center', wrap_text=False)

        thin = Side(style='thin', color='2A2A2A')
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Cabecera ──
        fixed_cols = ['SKU', 'Producto', 'Inventario']
        mes_labels = MESES_LABEL  # 12 meses
        tail_cols  = ['Total Proyectado', 'Total Cubierto', 'Déficit', 'Sobrante', 'Estado']
        headers    = fixed_cols + mes_labels + tail_cols

        ws.append(headers)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill   = C_HEADER
            cell.font   = F_WHITE
            cell.border = brd
            cell.alignment = AL_C

        # ── Filas ──
        for row_i, sku_data in enumerate(cobertura, 2):
            deficit_total = sku_data.get('total_deficit', 0)
            sobrante      = sku_data.get('sobrante', 0)
            estado_str    = ('Sin demanda'    if sku_data['total_proyectado'] == 0
                             else 'Cubierto'  if deficit_total == 0
                             else 'Faltante')

            row_vals = [
                sku_data['sku'],
                sku_data.get('producto', ''),
                sku_data['cantidad_entrante'],
            ]
            for mes in MESES_ORDEN:
                mc = next((c for c in sku_data['cobertura'] if c['mes'] == mes), None)
                row_vals.append(mc['proyectado'] if mc else 0)
            row_vals += [
                sku_data['total_proyectado'],
                sku_data['total_cubierto'],
                deficit_total,
                sobrante,
                estado_str,
            ]

            ws.append(row_vals)

            # Color de fila base
            if sku_data['total_proyectado'] == 0:
                row_fill = C_NEUTRAL
            elif deficit_total == 0:
                row_fill = C_GREEN
            elif deficit_total > 0:
                row_fill = C_RED
            else:
                row_fill = C_NEUTRAL

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_i, column=col_idx)
                cell.fill   = row_fill
                cell.border = brd
                cell.alignment = AL_C

                # Fuente según posición
                if col_idx == 1:  # SKU
                    cell.font = Font(color='EB5E28', bold=True, size=9,
                                     name='Courier New')
                elif col_idx == 2:  # Producto
                    cell.font = F_MUTED
                    cell.alignment = AL_L
                elif col_idx == 3:  # Inventario
                    cell.font = F_WHITE
                elif col_idx in range(4, 4 + 12):  # Meses
                    v = cell.value or 0
                    cell.font = (F_NORMAL if v == 0 else
                                 Font(color='FFFCF2', size=9))
                elif col_idx == len(headers) - 2:  # Déficit
                    cell.font = F_RED if (cell.value or 0) > 0 else F_MUTED
                elif col_idx == len(headers) - 1:  # Sobrante
                    cell.font = F_AMBER if (cell.value or 0) > 0 else F_MUTED
                elif col_idx == len(headers):       # Estado
                    cell.font = (F_GREEN if estado_str == 'Cubierto'
                                 else F_RED if estado_str == 'Faltante'
                                 else F_MUTED)
                else:
                    cell.font = F_NORMAL

        # Anchos de columna
        ws.column_dimensions[get_column_letter(1)].width = 16   # SKU
        ws.column_dimensions[get_column_letter(2)].width = 30   # Producto
        ws.column_dimensions[get_column_letter(3)].width = 13   # Inventario
        for c in range(4, 4 + 12):
            ws.column_dimensions[get_column_letter(c)].width = 8
        for c in range(4 + 12, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 15

        ws.row_dimensions[1].height = 22
        ws.freeze_panes = 'D2'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        nombre = f"InventarioEntrante_{periodo}_{ahora_str('%Y%m%d')}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=nombre,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    except Exception as e:
        logging.exception('[proyecciones_my27] exportar_cobertura error: %s', e)
        return jsonify({'error': str(e)}), 500
