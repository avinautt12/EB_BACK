from __future__ import annotations
"""
Forecast / Proyecciones B2B
Gestión del forecast anual de compra por distribuidor.
Periodo comercial: Mayo–Abril (e.g., "2026-2027")
"""
from flask import Blueprint, jsonify, request, send_file
from db_conexion import obtener_conexion
from services.forecast_excel_service import (
    load_excel_products,
    load_csv_apparel_products,
    list_excel_products,
    delete_excel_product,
    clear_excel_catalog,
    get_valid_skus
)
import io
import re
import threading
import logging
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

forecast_bp = Blueprint('forecast', __name__, url_prefix='')

# ── Redis (L2) + memoria (L1) para evitar llamadas repetidas a Odoo ──────────
import os as _os
import time as _time
import json as _json

_REDIS_URL      = _os.getenv('CELERY_BROKER_URL', 'redis://localhost:6379/0')
_FORECAST_R_TTL = 1500   # 25 min en Redis
_AVANCE_R_TTL   = 1500
_FORECAST_TTL   = 1500   # 25 min en memoria L1 (pre-warming renueva c/20 min)
_AVANCE_TTL     = 1500

try:
    import redis as _redis_lib
    _redis = _redis_lib.from_url(_REDIS_URL, decode_responses=True, socket_connect_timeout=2)
    _redis.ping()
    logging.info('[forecast] Redis activo: %s', _REDIS_URL)
except Exception as _re:
    _redis = None
    logging.warning('[forecast] Redis no disponible: %s', _re)

_avance_cache: dict   = {}   # key: (clave, periodo) → (timestamp, result_list)
_forecast_cache: dict = {}   # key: (clave, periodo) → (timestamp, result_list)


def _rkey_forecast(clave: str, periodo: str) -> str:
    return f'forecast:{clave}:{periodo}'

def _rkey_avance(clave: str, periodo: str) -> str:
    return f'forecast_avance:{clave}:{periodo}'


def _redis_get(key: str):
    """Devuelve el objeto deserializado o None."""
    if not _redis:
        return None
    try:
        raw = _redis.get(key)
        return _json.loads(raw) if raw else None
    except Exception:
        return None


def _redis_set(key: str, data, ttl: int) -> None:
    if not _redis:
        return
    try:
        _redis.setex(key, ttl, _json.dumps(data))
    except Exception:
        pass


def iniciar_precalentamiento_forecast(host: str = 'http://localhost:5000') -> int:
    """
    Lanza threads daemon que llaman /forecast y /forecast/avance para todos
    los clientes+periodos que tienen datos guardados en forecast_proyecciones.
    """
    from concurrent.futures import ThreadPoolExecutor
    import requests as _req

    try:
        conn = obtener_conexion()
        cur = conn.cursor(dictionary=True)
        # Solo periodos activos (últimos 2 años)
        cur.execute("""
            SELECT DISTINCT clave_cliente, periodo
            FROM forecast_proyecciones
            WHERE clave_cliente IS NOT NULL AND clave_cliente != ''
              AND periodo IS NOT NULL
            ORDER BY periodo DESC
        """)
        pairs = [(r['clave_cliente'], r['periodo']) for r in cur.fetchall()]
        cur.close()
        conn.close()
    except Exception as _e:
        logging.warning('[forecast] precalentamiento: no se pudo leer pares: %s', _e)
        return 0

    def _warm(clave: str, periodo: str) -> None:
        try:
            _req.get(f'{host}/forecast',       params={'clave': clave, 'periodo': periodo}, timeout=120)
            _req.get(f'{host}/forecast/avance', params={'clave': clave, 'periodo': periodo}, timeout=120)
        except Exception as _e:
            logging.debug('[forecast] warm error %s/%s: %s', clave, periodo, _e)

    def _run():
        logging.info('[forecast] Precalentamiento iniciado para %d pares clave+periodo', len(pairs))
        with ThreadPoolExecutor(max_workers=4) as pool:
            for clave, periodo in pairs:
                pool.submit(_warm, clave, periodo)
        logging.info('[forecast] Precalentamiento forecast terminado')

    threading.Thread(target=_run, daemon=True).start()
    return len(pairs)

# Orden de meses en el periodo comercial Mayo–Abril
MESES = ['mayo', 'junio', 'julio', 'agosto', 'septiembre', 'octubre',
         'noviembre', 'diciembre', 'enero', 'febrero', 'marzo', 'abril']
MESES_LABELS = ['May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic', 'Ene', 'Feb', 'Mar', 'Abr']

TIER_NAMES  = ['Partner Elite Plus!', 'Partner Elite', 'Partner', 'Distribuidor']
CAMPOS_INFO = ['SKU', 'Producto', 'Marca', 'Modelo', 'Color', 'Talla']
IVA_FACTOR  = 1.16   # precios en SKU_CATALOG y Odoo son sin IVA

# Mapeo tier (MySQL) → nombre exacto de lista de precios en Odoo
_TIER_TO_ODOO_PL: dict[str, str] = {
    'Partner Elite Plus!': 'PARTNER ELITE PLUS!',
    'Partner Elite':       'PARTNER ELITE',
    'Partner':             'PARTNER',
    'Distribuidor':        'DISTRIBUIDOR',
}

# SKU whitelist para proyecciones (solo estos productos se muestran)
FORECAST_SKU_WHITELIST = [
    # ── Scott MY27 ──────────────────────────────────────────────────────────
    '427981-5814004', '427981-5814006', '427981-5814008', '427981-5814010', '427981-5814012',
    '427982-8561004', '427982-8561006', '427982-8561008', '427982-8561010', '427982-8561012',
    '427982-0002004', '427982-0002006', '427982-0002008', '427982-0002010', '427982-0002012',
    '427902-8551004', '427902-8551006', '427902-8551008', '427902-8551010', '427902-8551012',
    '427902-8512004', '427902-8512006', '427902-8512008', '427902-8512010', '427902-8512012',
    '427983-8587004', '427983-8587006', '427983-8587008', '427983-8587010', '427983-8587012',
    '427983-8266004', '427983-8266006', '427983-8266008', '427983-8266010', '427983-8266012',
    '427984-8523004', '427984-8523006', '427984-8523008', '427984-8523010', '427984-8523012', '427984-8523014',
    '427984-8423004', '427984-8423006', '427984-8423008', '427984-8423010', '427984-8423012',
    '427984-8566004', '427984-8566006', '427984-8566008', '427984-8566010', '427984-8566012', '427984-8566014',
    '427794-8538004', '427794-8538006', '427794-8538008', '427794-8538010', '427794-8538012', '427794-8538014',
    '427794-8512004', '427794-8512006', '427794-8512008', '427794-8512010', '427794-8512012', '427794-8512014',
    '427794-8605004', '427794-8605006', '427794-8605008', '427794-8605010',
    '427102-0001004', '427102-0001006', '427102-0001008', '427102-0001010', '427102-0001012', '427102-0001014',
    '427102-0002004', '427102-0002006', '427102-0002008', '427102-0002010', '427102-0002012', '427102-0002014',
    '427102-2878004', '427102-2878006', '427102-2878008', '427102-2878010',
    '427102-1087004', '427102-1087006', '427102-1087008', '427102-1087010',
    '427985-3831006', '427985-3831008', '427985-3831010', '427985-3831012',
    # Megamo MY27 (completo desde CSV)
    # BICICLETA MEGAMO MY27 ALONG 01
    'MH46001205C', 'MH46001305C', 'MH46001405C', 'MH46001505C',
    'MH46001204M', 'MH46001304M', 'MH46001404M', 'MH46001504M',
    # MH09001 BICICLETA MEGAMO MY27 TRACK 01 SLR
    'MH09001205N', 'MH09001305N', 'MH09001405N', 'MH09001505N',
    # MH09006 BICICLETA MEGAMO MY27 TRACK 06
    'MH09006203A', 'MH09006303A', 'MH09006403A', 'MH09006503A',
    'MH09006206M', 'MH09006306M', 'MH09006406M', 'MH09006506M',
    # MH09008 BICICLETA MEGAMO MY27 TRACK 08
    'MH09008203A', 'MH09008303A', 'MH09008403A', 'MH09008503A',
    'MH09008206M', 'MH09008306M', 'MH09008406M', 'MH09008506M',
    # MH09126 BICICLETA MEGAMO MY27 TRACK 00 SLR FA
    'MH09126205N', 'MH09126305N', 'MH09126405N', 'MH09126505N',
    # MH09127 BICICLETA MEGAMO MY27 TRACK 04 FA CW
    'MH09127203A', 'MH09127303A', 'MH09127403A', 'MH09127503A',
    'MH09127206M', 'MH09127306M', 'MH09127406M', 'MH09127506M',
    # MH09128 BICICLETA MEGAMO MY27 TRACK 04 FA
    'MH09128203A', 'MH09128303A', 'MH09128403A', 'MH09128503A',
    'MH09128206M', 'MH09128306M', 'MH09128406M', 'MH09128506M',
    # MH10015 BICICLETA MEGAMO MY27 FACTORY 15
    'MH10015200O', 'MH10015300O', 'MH10015400O', 'MH10015500O',
    # MH10030 BICICLETA MEGAMO MY27 FACTORY 30
    'MH10030200O', 'MH10030300O', 'MH10030400O', 'MH10030500O',
    # MH10037 BICICLETA MEGAMO MY27 FACTORY Deore DI2
    'MH10037200O', 'MH10037300O', 'MH10037400O', 'MH10037500O',
    # MH11030 BICICLETA MEGAMO MY27 NATURAL 30
    'MH11030205A', 'MH11030305A', 'MH11030405A', 'MH11030505A',
    'MH11030207A', 'MH11030307A', 'MH11030407A', 'MH11030507A',
    # MH11040 BICICLETA MEGAMO MY27 NATURAL 40
    'MH11040205A', 'MH11040305A', 'MH11040405A', 'MH11040505A',
    'MH11040207A', 'MH11040307A', 'MH11040407A', 'MH11040507A',
    # MH11043 BICICLETA MEGAMO MY27 NATURAL 60
    'MH11044105A', 'MH11043205A', 'MH11044205A', 'MH11043305A',
    'MH11043405A', 'MH11043505A', 'MH11044107A', 'MH11043207A',
    'MH11044207A', 'MH11043307A', 'MH11043407A', 'MH11043507A',
    # MH12000 BICICLETA MEGAMO MY27 PULSE 00 SLR
    'MH12000109M', 'MH12000209M', 'MH12000309M', 'MH12000409M',
    'MH12000509M', 'MH12000108M', 'MH12000208M', 'MH12000308M',
    'MH12000408M', 'MH12000508M', 'MH12000100N', 'MH12000200N',
    'MH12000300N', 'MH12000400N', 'MH12000500N',
    # MH12001 BICICLETA MEGAMO MY27 PULSE 01 SLR
    'MH12001109M', 'MH12001209M', 'MH12001309M', 'MH12001409M',
    'MH12001509M', 'MH12001108M', 'MH12001208M', 'MH12001308M',
    'MH12001408M', 'MH12001508M', 'MH12001100N', 'MH12001200N',
    'MH12001300N', 'MH12001400N', 'MH12001500N',
    # MH12002 BICICLETA MEGAMO MY27 PULSE 02 SLR
    'MH12002109M', 'MH12002209M', 'MH12002309M', 'MH12002409M',
    'MH12002509M', 'MH12002108M', 'MH12002208M', 'MH12002308M',
    'MH12002408M', 'MH12002508M', 'MH12002100N', 'MH12002200N',
    'MH12002300N', 'MH12002400N', 'MH12002500N',
    # MH12015 BICICLETA MEGAMO MY27 PULSE 15
    'MH12015102N', 'MH12015202N', 'MH12015302N', 'MH12015402N',
    'MH12015502N', 'MH12015103N', 'MH12015203N', 'MH12015303N',
    'MH12015403N', 'MH12015503N', 'MH12015104N', 'MH12015204N',
    'MH12015304N', 'MH12015404N', 'MH12015504N',
    # MH12095 BICICLETA MEGAMO MY27 PULSE 05 CW
    'MH12095102N', 'MH12095202N', 'MH12095302N', 'MH12095402N',
    'MH12095502N', 'MH12095103N', 'MH12095203N', 'MH12095303N',
    'MH12095403N', 'MH12095503N', 'MH12095104N', 'MH12095204N',
    'MH12095304N', 'MH12095404N', 'MH12095504N',
    # MH12097 BICICLETA MEGAMO MY27 PULSE 15 CW
    'MH12097102N', 'MH12097202N', 'MH12097302N', 'MH12097402N',
    'MH12097502N', 'MH12097103N', 'MH12097203N', 'MH12097303N',
    'MH12097403N', 'MH12097503N', 'MH12097104N', 'MH12097204N',
    'MH12097304N', 'MH12097404N', 'MH12097504N',
    # MH12117 BICICLETA MEGAMO MY27 PULSE 04 SLR
    'MH12117109M', 'MH12117209M', 'MH12117309M', 'MH12117409M',
    'MH12117509M', 'MH12117108M', 'MH12117208M', 'MH12117308M',
    'MH12117408M', 'MH12117508M', 'MH12117100N', 'MH12117200N',
    'MH12117300N', 'MH12117400N', 'MH12117500N',
    # MH12118 BICICLETA MEGAMO MY27 PULSE 07 SLR
    'MH12118109M', 'MH12118209M', 'MH12118309M', 'MH12118409M',
    'MH12118509M', 'MH12118108M', 'MH12118208M', 'MH12118308M',
    'MH12118408M', 'MH12118508M', 'MH12118100N', 'MH12118200N',
    'MH12118300N', 'MH12118400N', 'MH12118500N',
    # MH12123 BICICLETA MEGAMO MY27 PULSE 03 CW LTD
    'MH12123102N', 'MH12123202N', 'MH12123302N', 'MH12123402N',
    'MH12123502N', 'MH12123103N', 'MH12123203N', 'MH12123303N',
    'MH12123403N', 'MH12123503N', 'MH12123104N', 'MH12123204N',
    'MH12123304N', 'MH12123404N', 'MH12123504N',
    # MH13000 BICICLETA MEGAMO MY27 RAISE 00 SLR
    'MH13000104J', 'MH13000204J', 'MH13000304J', 'MH13000404J',
    'MH13000504J', 'MH13000106N', 'MH13000206N', 'MH13000306N',
    'MH13000406N', 'MH13000506N',
    # MH13001 BICICLETA MEGAMO MY27 RAISE 01 SLR
    'MH13001104J', 'MH13001204J', 'MH13001304J', 'MH13001404J',
    'MH13001504J', 'MH13001106N', 'MH13001206N', 'MH13001306N',
    'MH13001406N', 'MH13001506N',
    # MH13002 BICICLETA MEGAMO MY27 RAISE 02 SLR
    'MH13002104J', 'MH13002204J', 'MH13002304J', 'MH13002404J',
    'MH13002504J', 'MH13002106N', 'MH13002206N', 'MH13002306N',
    'MH13002406N', 'MH13002506N',
    # MH13015 BICICLETA MEGAMO MY27 RAISE 15
    'MH13015104A', 'MH13015204A', 'MH13015304A', 'MH13015404A',
    'MH13015504A', 'MH13015103A', 'MH13015203A', 'MH13015303A',
    'MH13015403A', 'MH13015503A', 'MH13015107N', 'MH13015207N',
    'MH13015307N', 'MH13015407N', 'MH13015507N',
    # MH13020 BICICLETA MEGAMO MY27 RAISE 20
    'MH13020104A', 'MH13020204A', 'MH13020304A', 'MH13020404A',
    'MH13020504A', 'MH13020103A', 'MH13020203A', 'MH13020303A',
    'MH13020403A', 'MH13020503A', 'MH13020107N', 'MH13020207N',
    'MH13020307N', 'MH13020407N', 'MH13020507N',
    # MH13095 BICICLETA MEGAMO MY27 RAISE 05 CW
    'MH13095104A', 'MH13095204A', 'MH13095304A', 'MH13095404A',
    'MH13095504A', 'MH13095103A', 'MH13095203A', 'MH13095303A',
    'MH13095403A', 'MH13095503A', 'MH13095107N', 'MH13095207N',
    'MH13095307N', 'MH13095407N', 'MH13095507N',
    # MH13097 BICICLETA MEGAMO MY27 RAISE 15 CW
    'MH13097104A', 'MH13097204A', 'MH13097304A', 'MH13097404A',
    'MH13097504A', 'MH13097103A', 'MH13097203A', 'MH13097303A',
    'MH13097403A', 'MH13097503A', 'MH13097107N', 'MH13097207N',
    'MH13097307N', 'MH13097407N', 'MH13097507N',
    # MH13117 BICICLETA MEGAMO MY27 RAISE 04 SLR
    'MH13117104J', 'MH13117204J', 'MH13117304J', 'MH13117404J',
    'MH13117504J', 'MH13117106N', 'MH13117206N', 'MH13117306N',
    'MH13117406N', 'MH13117506N',
    # MH13123 BICICLETA MEGAMO MY27 RAISE 03 CW LTD
    'MH13123104A', 'MH13123204A', 'MH13123304A', 'MH13123404A',
    'MH13123504A', 'MH13123103A', 'MH13123203A', 'MH13123303A',
    'MH13123403A', 'MH13123503A', 'MH13123107N', 'MH13123207N',
    'MH13123307N', 'MH13123407N', 'MH13123507N',
    # MH13129 BICICLETA MEGAMO MY27 RAISE 07 CW
    'MH13129104A', 'MH13129204A', 'MH13129304A', 'MH13129404A',
    'MH13129504A', 'MH13129103A', 'MH13129203A', 'MH13129303A',
    'MH13129403A', 'MH13129503A', 'MH13129107N', 'MH13129207N',
    'MH13129307N', 'MH13129407N', 'MH13129507N',
    # MH15001 BICICLETA MEGAMO MY27 WEST 01
    'MH15001105C', 'MH15001205C', 'MH15001305C', 'MH15001405C',
    'MH15001505C', 'MH15001107L', 'MH15001207L', 'MH15001307L',
    'MH15001407L', 'MH15001507L', 'MH15001100O', 'MH15001200O',
    'MH15001300O', 'MH15001400O', 'MH15001500O',
    # MH15003 BICICLETA MEGAMO MY27 WEST 03
    'MH15003105C', 'MH15003205C', 'MH15003305C', 'MH15003405C',
    'MH15003505C', 'MH15003107L', 'MH15003207L', 'MH15003307L',
    'MH15003407L', 'MH15003507L', 'MH15003100O', 'MH15003200O',
    'MH15003300O', 'MH15003400O', 'MH15003500O',
    # MH15005 BICICLETA MEGAMO MY27 WEST 05
    'MH15005105C', 'MH15005205C', 'MH15005305C', 'MH15005405C',
    'MH15005505C', 'MH15005107L', 'MH15005207L', 'MH15005307L',
    'MH15005407L', 'MH15005507L', 'MH15005100O', 'MH15005200O',
    'MH15005300O', 'MH15005400O', 'MH15005500O',
    # MH15010 BICICLETA MEGAMO MY27 WEST 10
    'MH15010105C', 'MH15010205C', 'MH15010305C', 'MH15010405C',
    'MH15010505C', 'MH15010107L', 'MH15010207L', 'MH15010307L',
    'MH15010407L', 'MH15010507L', 'MH15010100O', 'MH15010200O',
    'MH15010300O', 'MH15010400O', 'MH15010500O',
    # MH15015 BICICLETA MEGAMO MY27 WEST 15
    'MH15015105C', 'MH15015205C', 'MH15015305C', 'MH15015405C',
    'MH15015505C', 'MH15015107L', 'MH15015207L', 'MH15015307L',
    'MH15015407L', 'MH15015507L', 'MH15015100O', 'MH15015200O',
    'MH15015300O', 'MH15015400O', 'MH15015500O',
    # MH16020 BICICLETA MEGAMO MY27 JAKAR 20
    'MH16020107M', 'MH16020207M', 'MH16020307M', 'MH16020407M',
    'MH16020507M', 'MH16020103J', 'MH16020203J', 'MH16020303J',
    'MH16020403J', 'MH16020503J', 'MH16020105M', 'MH16020205M',
    'MH16020305M', 'MH16020405M', 'MH16020505M',
    # MH16030 BICICLETA MEGAMO MY27 JAKAR 30
    'MH16030107M', 'MH16030207M', 'MH16030307M', 'MH16030407M',
    'MH16030507M', 'MH16030103J', 'MH16030203J', 'MH16030303J',
    'MH16030403J', 'MH16030503J', 'MH16030105M', 'MH16030205M',
    'MH16030305M', 'MH16030405M', 'MH16030505M',
    # MH16032 BICICLETA MEGAMO MY27 JAKAR 30 Flat-Bar
    'MH16032107M', 'MH16032207M', 'MH16032307M', 'MH16032407M',
    'MH16032507M', 'MH16032103J', 'MH16032203J', 'MH16032303J',
    'MH16032403J', 'MH16032503J', 'MH16032105M', 'MH16032205M',
    'MH16032305M', 'MH16032405M', 'MH16032505M',
    # MH17089 BICICLETA MEGAMO MY27 DX3
    'MH17089106A', 'MH17089206A', 'MH17090206A', 'MH17090306A',
    'MH17090406A', 'MH17090506A', 'MH17089103B', 'MH17089203B',
    'MH17090203B', 'MH17090303B', 'MH17090403B', 'MH17090503B',
    # MH18099 BICICLETA MEGAMO MY27 KU4
    'MH18099002H', 'MH18099002B', 'MH18099003B',
    # MH19099 BICICLETA MEGAMO MY27 KU2
    'MH19099004A', 'MH19099003A',
    # MH20078 BICICLETA MEGAMO MY27 20" GO Race
    'MH20078003A', 'MH20078000A',
    # MH20099 BICICLETA MEGAMO MY27 20" GO
    'MH20099003A', 'MH20099000A',
    # MH21099 BICICLETA MEGAMO MY27 18" GO
    'MH21099006A', 'MH21099002H',
    # MH22099 BICICLETA MEGAMO MY27 16" GO
    'MH22099004A', 'MH22099007A',
    # MH23099 BICICLETA MEGAMO MY27 14" GO
    'MH23099009A', 'MH23099003A',
    # MH24099 BICICLETA MEGAMO MY27 12" GO
    'MH24099004A', 'MH24099003A',
    # MH25015 BICICLETA MEGAMO MY27 NATURAL ELITE 15
    'MH25015205A', 'MH25015305A', 'MH25015405A', 'MH25015505A',
    'MH25015207A', 'MH25015307A', 'MH25015407A', 'MH25015507A',
    # MH28099 BICICLETA MEGAMO MY27 JAKAR Base
    'MH28099106B', 'MH28099106L',
    # MH37000 BICICLETA MEGAMO MY27 REASON CRB 00
    'MH37000203M', 'MH37000303M', 'MH37000403M', 'MH37000503M',
    # MH37001 BICICLETA MEGAMO MY27 REASON CRB 01
    'MH37001207K', 'MH37001307K', 'MH37001407K', 'MH37001507K',
    'MH37001200M', 'MH37001300M', 'MH37001400M', 'MH37001500M',
    'MH37001207A', 'MH37001307A', 'MH37001407A', 'MH37001507A',
    # MH37002 BICICLETA MEGAMO MY27 REASON CRB 02
    'MH37002204J', 'MH37002304J', 'MH37002404J', 'MH37002504J',
    # MH37003 BICICLETA MEGAMO MY27 REASON CRB 03
    'MH37003207K', 'MH37003307K', 'MH37003407K', 'MH37003507K',
    'MH37003200M', 'MH37003300M', 'MH37003400M', 'MH37003500M',
    'MH37003207A', 'MH37003307A', 'MH37003407A', 'MH37003507A',
    # MH37005 BICICLETA MEGAMO MY27 REASON CRB 05
    'MH37005207K', 'MH37005307K', 'MH37005407K', 'MH37005507K',
    'MH37005200M', 'MH37005300M', 'MH37005400M', 'MH37005500M',
    'MH37005207A', 'MH37005307A', 'MH37005407A', 'MH37005507A',
    # MH37007 BICICLETA MEGAMO MY27 REASON CRB 07
    'MH37007207K', 'MH37007307K', 'MH37007407K', 'MH37007507K',
    'MH37007200M', 'MH37007300M', 'MH37007400M', 'MH37007500M',
    'MH37007207A', 'MH37007307A', 'MH37007407A', 'MH37007507A',
    # MH37013 BICICLETA MEGAMO MY27 REASON CRB 03 AXS
    'MH37013207K', 'MH37013307K', 'MH37013407K', 'MH37013507K',
    'MH37013200M', 'MH37013300M', 'MH37013400M', 'MH37013500M',
    'MH37013207A', 'MH37013307A', 'MH37013407A', 'MH37013507A',
    # MH38003 BICICLETA MEGAMO MY27 REASON AL 03
    'MH38003201M', 'MH38003301M', 'MH38003401M', 'MH38003501M',
    'MH38003202M', 'MH38003302M', 'MH38003402M', 'MH38003502M',
    'MH38003205A', 'MH38003305A', 'MH38003405A', 'MH38003505A',
    # MH38005 BICICLETA MEGAMO MY27 REASON AL 05
    'MH38005201M', 'MH38005301M', 'MH38005401M', 'MH38005501M',
    'MH38005202M', 'MH38005302M', 'MH38005402M', 'MH38005502M',
    'MH38005205A', 'MH38005305A', 'MH38005405A', 'MH38005505A',
    # MH38007 BICICLETA MEGAMO MY27 REASON AL 07
    'MH38007201M', 'MH38007301M', 'MH38007401M', 'MH38007501M',
    'MH38007202M', 'MH38007302M', 'MH38007402M', 'MH38007502M',
    'MH38007205A', 'MH38007305A', 'MH38007405A', 'MH38007505A',
    # MH38013 BICICLETA MEGAMO MY27 REASON AL 03 AXS
    'MH38013201M', 'MH38013301M', 'MH38013401M', 'MH38013501M',
    'MH38013202M', 'MH38013302M', 'MH38013402M', 'MH38013502M',
    'MH38013205A', 'MH38013305A', 'MH38013405A', 'MH38013505A',
    # MH39000 BICICLETA MEGAMO MY27 REASON AIR CRB 00
    'MH39000207K', 'MH39000307K', 'MH39000407K', 'MH39000507K',
    'MH39000200M', 'MH39000300M', 'MH39000400M', 'MH39000500M',
    'MH39000207A', 'MH39000307A', 'MH39000407A', 'MH39000507A',
    # MH39005 BICICLETA MEGAMO MY27 REASON AIR CRB 05
    'MH39005207K', 'MH39005307K', 'MH39005407K', 'MH39005507K',
    'MH39005200M', 'MH39005300M', 'MH39005400M', 'MH39005500M',
    'MH39005207A', 'MH39005307A', 'MH39005407A', 'MH39005507A',
    # MH39007 BICICLETA MEGAMO MY27 REASON AIR CRB 07
    'MH39007207K', 'MH39007307K', 'MH39007407K', 'MH39007507K',
    'MH39007200M', 'MH39007300M', 'MH39007400M', 'MH39007500M',
    'MH39007207A', 'MH39007307A', 'MH39007407A', 'MH39007507A',
    # MH39013 BICICLETA MEGAMO MY27 REASON AIR CRB 03 AXS
    'MH39013207K', 'MH39013307K', 'MH39013407K', 'MH39013507K',
    'MH39013200M', 'MH39013300M', 'MH39013400M', 'MH39013500M',
    'MH39013207A', 'MH39013307A', 'MH39013407A', 'MH39013507A',
    # MH40005 BICICLETA MEGAMO MY27 REASON AIR AL 05
    'MH40005201M', 'MH40005301M', 'MH40005401M', 'MH40005501M',
    'MH40005202M', 'MH40005302M', 'MH40005402M', 'MH40005502M',
    'MH40005205A', 'MH40005305A', 'MH40005405A', 'MH40005505A',
    # MH40007 BICICLETA MEGAMO MY27 REASON AIR AL 07
    'MH40007201M', 'MH40007301M', 'MH40007401M', 'MH40007501M',
    'MH40007202M', 'MH40007302M', 'MH40007402M', 'MH40007502M',
    'MH40007205A', 'MH40007305A', 'MH40007405A', 'MH40007505A',
    # MH41099 BICICLETA MEGAMO MY27 JAKAR Junior
    'MH41099005C', 'MH41099004B',
    # MH44005 BICICLETA MEGAMO MY27 RYAL 05
    'MH44005203A', 'MH44005303A', 'MH44005403A', 'MH44005503A',
    'MH44005205A', 'MH44005305A', 'MH44005405A', 'MH44005505A',
    'MH44005203B', 'MH44005303B', 'MH44005403B', 'MH44005503B',
    # MH44008 BICICLETA MEGAMO MY27 RYAL 08
    'MH44008203A', 'MH44008303A', 'MH44008403A', 'MH44008503A',
    'MH44008205A', 'MH44008305A', 'MH44008405A', 'MH44008505A',
    'MH44008203B', 'MH44008303B', 'MH44008403B', 'MH44008503B',
    # MH44010 BICICLETA MEGAMO MY27 RYAL 10
    'MH44010203A', 'MH44010303A', 'MH44010403A', 'MH44010503A',
    'MH44010205A', 'MH44010305A', 'MH44010405A', 'MH44010505A',
    'MH44010203B', 'MH44010303B', 'MH44010403B', 'MH44010503B',
    # MH44013 BICICLETA MEGAMO MY27 RYAL 03 AXS
    'MH44013203A', 'MH44013303A', 'MH44013403A', 'MH44013503A',
    'MH44013205A', 'MH44013305A', 'MH44013405A', 'MH44013505A',
    'MH44013203B', 'MH44013303B', 'MH44013403B', 'MH44013503B',
    # MH45005 BICICLETA MEGAMO MY27 UPON 05
    'MH45005203A', 'MH45005303A', 'MH45005403A', 'MH45005503A',
    'MH45005205A', 'MH45005305A', 'MH45005405A', 'MH45005505A',
    # MH45020 BICICLETA MEGAMO MY27 UPON 20
    'MH45020203A', 'MH45020303A', 'MH45020403A', 'MH45020503A',
    'MH45020205A', 'MH45020305A', 'MH45020405A', 'MH45020505A',
    # MH45097 BICICLETA MEGAMO MY27 UPON 15 CW
    'MH45097203A', 'MH45097303A', 'MH45097403A', 'MH45097503A',
    'MH45097205A', 'MH45097305A', 'MH45097405A', 'MH45097505A',
    # MH46003 BICICLETA MEGAMO MY27 ALONG 03
    'MH46003205C', 'MH46003305C', 'MH46003405C', 'MH46003505C',
    'MH46003204M', 'MH46003304M', 'MH46003404M', 'MH46003504M',
    # MH46010 BICICLETA MEGAMO MY27 ALONG 10
    'MH46010205C', 'MH46010305C', 'MH46010405C', 'MH46010505C',
    'MH46010204M', 'MH46010304M', 'MH46010404M', 'MH46010504M',
    # MH46036 BICICLETA MEGAMO MY27 ALONG 03 CW
    'MH46036204M', 'MH46036304M', 'MH46036404M', 'MH46036504M',
    # MH46109 BICICLETA MEGAMO MY27 ALONG FLAT-BAR
    'MH46109205C', 'MH46109305C', 'MH46109405C', 'MH46109505C',
    'MH46109204M', 'MH46109304M', 'MH46109404M', 'MH46109504M',
    # MH47000 BICICLETA MEGAMO MY27 SILK 00 SLR
    'MH47000108N', 'MH47000208N', 'MH47000308N', 'MH47000408N',
    'MH47000508N', 'MH47000104L', 'MH47000204L', 'MH47000304L',
    'MH47000404L', 'MH47000504L', 'MH47000105A', 'MH47000205A',
    'MH47000305A', 'MH47000405A', 'MH47000505A',
    # MH47001 BICICLETA MEGAMO MY27 SILK 01 SLR
    'MH47001108N', 'MH47001208N', 'MH47001308N', 'MH47001408N',
    'MH47001508N', 'MH47001104L', 'MH47001204L', 'MH47001304L',
    'MH47001404L', 'MH47001504L', 'MH47001105A', 'MH47001205A',
    'MH47001305A', 'MH47001405A', 'MH47001505A',
    # MH47003 BICICLETA MEGAMO MY27 SILK 03 SLR
    'MH47003108N', 'MH47003208N', 'MH47003308N', 'MH47003408N',
    'MH47003508N', 'MH47003104L', 'MH47003204L', 'MH47003304L',
    'MH47003404L', 'MH47003504L', 'MH47003105A', 'MH47003205A',
    'MH47003305A', 'MH47003405A', 'MH47003505A',
    # MH47005 BICICLETA MEGAMO MY27 SILK 05 SLR
    'MH47005108N', 'MH47005208N', 'MH47005308N', 'MH47005408N',
    'MH47005508N', 'MH47005104L', 'MH47005204L', 'MH47005304L',
    'MH47005404L', 'MH47005504L', 'MH47005105A', 'MH47005205A',
    'MH47005305A', 'MH47005405A', 'MH47005505A',
    # MH47006 BICICLETA MEGAMO MY27 SILK 06
    'MH47006105C', 'MH47006205C', 'MH47006305C', 'MH47006405C',
    'MH47006505C', 'MH47006109N', 'MH47006209N', 'MH47006309N',
    'MH47006409N', 'MH47006509N',
    # MH47007 BICICLETA MEGAMO MY27 SILK 07
    'MH47007105C', 'MH47007205C', 'MH47007305C', 'MH47007405C',
    'MH47007505C', 'MH47007109N', 'MH47007209N', 'MH47007309N',
    'MH47007409N', 'MH47007509N',
    # MH47117 BICICLETA MEGAMO MY27 SILK 04 SLR
    'MH47117108N', 'MH47117208N', 'MH47117308N', 'MH47117408N',
    'MH47117508N', 'MH47117104L', 'MH47117204L', 'MH47117304L',
    'MH47117404L', 'MH47117504L', 'MH47117105A', 'MH47117205A',
    'MH47117305A', 'MH47117405A', 'MH47117505A',
    # ── Scott MY27 nuevos ──────────────────────────────────────────────────
    # SCOTT SPARK RC WC
    '427536-8086006', '427536-8086008', '427536-8086010',
    # SCOTT SPARK RC EXPERT
    '427539-8350006', '427539-8350008', '427539-8350010',
    # SCOTT SPARK RC TEAM ISSUE (427537)
    '427537-8532006', '427537-8532008', '427537-8532010',
    # SCOTT SPARK RC TEAM ISSUE (427540)
    '427540-8527006', '427540-8527008', '427540-8527010', '427540-8527012',
    # SCOTT RANSOM RC
    '427563-8535008', '427563-8535010',
    # SCOTT FOIL RC 10
    '427590-8522002', '427590-8522004', '427590-8522006', '427590-8522008', '427590-8522010',
    '427590-3020002', '427590-3020004', '427590-3020006', '427590-3020008', '427590-3020010',
    # SCOTT FOIL RC 20
    '427591-8086002', '427591-8086004', '427591-8086006', '427591-8086008', '427591-8086010',
    # SCOTT FOIL RC PRO
    '427588-8575004', '427588-8575006', '427588-8575008', '427588-8575010',
    # SCOTT FOIL RC TEAM
    '427589-8350004', '427589-8350006', '427589-8350008', '427589-8350010',
    # SCOTT SPARK RC COMP
    '427541-0002006', '427541-0002008', '427541-0002010', '427541-0002012',
    # SCOTT ADDICT RC PRO
    '427598-8575004', '427598-8575006', '427598-8575008',
    # SCOTT SPARK RC ELITE
    '427995-8585006', '427995-8585008', '427995-8585010',
    # SCOTT ROXTER 200
    '427986-8565222', '427986-8588222',
    # SCOTT ROXTER 400
    '427987-1494222', '427987-8606222',
    # ── Scott adicionales (Contessa / Contrail / Scale / Roxter MY26) ────────
    '286383-704', '286383-706',
    '290310-704', '290310-706', '290310-908',
    '425790-3761222', '425790-8269222',
    '425791-3028222', '425791-8268222',
    '425792-2308222', '425792-4173222',
    '425793-8265222',
    '425794-3774222',
]

# Catálogo oficial MY27 — precios reales + disponibilidad mensual (May-Ago).
# Meses NOT en este dict (Sep-Abr) siempre son disponibles.
# avail: True = puede pedirse, False = mes bloqueado (celda oscura, solo lectura).
SKU_CATALOG: dict = {
    # ── Scott MY27 nuevos ─────────────────────────────────────────────────────
    # SPARK RC WC — llega Julio
    '427536-8086006': {'prices': {'Distribuidor': 140862.07, 'Partner': 137155.17, 'Partner Elite': 132521.55, 'Partner Elite Plus!': 128814.66, 'list_price': 185344.83}, 'avail': {'mayo': False, 'junio': False}},
    '427536-8086008': {'prices': {'Distribuidor': 140862.07, 'Partner': 137155.17, 'Partner Elite': 132521.55, 'Partner Elite Plus!': 128814.66, 'list_price': 185344.83}, 'avail': {'mayo': False, 'junio': False}},
    '427536-8086010': {'prices': {'Distribuidor': 140862.07, 'Partner': 137155.17, 'Partner Elite': 132521.55, 'Partner Elite Plus!': 128814.66, 'list_price': 185344.83}, 'avail': {'mayo': False, 'junio': False}},
    # SPARK RC EXPERT — llega Julio
    '427539-8350006': {'prices': {'Distribuidor': 94127.59, 'Partner': 91803.45, 'Partner Elite': 88898.28, 'Partner Elite Plus!': 86574.14, 'list_price': 116206.9}, 'avail': {'mayo': False, 'junio': False}},
    '427539-8350008': {'prices': {'Distribuidor': 94127.59, 'Partner': 91803.45, 'Partner Elite': 88898.28, 'Partner Elite Plus!': 86574.14, 'list_price': 116206.9}, 'avail': {'mayo': False, 'junio': False}},
    '427539-8350010': {'prices': {'Distribuidor': 94127.59, 'Partner': 91803.45, 'Partner Elite': 88898.28, 'Partner Elite Plus!': 86574.14, 'list_price': 116206.9}, 'avail': {'mayo': False, 'junio': False}},
    # SPARK RC TEAM ISSUE (427537) — llega Septiembre
    '427537-8532006': {'prices': {'Distribuidor': 129530.17, 'Partner': 126331.9, 'Partner Elite': 122334.05, 'Partner Elite Plus!': 119135.78, 'list_price': 159913.79}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False}},
    '427537-8532008': {'prices': {'Distribuidor': 129530.17, 'Partner': 126331.9, 'Partner Elite': 122334.05, 'Partner Elite Plus!': 119135.78, 'list_price': 159913.79}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False}},
    '427537-8532010': {'prices': {'Distribuidor': 129530.17, 'Partner': 126331.9, 'Partner Elite': 122334.05, 'Partner Elite Plus!': 119135.78, 'list_price': 159913.79}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False}},
    # SPARK RC TEAM ISSUE (427540) — Sep todas
    '427540-8527006': {'prices': {'Distribuidor': 76670.69, 'Partner': 74777.59, 'Partner Elite': 72411.21, 'Partner Elite Plus!': 70518.1, 'list_price': 94655.17}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False}},
    '427540-8527008': {'prices': {'Distribuidor': 76670.69, 'Partner': 74777.59, 'Partner Elite': 72411.21, 'Partner Elite Plus!': 70518.1, 'list_price': 94655.17}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False}},
    '427540-8527010': {'prices': {'Distribuidor': 76670.69, 'Partner': 74777.59, 'Partner Elite': 72411.21, 'Partner Elite Plus!': 70518.1, 'list_price': 94655.17}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False}},
    '427540-8527012': {'prices': {'Distribuidor': 76670.69, 'Partner': 74777.59, 'Partner Elite': 72411.21, 'Partner Elite Plus!': 70518.1, 'list_price': 94655.17}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False}},
    # RANSOM RC — Dic/Sep por talla
    '427563-8535008': {'prices': {'Distribuidor': 143216.38, 'Partner': 139680.17, 'Partner Elite': 135259.91, 'Partner Elite Plus!': 131723.71, 'list_price': 176810.34}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False, 'octubre': False, 'noviembre': False}},
    '427563-8535010': {'prices': {'Distribuidor': 143216.38, 'Partner': 139680.17, 'Partner Elite': 135259.91, 'Partner Elite Plus!': 131723.71, 'list_price': 176810.34}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False, 'octubre': False, 'noviembre': False}},
    # FOIL RC 10 (8522) — Sep todas
    '427590-8522002': {'prices': {'Distribuidor': 92102.59, 'Partner': 89828.45, 'Partner Elite': 86985.78, 'Partner Elite Plus!': 84711.64, 'list_price': 113706.9}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False}},
    '427590-8522004': {'prices': {'Distribuidor': 92102.59, 'Partner': 89828.45, 'Partner Elite': 86985.78, 'Partner Elite Plus!': 84711.64, 'list_price': 113706.9}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False}},
    '427590-8522006': {'prices': {'Distribuidor': 92102.59, 'Partner': 89828.45, 'Partner Elite': 86985.78, 'Partner Elite Plus!': 84711.64, 'list_price': 113706.9}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False}},
    '427590-8522008': {'prices': {'Distribuidor': 92102.59, 'Partner': 89828.45, 'Partner Elite': 86985.78, 'Partner Elite Plus!': 84711.64, 'list_price': 113706.9}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False}},
    '427590-8522010': {'prices': {'Distribuidor': 92102.59, 'Partner': 89828.45, 'Partner Elite': 86985.78, 'Partner Elite Plus!': 84711.64, 'list_price': 113706.9}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False}},
    # FOIL RC 10 (3020) — Dic todas
    '427590-3020002': {'prices': {'Distribuidor': 92102.59, 'Partner': 89828.45, 'Partner Elite': 86985.78, 'Partner Elite Plus!': 84711.64, 'list_price': 113706.9}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False, 'octubre': False, 'noviembre': False}},
    '427590-3020004': {'prices': {'Distribuidor': 92102.59, 'Partner': 89828.45, 'Partner Elite': 86985.78, 'Partner Elite Plus!': 84711.64, 'list_price': 113706.9}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False, 'octubre': False, 'noviembre': False}},
    '427590-3020006': {'prices': {'Distribuidor': 92102.59, 'Partner': 89828.45, 'Partner Elite': 86985.78, 'Partner Elite Plus!': 84711.64, 'list_price': 113706.9}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False, 'octubre': False, 'noviembre': False}},
    '427590-3020008': {'prices': {'Distribuidor': 92102.59, 'Partner': 89828.45, 'Partner Elite': 86985.78, 'Partner Elite Plus!': 84711.64, 'list_price': 113706.9}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False, 'octubre': False, 'noviembre': False}},
    '427590-3020010': {'prices': {'Distribuidor': 92102.59, 'Partner': 89828.45, 'Partner Elite': 86985.78, 'Partner Elite Plus!': 84711.64, 'list_price': 113706.9}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False, 'octubre': False, 'noviembre': False}},
    # FOIL RC 20 — Sep todas
    '427591-8086002': {'prices': {'Distribuidor': 78835.34, 'Partner': 76888.79, 'Partner Elite': 74455.6, 'Partner Elite Plus!': 72509.05, 'list_price': 97327.59}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False}},
    '427591-8086004': {'prices': {'Distribuidor': 78835.34, 'Partner': 76888.79, 'Partner Elite': 74455.6, 'Partner Elite Plus!': 72509.05, 'list_price': 97327.59}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False}},
    '427591-8086006': {'prices': {'Distribuidor': 78835.34, 'Partner': 76888.79, 'Partner Elite': 74455.6, 'Partner Elite Plus!': 72509.05, 'list_price': 97327.59}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False}},
    '427591-8086008': {'prices': {'Distribuidor': 78835.34, 'Partner': 76888.79, 'Partner Elite': 74455.6, 'Partner Elite Plus!': 72509.05, 'list_price': 97327.59}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False}},
    '427591-8086010': {'prices': {'Distribuidor': 78835.34, 'Partner': 76888.79, 'Partner Elite': 74455.6, 'Partner Elite Plus!': 72509.05, 'list_price': 97327.59}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False}},
    # FOIL RC PRO — Diciembre
    '427588-8575004': {'prices': {'Distribuidor': 142168.97, 'Partner': 138658.62, 'Partner Elite': 134270.69, 'Partner Elite Plus!': 130760.34, 'list_price': 175517.24}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False, 'octubre': False, 'noviembre': False}},
    '427588-8575006': {'prices': {'Distribuidor': 142168.97, 'Partner': 138658.62, 'Partner Elite': 134270.69, 'Partner Elite Plus!': 130760.34, 'list_price': 175517.24}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False, 'octubre': False, 'noviembre': False}},
    '427588-8575008': {'prices': {'Distribuidor': 142168.97, 'Partner': 138658.62, 'Partner Elite': 134270.69, 'Partner Elite Plus!': 130760.34, 'list_price': 175517.24}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False, 'octubre': False, 'noviembre': False}},
    '427588-8575010': {'prices': {'Distribuidor': 142168.97, 'Partner': 138658.62, 'Partner Elite': 134270.69, 'Partner Elite Plus!': 130760.34, 'list_price': 175517.24}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False, 'octubre': False, 'noviembre': False}},
    # FOIL RC TEAM — Dic todas
    '427589-8350004': {'prices': {'Distribuidor': 119614.66, 'Partner': 116661.21, 'Partner Elite': 112969.4, 'Partner Elite Plus!': 110015.95, 'list_price': 147672.41}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False, 'octubre': False, 'noviembre': False}},
    '427589-8350006': {'prices': {'Distribuidor': 119614.66, 'Partner': 116661.21, 'Partner Elite': 112969.4, 'Partner Elite Plus!': 110015.95, 'list_price': 147672.41}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False, 'octubre': False, 'noviembre': False}},
    '427589-8350008': {'prices': {'Distribuidor': 119614.66, 'Partner': 116661.21, 'Partner Elite': 112969.4, 'Partner Elite Plus!': 110015.95, 'list_price': 147672.41}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False, 'octubre': False, 'noviembre': False}},
    '427589-8350010': {'prices': {'Distribuidor': 119614.66, 'Partner': 116661.21, 'Partner Elite': 112969.4, 'Partner Elite Plus!': 110015.95, 'list_price': 147672.41}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False, 'octubre': False, 'noviembre': False}},
    # SPARK RC COMP — Oct todas
    '427541-0002006': {'prices': {'Distribuidor': 61455.17, 'Partner': 59837.93, 'Partner Elite': 57816.38, 'Partner Elite Plus!': 56199.14, 'list_price': 80862.07}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False}},
    '427541-0002008': {'prices': {'Distribuidor': 61455.17, 'Partner': 59837.93, 'Partner Elite': 57816.38, 'Partner Elite Plus!': 56199.14, 'list_price': 80862.07}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False}},
    '427541-0002010': {'prices': {'Distribuidor': 61455.17, 'Partner': 59837.93, 'Partner Elite': 57816.38, 'Partner Elite Plus!': 56199.14, 'list_price': 80862.07}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False}},
    '427541-0002012': {'prices': {'Distribuidor': 61455.17, 'Partner': 59837.93, 'Partner Elite': 57816.38, 'Partner Elite Plus!': 56199.14, 'list_price': 80862.07}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False}},
    # ADDICT RC PRO — Octubre
    '427598-8575004': {'prices': {'Distribuidor': 143914.66, 'Partner': 140361.21, 'Partner Elite': 135919.4, 'Partner Elite Plus!': 132365.95, 'list_price': 177672.41}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False, 'octubre': False, 'noviembre': False}},
    '427598-8575006': {'prices': {'Distribuidor': 143914.66, 'Partner': 140361.21, 'Partner Elite': 135919.4, 'Partner Elite Plus!': 132365.95, 'list_price': 177672.41}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False, 'octubre': False, 'noviembre': False}},
    '427598-8575008': {'prices': {'Distribuidor': 143914.66, 'Partner': 140361.21, 'Partner Elite': 135919.4, 'Partner Elite Plus!': 132365.95, 'list_price': 177672.41}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False, 'octubre': False, 'noviembre': False}},
    # SPARK RC ELITE — Octubre
    '427995-8585006': {'prices': {'Distribuidor': 76670.69, 'Partner': 74777.59, 'Partner Elite': 72411.21, 'Partner Elite Plus!': 70518.1, 'list_price': 94655.17}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False}},
    '427995-8585008': {'prices': {'Distribuidor': 76670.69, 'Partner': 74777.59, 'Partner Elite': 72411.21, 'Partner Elite Plus!': 70518.1, 'list_price': 94655.17}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False}},
    '427995-8585010': {'prices': {'Distribuidor': 76670.69, 'Partner': 74777.59, 'Partner Elite': 72411.21, 'Partner Elite Plus!': 70518.1, 'list_price': 94655.17}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False}},
    # ROXTER 200 — Octubre
    '427986-8565222': {'prices': {'Distribuidor': 7600.0, 'Partner': 7400.0, 'Partner Elite': 7150.0, 'Partner Elite Plus!': 6950.0, 'list_price': 10000.0}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False}},
    '427986-8588222': {'prices': {'Distribuidor': 7600.0, 'Partner': 7400.0, 'Partner Elite': 7150.0, 'Partner Elite Plus!': 6950.0, 'list_price': 10000.0}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False}},
    # ROXTER 400 — Octubre
    '427987-1494222': {'prices': {'Distribuidor': 7075.86, 'Partner': 6889.66, 'Partner Elite': 6656.9, 'Partner Elite Plus!': 6470.69, 'list_price': 9310.34}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False}},
    '427987-8606222': {'prices': {'Distribuidor': 7075.86, 'Partner': 6889.66, 'Partner Elite': 6656.9, 'Partner Elite Plus!': 6470.69, 'list_price': 9310.34}, 'avail': {'mayo': False, 'junio': False, 'julio': False, 'agosto': False, 'septiembre': False}},
    # ── Scott adicionales (Contessa / Contrail / Scale / Roxter MY26) — disponibles desde julio
    '286383-704':     {'prices': {'Distribuidor': 7241.38, 'Partner': 7241.38, 'Partner Elite': 7241.38, 'Partner Elite Plus!': 7241.38, 'list_price': 9896.55}, 'avail': {'mayo': False, 'junio': False}},
    '286383-706':     {'prices': {'Distribuidor': 7241.38, 'Partner': 7241.38, 'Partner Elite': 7241.38, 'Partner Elite Plus!': 7241.38, 'list_price': 9896.55}, 'avail': {'mayo': False, 'junio': False}},
    '290310-704':     {'prices': {'Distribuidor': 7241.38, 'Partner': 7241.38, 'Partner Elite': 7241.38, 'Partner Elite Plus!': 7241.38, 'list_price': 9896.55}, 'avail': {'mayo': False, 'junio': False}},
    '290310-706':     {'prices': {'Distribuidor': 7241.38, 'Partner': 7241.38, 'Partner Elite': 7241.38, 'Partner Elite Plus!': 7241.38, 'list_price': 9896.55}, 'avail': {'mayo': False, 'junio': False}},
    '290310-908':     {'prices': {'Distribuidor': 7241.38, 'Partner': 7241.38, 'Partner Elite': 7241.38, 'Partner Elite Plus!': 7241.38, 'list_price': 9896.55}, 'avail': {'mayo': False, 'junio': False}},
    '425790-3761222': {'prices': {'Distribuidor': 6077.59, 'Partner': 6077.59, 'Partner Elite': 6077.59, 'Partner Elite Plus!': 6077.59, 'list_price': 8280.17}, 'avail': {'mayo': False, 'junio': False}},
    '425790-8269222': {'prices': {'Distribuidor': 6077.59, 'Partner': 6077.59, 'Partner Elite': 6077.59, 'Partner Elite Plus!': 6077.59, 'list_price': 8280.17}, 'avail': {'mayo': False, 'junio': False}},
    '425791-3028222': {'prices': {'Distribuidor': 5646.55, 'Partner': 5646.55, 'Partner Elite': 5646.55, 'Partner Elite Plus!': 5646.55, 'list_price': 7693.97}, 'avail': {'mayo': False, 'junio': False}},
    '425791-8268222': {'prices': {'Distribuidor': 5646.55, 'Partner': 5646.55, 'Partner Elite': 5646.55, 'Partner Elite Plus!': 5646.55, 'list_price': 7693.97}, 'avail': {'mayo': False, 'junio': False}},
    '425792-2308222': {'prices': {'Distribuidor': 6810.34, 'Partner': 6810.34, 'Partner Elite': 6810.34, 'Partner Elite Plus!': 6810.34, 'list_price': 9306.03}, 'avail': {'mayo': False, 'junio': False}},
    '425792-4173222': {'prices': {'Distribuidor': 6810.34, 'Partner': 6810.34, 'Partner Elite': 6810.34, 'Partner Elite Plus!': 6810.34, 'list_price': 9306.03}, 'avail': {'mayo': False, 'junio': False}},
    '425793-8265222': {'prices': {'Distribuidor': 8258.62, 'Partner': 8043.1,  'Partner Elite': 7767.24, 'Partner Elite Plus!': 7551.72, 'list_price': 10862.07}, 'avail': {'mayo': False, 'junio': False}},
    '425794-3774222': {'prices': {'Distribuidor': 7801.72, 'Partner': 7594.83, 'Partner Elite': 7336.21, 'Partner Elite Plus!': 7551.72, 'list_price': 10258.62}, 'avail': {'mayo': False, 'junio': False}},
    # ── CONTRAIL 40 ── llega MAYO(F): disponible desde junio (junio(F) no bloquea — ya llegó en mayo)
    '427102-0001004': {'prices': {'Distribuidor': 10155.17, 'Partner': 9887.93, 'Partner Elite': 9553.88, 'Partner Elite Plus!': 9286.64, 'list_price': 13362.07}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427102-0001006': {'prices': {'Distribuidor': 10155.17, 'Partner': 9887.93, 'Partner Elite': 9553.88, 'Partner Elite Plus!': 9286.64, 'list_price': 13362.07}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427102-0001008': {'prices': {'Distribuidor': 10155.17, 'Partner': 9887.93, 'Partner Elite': 9553.88, 'Partner Elite Plus!': 9286.64, 'list_price': 13362.07}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427102-0001010': {'prices': {'Distribuidor': 10155.17, 'Partner': 9887.93, 'Partner Elite': 9553.88, 'Partner Elite Plus!': 9286.64, 'list_price': 13362.07}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427102-0001012': {'prices': {'Distribuidor': 10155.17, 'Partner': 9887.93, 'Partner Elite': 9553.88, 'Partner Elite Plus!': 9286.64, 'list_price': 13362.07}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    # talla 14: mayo(F) sí, junio NO → primer mes ordenable: junio
    '427102-0001014': {'prices': {'Distribuidor': 10155.17, 'Partner': 9887.93, 'Partner Elite': 9553.88, 'Partner Elite Plus!': 9286.64, 'list_price': 13362.07}, 'avail': {'mayo': False, 'junio': True,  'julio': True, 'agosto': True}},
    # ── CONTRAIL 40 color 2 ── mayo(F) pero junio sin llegada → primer mes ordenable: junio
    '427102-0002004': {'prices': {'Distribuidor': 10155.17, 'Partner': 9887.93, 'Partner Elite': 9553.88, 'Partner Elite Plus!': 9286.64, 'list_price': 13362.07}, 'avail': {'mayo': False, 'junio': True,  'julio': True, 'agosto': True}},
    '427102-0002006': {'prices': {'Distribuidor': 10155.17, 'Partner': 9887.93, 'Partner Elite': 9553.88, 'Partner Elite Plus!': 9286.64, 'list_price': 13362.07}, 'avail': {'mayo': False, 'junio': True,  'julio': True, 'agosto': True}},
    '427102-0002008': {'prices': {'Distribuidor': 10155.17, 'Partner': 9887.93, 'Partner Elite': 9553.88, 'Partner Elite Plus!': 9286.64, 'list_price': 13362.07}, 'avail': {'mayo': False, 'junio': True,  'julio': True, 'agosto': True}},
    '427102-0002010': {'prices': {'Distribuidor': 10155.17, 'Partner': 9887.93, 'Partner Elite': 9553.88, 'Partner Elite Plus!': 9286.64, 'list_price': 13362.07}, 'avail': {'mayo': False, 'junio': True,  'julio': True, 'agosto': True}},
    '427102-0002012': {'prices': {'Distribuidor': 10155.17, 'Partner': 9887.93, 'Partner Elite': 9553.88, 'Partner Elite Plus!': 9286.64, 'list_price': 13362.07}, 'avail': {'mayo': False, 'junio': True,  'julio': True, 'agosto': True}},
    '427102-0002014': {'prices': {'Distribuidor': 10155.17, 'Partner': 9887.93, 'Partner Elite': 9553.88, 'Partner Elite Plus!': 9286.64, 'list_price': 13362.07}, 'avail': {'mayo': False, 'junio': True,  'julio': True, 'agosto': True}},
    # ── CONTRAIL 40 color 3 ── llega MAYO(F): disponible desde junio
    '427102-1087004': {'prices': {'Distribuidor': 10155.17, 'Partner': 9887.93, 'Partner Elite': 9553.88, 'Partner Elite Plus!': 9286.64, 'list_price': 13362.07}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427102-1087006': {'prices': {'Distribuidor': 10155.17, 'Partner': 9887.93, 'Partner Elite': 9553.88, 'Partner Elite Plus!': 9286.64, 'list_price': 13362.07}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427102-1087008': {'prices': {'Distribuidor': 10155.17, 'Partner': 9887.93, 'Partner Elite': 9553.88, 'Partner Elite Plus!': 9286.64, 'list_price': 13362.07}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427102-1087010': {'prices': {'Distribuidor': 10155.17, 'Partner': 9887.93, 'Partner Elite': 9553.88, 'Partner Elite Plus!': 9286.64, 'list_price': 13362.07}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    # ── CONTRAIL 40 color 4 ── llega MAYO(F): disponible desde junio
    '427102-2878004': {'prices': {'Distribuidor': 10155.17, 'Partner': 9887.93, 'Partner Elite': 9553.88, 'Partner Elite Plus!': 9286.64, 'list_price': 13362.07}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427102-2878006': {'prices': {'Distribuidor': 10155.17, 'Partner': 9887.93, 'Partner Elite': 9553.88, 'Partner Elite Plus!': 9286.64, 'list_price': 13362.07}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427102-2878008': {'prices': {'Distribuidor': 10155.17, 'Partner': 9887.93, 'Partner Elite': 9553.88, 'Partner Elite Plus!': 9286.64, 'list_price': 13362.07}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427102-2878010': {'prices': {'Distribuidor': 10155.17, 'Partner': 9887.93, 'Partner Elite': 9553.88, 'Partner Elite Plus!': 9286.64, 'list_price': 13362.07}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    # ── SCALE 920 ── primera llegada: JULIO (no finales) → primer mes ordenable: julio
    '427983-8587004': {'prices': {'Distribuidor': 14020.69, 'Partner': 13651.72, 'Partner Elite': 13190.52, 'Partner Elite Plus!': 12821.55, 'list_price': 18448.28}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427983-8587006': {'prices': {'Distribuidor': 14020.69, 'Partner': 13651.72, 'Partner Elite': 13190.52, 'Partner Elite Plus!': 12821.55, 'list_price': 18448.28}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427983-8587008': {'prices': {'Distribuidor': 14020.69, 'Partner': 13651.72, 'Partner Elite': 13190.52, 'Partner Elite Plus!': 12821.55, 'list_price': 18448.28}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427983-8587010': {'prices': {'Distribuidor': 14020.69, 'Partner': 13651.72, 'Partner Elite': 13190.52, 'Partner Elite Plus!': 12821.55, 'list_price': 18448.28}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427983-8587012': {'prices': {'Distribuidor': 14020.69, 'Partner': 13651.72, 'Partner Elite': 13190.52, 'Partner Elite Plus!': 12821.55, 'list_price': 18448.28}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427983-8266004': {'prices': {'Distribuidor': 14020.69, 'Partner': 13651.72, 'Partner Elite': 13190.52, 'Partner Elite Plus!': 12821.55, 'list_price': 18448.28}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427983-8266006': {'prices': {'Distribuidor': 14020.69, 'Partner': 13651.72, 'Partner Elite': 13190.52, 'Partner Elite Plus!': 12821.55, 'list_price': 18448.28}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427983-8266008': {'prices': {'Distribuidor': 14020.69, 'Partner': 13651.72, 'Partner Elite': 13190.52, 'Partner Elite Plus!': 12821.55, 'list_price': 18448.28}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427983-8266010': {'prices': {'Distribuidor': 14020.69, 'Partner': 13651.72, 'Partner Elite': 13190.52, 'Partner Elite Plus!': 12821.55, 'list_price': 18448.28}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427983-8266012': {'prices': {'Distribuidor': 14020.69, 'Partner': 13651.72, 'Partner Elite': 13190.52, 'Partner Elite Plus!': 12821.55, 'list_price': 18448.28}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    # ── SCALE 980 ── primera llegada: JUNIO(F) → primer mes ordenable: junio
    '427984-8523004': {'prices': {'Distribuidor': 12317.24, 'Partner': 11993.10, 'Partner Elite': 11587.93, 'Partner Elite Plus!': 11263.79, 'list_price': 16206.90}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427984-8523006': {'prices': {'Distribuidor': 12317.24, 'Partner': 11993.10, 'Partner Elite': 11587.93, 'Partner Elite Plus!': 11263.79, 'list_price': 16206.90}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427984-8523008': {'prices': {'Distribuidor': 12317.24, 'Partner': 11993.10, 'Partner Elite': 11587.93, 'Partner Elite Plus!': 11263.79, 'list_price': 16206.90}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427984-8523010': {'prices': {'Distribuidor': 12317.24, 'Partner': 11993.10, 'Partner Elite': 11587.93, 'Partner Elite Plus!': 11263.79, 'list_price': 16206.90}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427984-8523012': {'prices': {'Distribuidor': 12317.24, 'Partner': 11993.10, 'Partner Elite': 11587.93, 'Partner Elite Plus!': 11263.79, 'list_price': 16206.90}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427984-8523014': {'prices': {'Distribuidor': 12317.24, 'Partner': 11993.10, 'Partner Elite': 11587.93, 'Partner Elite Plus!': 11263.79, 'list_price': 16206.90}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427984-8423004': {'prices': {'Distribuidor': 12317.24, 'Partner': 11993.10, 'Partner Elite': 11587.93, 'Partner Elite Plus!': 11263.79, 'list_price': 16206.90}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427984-8423006': {'prices': {'Distribuidor': 12317.24, 'Partner': 11993.10, 'Partner Elite': 11587.93, 'Partner Elite Plus!': 11263.79, 'list_price': 16206.90}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427984-8423008': {'prices': {'Distribuidor': 12317.24, 'Partner': 11993.10, 'Partner Elite': 11587.93, 'Partner Elite Plus!': 11263.79, 'list_price': 16206.90}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427984-8423010': {'prices': {'Distribuidor': 12317.24, 'Partner': 11993.10, 'Partner Elite': 11587.93, 'Partner Elite Plus!': 11263.79, 'list_price': 16206.90}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427984-8423012': {'prices': {'Distribuidor': 12317.24, 'Partner': 11993.10, 'Partner Elite': 11587.93, 'Partner Elite Plus!': 11263.79, 'list_price': 16206.90}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427984-8566004': {'prices': {'Distribuidor': 12317.24, 'Partner': 11993.10, 'Partner Elite': 11587.93, 'Partner Elite Plus!': 11263.79, 'list_price': 16206.90}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427984-8566006': {'prices': {'Distribuidor': 12317.24, 'Partner': 11993.10, 'Partner Elite': 11587.93, 'Partner Elite Plus!': 11263.79, 'list_price': 16206.90}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427984-8566008': {'prices': {'Distribuidor': 12317.24, 'Partner': 11993.10, 'Partner Elite': 11587.93, 'Partner Elite Plus!': 11263.79, 'list_price': 16206.90}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427984-8566010': {'prices': {'Distribuidor': 12317.24, 'Partner': 11993.10, 'Partner Elite': 11587.93, 'Partner Elite Plus!': 11263.79, 'list_price': 16206.90}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427984-8566012': {'prices': {'Distribuidor': 12317.24, 'Partner': 11993.10, 'Partner Elite': 11587.93, 'Partner Elite Plus!': 11263.79, 'list_price': 16206.90}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427984-8566014': {'prices': {'Distribuidor': 12317.24, 'Partner': 11993.10, 'Partner Elite': 11587.93, 'Partner Elite Plus!': 11263.79, 'list_price': 16206.90}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    # ── SPARK RC ── primera llegada: JUNIO(F) → primer mes ordenable: junio
    '427794-8538004': {'prices': {'Distribuidor': 11072.41, 'Partner': 10781.03, 'Partner Elite': 10416.81, 'Partner Elite Plus!': 10125.43, 'list_price': 14568.97}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427794-8538006': {'prices': {'Distribuidor': 11072.41, 'Partner': 10781.03, 'Partner Elite': 10416.81, 'Partner Elite Plus!': 10125.43, 'list_price': 14568.97}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427794-8538008': {'prices': {'Distribuidor': 11072.41, 'Partner': 10781.03, 'Partner Elite': 10416.81, 'Partner Elite Plus!': 10125.43, 'list_price': 14568.97}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427794-8538010': {'prices': {'Distribuidor': 11072.41, 'Partner': 10781.03, 'Partner Elite': 10416.81, 'Partner Elite Plus!': 10125.43, 'list_price': 14568.97}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427794-8538012': {'prices': {'Distribuidor': 11072.41, 'Partner': 10781.03, 'Partner Elite': 10416.81, 'Partner Elite Plus!': 10125.43, 'list_price': 14568.97}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427794-8538014': {'prices': {'Distribuidor': 11072.41, 'Partner': 10781.03, 'Partner Elite': 10416.81, 'Partner Elite Plus!': 10125.43, 'list_price': 14568.97}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427794-8512004': {'prices': {'Distribuidor': 11072.41, 'Partner': 10781.03, 'Partner Elite': 10416.81, 'Partner Elite Plus!': 10125.43, 'list_price': 14568.97}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427794-8512006': {'prices': {'Distribuidor': 11072.41, 'Partner': 10781.03, 'Partner Elite': 10416.81, 'Partner Elite Plus!': 10125.43, 'list_price': 14568.97}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427794-8512008': {'prices': {'Distribuidor': 11072.41, 'Partner': 10781.03, 'Partner Elite': 10416.81, 'Partner Elite Plus!': 10125.43, 'list_price': 14568.97}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427794-8512010': {'prices': {'Distribuidor': 11072.41, 'Partner': 10781.03, 'Partner Elite': 10416.81, 'Partner Elite Plus!': 10125.43, 'list_price': 14568.97}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427794-8512012': {'prices': {'Distribuidor': 11072.41, 'Partner': 10781.03, 'Partner Elite': 10416.81, 'Partner Elite Plus!': 10125.43, 'list_price': 14568.97}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427794-8512014': {'prices': {'Distribuidor': 11072.41, 'Partner': 10781.03, 'Partner Elite': 10416.81, 'Partner Elite Plus!': 10125.43, 'list_price': 14568.97}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427794-8605004': {'prices': {'Distribuidor': 11072.41, 'Partner': 10781.03, 'Partner Elite': 10416.81, 'Partner Elite Plus!': 10125.43, 'list_price': 14568.97}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427794-8605006': {'prices': {'Distribuidor': 11072.41, 'Partner': 10781.03, 'Partner Elite': 10416.81, 'Partner Elite Plus!': 10125.43, 'list_price': 14568.97}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427794-8605008': {'prices': {'Distribuidor': 11072.41, 'Partner': 10781.03, 'Partner Elite': 10416.81, 'Partner Elite Plus!': 10125.43, 'list_price': 14568.97}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427794-8605010': {'prices': {'Distribuidor': 11072.41, 'Partner': 10781.03, 'Partner Elite': 10416.81, 'Partner Elite Plus!': 10125.43, 'list_price': 14568.97}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    # ── SCALE 910 ── primera llegada: JULIO (no finales) → primer mes ordenable: julio
    '427981-5814004': {'prices': {'Distribuidor': 26796.55, 'Partner': 26091.38, 'Partner Elite': 25209.91, 'Partner Elite Plus!': 24504.74, 'list_price': 35258.62}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427981-5814006': {'prices': {'Distribuidor': 26796.55, 'Partner': 26091.38, 'Partner Elite': 25209.91, 'Partner Elite Plus!': 24504.74, 'list_price': 35258.62}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427981-5814008': {'prices': {'Distribuidor': 26796.55, 'Partner': 26091.38, 'Partner Elite': 25209.91, 'Partner Elite Plus!': 24504.74, 'list_price': 35258.62}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427981-5814010': {'prices': {'Distribuidor': 26796.55, 'Partner': 26091.38, 'Partner Elite': 25209.91, 'Partner Elite Plus!': 24504.74, 'list_price': 35258.62}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427981-5814012': {'prices': {'Distribuidor': 26796.55, 'Partner': 26091.38, 'Partner Elite': 25209.91, 'Partner Elite Plus!': 24504.74, 'list_price': 35258.62}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    # ── SCALE 930 ── primera llegada: JULIO → primer mes ordenable: julio
    '427982-8561004': {'prices': {'Distribuidor': 21489.66, 'Partner': 20924.14, 'Partner Elite': 20217.24, 'Partner Elite Plus!': 19651.72, 'list_price': 28275.86}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427982-8561006': {'prices': {'Distribuidor': 21489.66, 'Partner': 20924.14, 'Partner Elite': 20217.24, 'Partner Elite Plus!': 19651.72, 'list_price': 28275.86}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427982-8561008': {'prices': {'Distribuidor': 21489.66, 'Partner': 20924.14, 'Partner Elite': 20217.24, 'Partner Elite Plus!': 19651.72, 'list_price': 28275.86}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427982-8561010': {'prices': {'Distribuidor': 21489.66, 'Partner': 20924.14, 'Partner Elite': 20217.24, 'Partner Elite Plus!': 19651.72, 'list_price': 28275.86}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427982-8561012': {'prices': {'Distribuidor': 21489.66, 'Partner': 20924.14, 'Partner Elite': 20217.24, 'Partner Elite Plus!': 19651.72, 'list_price': 28275.86}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427982-0002004': {'prices': {'Distribuidor': 21489.66, 'Partner': 20924.14, 'Partner Elite': 20217.24, 'Partner Elite Plus!': 19651.72, 'list_price': 28275.86}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427982-0002006': {'prices': {'Distribuidor': 21489.66, 'Partner': 20924.14, 'Partner Elite': 20217.24, 'Partner Elite Plus!': 19651.72, 'list_price': 28275.86}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427982-0002008': {'prices': {'Distribuidor': 21489.66, 'Partner': 20924.14, 'Partner Elite': 20217.24, 'Partner Elite Plus!': 19651.72, 'list_price': 28275.86}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427982-0002010': {'prices': {'Distribuidor': 21489.66, 'Partner': 20924.14, 'Partner Elite': 20217.24, 'Partner Elite Plus!': 19651.72, 'list_price': 28275.86}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427982-0002012': {'prices': {'Distribuidor': 21489.66, 'Partner': 20924.14, 'Partner Elite': 20217.24, 'Partner Elite Plus!': 19651.72, 'list_price': 28275.86}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    # ── SCALE 900 ── primera llegada: JULIO → primer mes ordenable: julio
    '427902-8551004': {'prices': {'Distribuidor': 17362.07, 'Partner': 16905.17, 'Partner Elite': 16334.05, 'Partner Elite Plus!': 15877.16, 'list_price': 22844.83}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427902-8551006': {'prices': {'Distribuidor': 17362.07, 'Partner': 16905.17, 'Partner Elite': 16334.05, 'Partner Elite Plus!': 15877.16, 'list_price': 22844.83}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427902-8551008': {'prices': {'Distribuidor': 17362.07, 'Partner': 16905.17, 'Partner Elite': 16334.05, 'Partner Elite Plus!': 15877.16, 'list_price': 22844.83}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427902-8551010': {'prices': {'Distribuidor': 17362.07, 'Partner': 16905.17, 'Partner Elite': 16334.05, 'Partner Elite Plus!': 15877.16, 'list_price': 22844.83}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427902-8551012': {'prices': {'Distribuidor': 17362.07, 'Partner': 16905.17, 'Partner Elite': 16334.05, 'Partner Elite Plus!': 15877.16, 'list_price': 22844.83}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427902-8512004': {'prices': {'Distribuidor': 17362.07, 'Partner': 16905.17, 'Partner Elite': 16334.05, 'Partner Elite Plus!': 15877.16, 'list_price': 22844.83}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427902-8512006': {'prices': {'Distribuidor': 17362.07, 'Partner': 16905.17, 'Partner Elite': 16334.05, 'Partner Elite Plus!': 15877.16, 'list_price': 22844.83}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427902-8512008': {'prices': {'Distribuidor': 17362.07, 'Partner': 16905.17, 'Partner Elite': 16334.05, 'Partner Elite Plus!': 15877.16, 'list_price': 22844.83}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427902-8512010': {'prices': {'Distribuidor': 17362.07, 'Partner': 16905.17, 'Partner Elite': 16334.05, 'Partner Elite Plus!': 15877.16, 'list_price': 22844.83}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    '427902-8512012': {'prices': {'Distribuidor': 17362.07, 'Partner': 16905.17, 'Partner Elite': 16334.05, 'Partner Elite Plus!': 15877.16, 'list_price': 22844.83}, 'avail': {'mayo': False, 'junio': False, 'julio': True, 'agosto': True}},
    # ── SCOTT SUB CROSS 40 ── primera llegada: JUNIO(F) → primer mes ordenable: junio
    '427985-3831006': {'prices': {'Distribuidor': 9696.55, 'Partner': 9441.38, 'Partner Elite': 9122.41, 'Partner Elite Plus!': 8867.24, 'list_price': 12758.62}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427985-3831008': {'prices': {'Distribuidor': 9696.55, 'Partner': 9441.38, 'Partner Elite': 9122.41, 'Partner Elite Plus!': 8867.24, 'list_price': 12758.62}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427985-3831010': {'prices': {'Distribuidor': 9696.55, 'Partner': 9441.38, 'Partner Elite': 9122.41, 'Partner Elite Plus!': 8867.24, 'list_price': 12758.62}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
    '427985-3831012': {'prices': {'Distribuidor': 9696.55, 'Partner': 9441.38, 'Partner Elite': 9122.41, 'Partner Elite Plus!': 8867.24, 'list_price': 12758.62}, 'avail': {'mayo': False, 'junio': True, 'julio': True, 'agosto': True}},
}

# Cache de nombres, colores y tallas Scott (cargado una vez desde Odoo con lang=es_MX)
_SCOTT_CORRECT_NAMES: dict = {}
_SCOTT_COLORS: dict = {}
_SCOTT_TALLAS: dict = {}
_SCOTT_NAMES_LOADED: bool = False

def _ensure_scott_names():
    """Carga nombre, color y talla de cada SKU Scott desde Odoo (es_MX), una sola vez."""
    global _SCOTT_CORRECT_NAMES, _SCOTT_COLORS, _SCOTT_TALLAS, _SCOTT_NAMES_LOADED
    if _SCOTT_NAMES_LOADED:
        return
    try:
        from utils.odoo_utils import get_odoo_models, ODOO_DB, ODOO_PASSWORD
        uid_s, models_s, err_s = get_odoo_models()
        if err_s or not uid_s:
            return
        all_scott_skus = list(SKU_CATALOG.keys())
        _ctx = {'lang': 'es_MX'}
        odoo_prods = models_s.execute_kw(ODOO_DB, uid_s, ODOO_PASSWORD,
            'product.product', 'search_read',
            [[['default_code', 'in', all_scott_skus]]],
            {'fields': ['id', 'default_code', 'name',
                        'product_template_attribute_value_ids'],
             'limit': 300, 'context': _ctx}
        )
        # Obtener atributos (color, talla)
        all_avids = list({v for p in odoo_prods
                          for v in p.get('product_template_attribute_value_ids', [])})
        av_map: dict = {}
        ad_map: dict = {}
        if all_avids:
            attr_vals = models_s.execute_kw(ODOO_DB, uid_s, ODOO_PASSWORD,
                'product.template.attribute.value', 'search_read',
                [[['id', 'in', all_avids]]],
                {'fields': ['id', 'name', 'attribute_id'], 'limit': 1000, 'context': _ctx}
            )
            av_map = {a['id']: a for a in attr_vals}
            adids = list({a['attribute_id'][0] for a in attr_vals})
            if adids:
                ad_map = {a['id']: a['name'] for a in models_s.execute_kw(
                    ODOO_DB, uid_s, ODOO_PASSWORD,
                    'product.attribute', 'search_read',
                    [[['id', 'in', adids]]], {'fields': ['id', 'name'], 'limit': 100}
                )}

        # Para cada SKU, preferir el registro con id más alto (más reciente)
        best: dict = {}
        for p in odoo_prods:
            dc = p['default_code']
            if dc not in best or p['id'] > best[dc]['id']:
                attrs_lower: dict = {}
                for vid in p.get('product_template_attribute_value_ids', []):
                    av = av_map.get(vid)
                    if av:
                        aname = ad_map.get(av['attribute_id'][0], '').lower()
                        attrs_lower[aname] = av['name'].upper()
                raw = re.sub(r'\s*\([^)]+\)\s*', '', p['name']).strip()
                best[dc] = {
                    'id':    p['id'],
                    'name':  ' '.join(raw.split()).upper(),
                    'color': attrs_lower.get('color', '') or attrs_lower.get('colour', ''),
                    'talla': (attrs_lower.get('talla_bici', '')
                              or attrs_lower.get('talla', '')),
                }

        for dc, data in best.items():
            _SCOTT_CORRECT_NAMES[dc] = data['name']
            if data['color']:
                _SCOTT_COLORS[dc] = data['color']
            if data['talla']:
                _SCOTT_TALLAS[dc] = data['talla']
        _SCOTT_NAMES_LOADED = True
    except Exception:
        pass

# ─────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────

def _safe_obtener_conexion():
    try:
        return obtener_conexion()
    except Exception as e:
        logging.warning('[forecast] MySQL unavailable: %s', e)
        return None


def _ensure_table():
    """Create forecast_proyecciones if it doesn't exist (idempotent)."""
    conn = _safe_obtener_conexion()
    if conn is None:
        return

    cur = conn.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS forecast_proyecciones (
                id INT AUTO_INCREMENT PRIMARY KEY,
                id_cliente INT,
                clave_cliente VARCHAR(255),
                periodo VARCHAR(50),
                sku VARCHAR(255),
                producto VARCHAR(255),
                marca VARCHAR(255),
                modelo VARCHAR(255),
                color VARCHAR(255),
                talla VARCHAR(255),
                mayo INT DEFAULT 0,
                junio INT DEFAULT 0,
                julio INT DEFAULT 0,
                agosto INT DEFAULT 0,
                septiembre INT DEFAULT 0,
                octubre INT DEFAULT 0,
                noviembre INT DEFAULT 0,
                diciembre INT DEFAULT 0,
                enero INT DEFAULT 0,
                febrero INT DEFAULT 0,
                marzo INT DEFAULT 0,
                abril INT DEFAULT 0,
                creado_en DATETIME DEFAULT CURRENT_TIMESTAMP,
                actualizado_en DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_forecast_clave_periodo_sku (clave_cliente, periodo, sku)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()
    except Exception as e:
        logging.warning('[forecast] Could not ensure forecast_proyecciones table: %s', e)
    finally:
        cur.close()
        conn.close()

_ensure_table()


# ─────────────────────────────────────────────────────
# Odoo catalog sync (odoo_catalogo table)
# ─────────────────────────────────────────────────────

def _ensure_catalogo_table():
    """Create odoo_catalogo table if it doesn't exist."""
    conn = _safe_obtener_conexion()
    if conn is None:
        return

    cur = conn.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS odoo_catalogo (
                referencia_interna VARCHAR(255) PRIMARY KEY,
                nombre_producto VARCHAR(255),
                categoria VARCHAR(255),
                marca VARCHAR(255),
                color VARCHAR(255),
                talla VARCHAR(255),
                lst_price DECIMAL(12,2) DEFAULT NULL,
                actualizado_en DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                FULLTEXT INDEX idx_ft_nombre_producto (nombre_producto)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()
        # Migraciones para tablas ya existentes
        for _mig in [
            "ALTER TABLE odoo_catalogo ADD COLUMN lst_price DECIMAL(12,2) DEFAULT NULL",
        ]:
            try:
                cur.execute(_mig)
                conn.commit()
            except Exception:
                conn.rollback()
        # Migración: agregar FULLTEXT si la tabla ya existía sin él
        try:
            cur.execute("ALTER TABLE odoo_catalogo ADD FULLTEXT INDEX idx_ft_nombre_producto (nombre_producto)")
            conn.commit()
        except Exception:
            conn.rollback()  # índice ya existe, ignorar
    except Exception as e:
        logging.warning('[forecast] Could not ensure odoo_catalogo table: %s', e)
    finally:
        cur.close()
        conn.close()


_ensure_catalogo_table()
_catalogo_sync_lock = threading.Lock()
_catalogo_syncing   = False


# ─────────────────────────────────────────────────────
# Excel Product Catalog (forecast_excel_productos table)
# ─────────────────────────────────────────────────────
# Allows loading products from Excel before they exist in Odoo
# Used as validation source for product SKUs in forecasts

def _ensure_excel_producto_table():
    """Create forecast_excel_productos table if it doesn't exist."""
    conn = _safe_obtener_conexion()
    if conn is None:
        return

    cur = conn.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS forecast_excel_productos (
                sku VARCHAR(255) PRIMARY KEY,
                nombre VARCHAR(255),
                color VARCHAR(255),
                talla VARCHAR(255),
                marca VARCHAR(120) DEFAULT NULL,
                modelo VARCHAR(255) DEFAULT NULL,
                origen VARCHAR(50) DEFAULT 'excel',
                cargado_en DATETIME DEFAULT CURRENT_TIMESTAMP,
                actualizado_en DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()
        # Agregar columnas si la tabla ya existía sin ellas
        for col_sql in [
            "ALTER TABLE forecast_excel_productos ADD COLUMN origen VARCHAR(50) DEFAULT 'excel'",
            "ALTER TABLE forecast_excel_productos ADD COLUMN marca  VARCHAR(120) DEFAULT NULL",
            "ALTER TABLE forecast_excel_productos ADD COLUMN modelo VARCHAR(255) DEFAULT NULL",
            "ALTER TABLE forecast_excel_productos ADD COLUMN categoria VARCHAR(255) DEFAULT NULL",
            "ALTER TABLE forecast_excel_productos ADD COLUMN precio_distribuidor DECIMAL(10,2) DEFAULT NULL",
            "ALTER TABLE forecast_excel_productos ADD COLUMN precio_partner DECIMAL(10,2) DEFAULT NULL",
            "ALTER TABLE forecast_excel_productos ADD COLUMN precio_partner_elite DECIMAL(10,2) DEFAULT NULL",
            "ALTER TABLE forecast_excel_productos ADD COLUMN precio_partner_elite_plus DECIMAL(10,2) DEFAULT NULL",
            "ALTER TABLE forecast_excel_productos ADD COLUMN precio_publico DECIMAL(10,2) DEFAULT NULL",
        ]:
            try:
                cur.execute(col_sql)
                conn.commit()
            except Exception:
                conn.rollback()  # columna ya existe, ignorar
    except Exception as e:
        logging.warning('[forecast] Could not ensure forecast_excel_productos table: %s', e)
    finally:
        cur.close()
        conn.close()


_ensure_excel_producto_table()


# ─────────────────────────────────────────────────────
# SKU Whitelist (forecast_sku_whitelist table)
# ─────────────────────────────────────────────────────

def _ensure_sku_whitelist_table():
    conn = _safe_obtener_conexion()
    if conn is None:
        return
    cur = conn.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS forecast_sku_whitelist (
                sku VARCHAR(255) PRIMARY KEY,
                cargado_en DATETIME DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()
    except Exception as e:
        logging.warning('[forecast] Could not ensure forecast_sku_whitelist table: %s', e)
    finally:
        cur.close()
        conn.close()


_ensure_sku_whitelist_table()


_whitelist_db_loaded = False

def _update_whitelist_skus():
    """Populate forecast_sku_whitelist table with FORECAST_SKU_WHITELIST (cached once per process)."""
    global _whitelist_db_loaded
    if _whitelist_db_loaded:
        return
    conn = _safe_obtener_conexion()
    if conn is None:
        return
    cur = conn.cursor()
    try:
        # Clear existing
        cur.execute("DELETE FROM forecast_sku_whitelist")
        # Insert new
        for sku in FORECAST_SKU_WHITELIST:
            cur.execute("INSERT IGNORE INTO forecast_sku_whitelist (sku) VALUES (%s)", (sku,))
        conn.commit()
        logging.info('[whitelist] Updated with %d SKUs', len(FORECAST_SKU_WHITELIST))
        _whitelist_db_loaded = True
    except Exception as e:
        logging.warning('[whitelist] Error updating: %s', e)
        conn.rollback()
    finally:
        cur.close()
        conn.close()


_update_whitelist_skus()

def _get_whitelist_products() -> list:
    """Returns product info from odoo_catalogo for whitelist SKUs (stubs for missing ones)."""
    conn = _safe_obtener_conexion()
    if conn is None:
        return []
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT sku FROM forecast_sku_whitelist ORDER BY sku")
        all_skus = [r['sku'] for r in cur.fetchall()]
        if not all_skus:
            return []

        placeholders = ','.join(['%s'] * len(all_skus))
        cur.execute(f"""
            SELECT referencia_interna AS sku, nombre_producto AS nombre,
                   categoria, marca, color, talla,
                   COALESCE(lst_price, 0) AS lst_price
            FROM odoo_catalogo
            WHERE referencia_interna IN ({placeholders})
            ORDER BY marca, nombre_producto
        """, all_skus)
        rows       = cur.fetchall()
        found_skus = {r['sku'] for r in rows}

        result = []
        for r in rows:
            cat    = r.get('categoria') or ''
            modelo = cat.split(' / ')[-1].strip() if ' / ' in cat else ''
            result.append({
                'sku':       r['sku'] or '',
                'producto':  (r.get('nombre') or '').strip(),
                'marca':     (r.get('marca') or '').strip(),
                'modelo':    modelo,
                'color':     (r.get('color') or '').upper().strip(),
                'talla':     (r.get('talla') or '').upper().strip(),
                'lst_price': float(r.get('lst_price') or 0.0),
            })
        for sku in all_skus:
            if sku not in found_skus:
                result.append({'sku': sku, 'producto': '', 'marca': '',
                                'modelo': '', 'color': '', 'talla': ''})
        return result
    except Exception as e:
        logging.warning('[forecast] _get_whitelist_products error: %s', e)
        return []
    finally:
        cur.close()
        conn.close()


def _get_odoo_prices_for_skus(refs: list) -> dict:
    """
    Returns {sku: {list_price, 'Partner Elite Plus!', 'Partner Elite', 'Partner', 'Distribuidor'}}
    Reads pricelist items for the 4 TIER_NAMES pricelists.
    Falls back to zeros on any Odoo error.
    """
    empty = lambda: {'list_price': 0.0, **{t: 0.0 for t in TIER_NAMES}}
    result = {ref: empty() for ref in refs}
    if not refs:
        return result
    try:
        from utils.odoo_utils import get_odoo_models, ODOO_DB, ODOO_PASSWORD
        uid, models, err = get_odoo_models()
        if not uid:
            logging.warning('[prices] Cannot connect to Odoo: %s', err)
            return result

        prods = models.execute_kw(ODOO_DB, uid, ODOO_PASSWORD,
            'product.product', 'search_read',
            [[['default_code', 'in', refs]]],
            {'fields': ['id', 'default_code', 'lst_price', 'product_tmpl_id']}
        )
        prod_id_by_ref = {}
        tmpl_id_by_ref = {}
        for p in prods:
            ref = (p.get('default_code') or '').strip()
            if ref in result:
                result[ref]['list_price'] = float(p.get('lst_price') or 0.0)
                prod_id_by_ref[ref] = p['id']
                if p.get('product_tmpl_id'):
                    tmpl_id_by_ref[ref] = p['product_tmpl_id'][0]

        if not prod_id_by_ref:
            return result

        # Buscar todas las listas activas y mapear por keyword (robusto ante sufijos como "(MXN)")
        all_pls = models.execute_kw(ODOO_DB, uid, ODOO_PASSWORD,
            'product.pricelist', 'search_read',
            [[['active', '=', True]]],
            {'fields': ['id', 'name']}
        )
        pl_tier_list = []
        for pl in all_pls:
            n = pl['name'].upper()
            if 'PRECIO PUBLICO' in n or 'PRECIO PÚBLICO' in n:
                pl_tier_list.append({'id': pl['id'], 'name': '__public__'})
            elif 'PARTNER ELITE PLUS' in n:
                pl_tier_list.append({'id': pl['id'], 'name': 'Partner Elite Plus!'})
            elif 'PARTNER ELITE' in n:
                pl_tier_list.append({'id': pl['id'], 'name': 'Partner Elite'})
            elif 'PARTNER' in n:
                pl_tier_list.append({'id': pl['id'], 'name': 'Partner'})
            elif 'DISTRIBUIDOR' in n:
                pl_tier_list.append({'id': pl['id'], 'name': 'Distribuidor'})
        if not pl_tier_list:
            logging.warning('[prices] No pricelists matched in Odoo (active pricelists: %s)',
                            [p['name'] for p in all_pls])
            return result
        pricelists = pl_tier_list

        all_prod_ids = list(prod_id_by_ref.values())
        all_tmpl_ids = list(set(tmpl_id_by_ref.values()))
        PRIORITY = {'0_product_variant': 0, '1_product': 1,
                    '2_product_category': 2, '3_global': 3}

        for pl in pricelists:
            pl_id = pl['id']
            tier  = pl['name']
            items = models.execute_kw(ODOO_DB, uid, ODOO_PASSWORD,
                'product.pricelist.item', 'search_read',
                [[
                    ['pricelist_id', '=', pl_id],
                    '|', '|', '|',
                    ['product_id', 'in', all_prod_ids],
                    ['product_tmpl_id', 'in', all_tmpl_ids],
                    ['applied_on', '=', '2_product_category'],
                    ['applied_on', '=', '3_global'],
                ]],
                {'fields': ['applied_on', 'product_id', 'product_tmpl_id',
                            'compute_price', 'fixed_price', 'percent_price',
                            'price_discount', 'price_surcharge']}
            )
            for ref in refs:
                if ref not in result:
                    continue
                prod_id    = prod_id_by_ref.get(ref)
                tmpl_id    = tmpl_id_by_ref.get(ref)
                list_price = result[ref]['list_price']
                best_item, best_prio = None, 999
                for item in items:
                    prio = PRIORITY.get(item.get('applied_on', '3_global'), 999)
                    if prio >= best_prio:
                        continue
                    ao = item.get('applied_on')
                    if ao == '0_product_variant':
                        if not item.get('product_id') or item['product_id'][0] != prod_id:
                            continue
                    elif ao == '1_product':
                        if not item.get('product_tmpl_id') or item['product_tmpl_id'][0] != tmpl_id:
                            continue
                    best_item, best_prio = item, prio
                if best_item:
                    compute = best_item.get('compute_price', 'fixed')
                    if compute == 'fixed':
                        price = float(best_item.get('fixed_price') or 0.0)
                    elif compute == 'percentage':
                        pct   = float(best_item.get('percent_price') or 0.0)
                        price = round(list_price * (1 - pct / 100), 2)
                    else:
                        disc      = float(best_item.get('price_discount') or 0.0)
                        surcharge = float(best_item.get('price_surcharge') or 0.0)
                        price = round(list_price * (1 - disc / 100) + surcharge, 2)
                    if tier == '__public__':
                        # Precio Público → sobrescribe list_price (base para columna G)
                        if price > 0:
                            result[ref]['list_price'] = price
                    else:
                        result[ref][tier] = price
        return result
    except Exception as e:
        logging.exception('[prices] Error fetching Odoo prices: %s', e)
        return result


def _get_single_pricelist_prices(pricelist_id: int, refs: list) -> dict:
    """
    Retorna {sku: precio_sin_iva} para los refs usando la lista de precios indicada por ID.
    Misma lógica de resolución de ítems que _get_odoo_prices_for_skus().
    """
    result = {r: 0.0 for r in refs}
    if not refs or not pricelist_id:
        return result
    try:
        from utils.odoo_utils import get_odoo_models, ODOO_DB, ODOO_PASSWORD
        uid, models, err = get_odoo_models()
        if not uid:
            logging.warning('[prices] No Odoo connection for pricelist %s: %s', pricelist_id, err)
            return result

        prods = models.execute_kw(ODOO_DB, uid, ODOO_PASSWORD,
            'product.product', 'search_read',
            [[['default_code', 'in', refs]]],
            {'fields': ['id', 'default_code', 'lst_price', 'product_tmpl_id']}
        )
        prod_id_by_ref    = {}
        tmpl_id_by_ref    = {}
        list_price_by_ref = {}
        for p in prods:
            ref = (p.get('default_code') or '').strip()
            if ref in result:
                prod_id_by_ref[ref]    = p['id']
                list_price_by_ref[ref] = float(p.get('lst_price') or 0.0)
                if p.get('product_tmpl_id'):
                    tmpl_id_by_ref[ref] = p['product_tmpl_id'][0]

        if not prod_id_by_ref:
            return result

        all_prod_ids = list(prod_id_by_ref.values())
        all_tmpl_ids = list(set(tmpl_id_by_ref.values()))
        PRIORITY = {'0_product_variant': 0, '1_product': 1,
                    '2_product_category': 2, '3_global': 3}

        items = models.execute_kw(ODOO_DB, uid, ODOO_PASSWORD,
            'product.pricelist.item', 'search_read',
            [[
                ['pricelist_id', '=', pricelist_id],
                '|', '|', '|',
                ['product_id',       'in', all_prod_ids],
                ['product_tmpl_id',  'in', all_tmpl_ids],
                ['applied_on', '=', '2_product_category'],
                ['applied_on', '=', '3_global'],
            ]],
            {'fields': ['applied_on', 'product_id', 'product_tmpl_id',
                        'compute_price', 'fixed_price', 'percent_price',
                        'price_discount', 'price_surcharge']}
        )

        for ref in refs:
            prod_id    = prod_id_by_ref.get(ref)
            tmpl_id    = tmpl_id_by_ref.get(ref)
            list_price = list_price_by_ref.get(ref, 0.0)
            if not prod_id:
                continue
            best_item, best_prio = None, 999
            for item in items:
                prio = PRIORITY.get(item.get('applied_on', '3_global'), 999)
                if prio >= best_prio:
                    continue
                ao = item.get('applied_on')
                if ao == '0_product_variant':
                    if not item.get('product_id') or item['product_id'][0] != prod_id:
                        continue
                elif ao == '1_product':
                    if not item.get('product_tmpl_id') or item['product_tmpl_id'][0] != tmpl_id:
                        continue
                best_item, best_prio = item, prio
            if best_item:
                compute = best_item.get('compute_price', 'fixed')
                if compute == 'fixed':
                    result[ref] = float(best_item.get('fixed_price') or 0.0)
                elif compute == 'percentage':
                    pct = float(best_item.get('percent_price') or 0.0)
                    result[ref] = round(list_price * (1 - pct / 100), 2)
                else:
                    disc      = float(best_item.get('price_discount') or 0.0)
                    surcharge = float(best_item.get('price_surcharge') or 0.0)
                    result[ref] = round(list_price * (1 - disc / 100) + surcharge, 2)
        return result
    except Exception as e:
        logging.exception('[prices] _get_single_pricelist_prices(%s) error: %s', pricelist_id, e)
        return result


def _get_product_from_sources(sku: str) -> dict | None:
    """
    Busca un producto por SKU en ambas fuentes (Excel primero, luego Odoo).
    Retorna dict con keys: sku, nombre, color, talla, origen
    o None si no existe en ninguna fuente.
    """
    conn = obtener_conexion()
    cur = conn.cursor(dictionary=True)
    try:
        # Buscar primero en Excel (tiene prioridad)
        cur.execute("""
            SELECT sku, nombre, color, talla, origen
            FROM forecast_excel_productos
            WHERE sku = %s AND origen = 'excel'
        """, (sku,))
        row = cur.fetchone()
        if row:
            return row

        # Fallback a Odoo catalog
        cur.execute("""
            SELECT referencia_interna AS sku, 
                   nombre_producto AS nombre,
                   color,
                   talla,
                   'odoo' AS origen
            FROM odoo_catalogo
            WHERE referencia_interna = %s
        """, (sku,))
        row = cur.fetchone()
        if row:
            return row

        return None
    finally:
        cur.close()
        conn.close()


def _sync_catalogo_odoo_task():
    """Fetch all active product variants from Odoo and upsert into odoo_catalogo."""
    global _catalogo_syncing
    try:
        from utils.odoo_utils import get_odoo_models, ODOO_DB, ODOO_PASSWORD
        uid, models, err = get_odoo_models()
        if not uid:
            logging.warning('[catalogo_sync] Could not connect to Odoo: %s', err)
            return

        batch_size = 500
        offset     = 0
        total_upserted = 0
        conn = obtener_conexion()
        cur  = conn.cursor()

        while True:
            records = models.execute_kw(
                ODOO_DB, uid, ODOO_PASSWORD,
                'product.product', 'search_read',
                [[['active', '=', True]]],
                {'fields': ['id', 'default_code', 'name', 'categ_id',
                            'product_template_attribute_value_ids', 'lst_price'],
                 'limit': batch_size, 'offset': offset,
                 'order': 'id asc'}
            )
            if not records:
                break

            # Batch-fetch variant attribute values (color, talla) for all products in this page
            all_ptav_ids = []
            for p in records:
                all_ptav_ids.extend(p.get('product_template_attribute_value_ids') or [])
            ptav_map = {}  # ptav_id → {'attr': str, 'val': str}
            if all_ptav_ids:
                ptav_recs = models.execute_kw(
                    ODOO_DB, uid, ODOO_PASSWORD,
                    'product.template.attribute.value', 'read',
                    [all_ptav_ids],
                    {'fields': ['id', 'attribute_id', 'name']}
                )
                for pv in ptav_recs:
                    attr_name = ((pv.get('attribute_id') or [None, ''])[1] or '').upper()
                    ptav_map[pv['id']] = {
                        'attr': attr_name,
                        'val':  (pv.get('name') or '').upper().strip()
                    }

            rows = []
            for p in records:
                ref  = (p.get('default_code') or '').strip()
                if not ref:
                    ref = f'ODOO:{p["id"]}'  # synthetic SKU for products without referencia_interna
                nombre   = (p.get('name') or '').upper().strip()
                categ    = p.get('categ_id', [None, ''])
                categoria = (categ[1] if categ and len(categ) > 1 else '').strip()
                # Odoo paths: 'All / MEGAMO / ...' or '25161506 / MEGAMO / ...'
                # Skip root "All" and numeric codes to get the real brand
                _parts = [s.strip() for s in categoria.split(' / ')] if categoria else []
                _meaningful = [s for s in _parts if s and not s.isdigit() and s.upper() != 'ALL']
                marca = _meaningful[0] if _meaningful else (_parts[0] if _parts else '')
                # Extract color and talla from Odoo variant attributes
                color = ''
                talla = ''
                for ptav_id in (p.get('product_template_attribute_value_ids') or []):
                    pv = ptav_map.get(ptav_id)
                    if not pv:
                        continue
                    if any(k in pv['attr'] for k in ('COLOR', 'COLO', 'COLOUR')):
                        color = pv['val']
                    elif any(k in pv['attr'] for k in ('TALLA', 'TAMAÑO', 'SIZE', 'TAMA')):
                        talla = pv['val']
                lst_price = float(p.get('lst_price') or 0.0)
                rows.append((ref, nombre, categoria, marca, color, talla, lst_price))

            if rows:
                cur.executemany(
                    """
                    INSERT INTO odoo_catalogo
                        (referencia_interna, nombre_producto, categoria, marca, color, talla, lst_price)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE
                        nombre_producto = VALUES(nombre_producto),
                        categoria       = VALUES(categoria),
                        marca           = VALUES(marca),
                        color           = VALUES(color),
                        talla           = VALUES(talla),
                        lst_price       = VALUES(lst_price),
                        actualizado_en  = NOW()
                    """,
                    rows
                )
                conn.commit()
                total_upserted += len(rows)

            if len(records) < batch_size:
                break
            offset += batch_size

        cur.close()
        conn.close()
        logging.info('[catalogo_sync] Done — %d products upserted.', total_upserted)
    except Exception as exc:
        logging.exception('[catalogo_sync] Error: %s', exc)
    finally:
        with _catalogo_sync_lock:
            _catalogo_syncing = False


def _trigger_catalogo_sync(force: bool = False):
    """Launch a background sync if not already running (and table is empty or force=True)."""
    global _catalogo_syncing
    with _catalogo_sync_lock:
        if _catalogo_syncing:
            return 'already_running'
        if not force:
            conn = _safe_obtener_conexion()
            if conn is None:
                return 'db_unavailable'
            try:
                cur = conn.cursor()
                cur.execute('SELECT COUNT(*) as cnt FROM odoo_catalogo')
                cnt = cur.fetchone()[0]
            except Exception as exc:
                logging.warning('[forecast] Could not evaluate odoo_catalogo count: %s', exc)
                return 'db_unavailable'
            finally:
                cur.close()
                conn.close()

            if cnt > 0:
                return 'already_populated'
        _catalogo_syncing = True

    t = threading.Thread(target=_sync_catalogo_odoo_task, daemon=True, name='catalogo_sync')
    t.start()
    return 'started'


# Auto-sync on startup if the catalog is empty
_trigger_catalogo_sync(force=False)


SIZE_RE = re.compile(r'^(XS|S|M|L|XL|XXL|XXXL|TU|\d{1,3})$', re.IGNORECASE)
CATEGORY_PREFIXES = [
    'BICICLETA', 'BICI', 'CASCO', 'GUANTE', 'GUANTES', 'LENTE', 'LENTES', 'BOLSO',
    'MOCHILA', 'ZAPATILLA', 'ZAPATILLAS', 'ZAPATO', 'ZAPATOS', 'JERSEY',
    'SHORTS', 'CHAMARRA', 'GORRA', 'GAFAS', 'ACCESORIO', 'ACCESORIOS',
    'ROPA', 'MANUBRIO', 'SILLA', 'SILLÍN', 'RUEDA',
]

def _parse_color_talla(descripcion: str, modelo: str) -> tuple:
    """Heuristically extract (color, talla) from a product descripcion string."""
    if not descripcion:
        return '', ''
    text = descripcion.upper().strip()
    modelo_up = modelo.upper().strip() if modelo else ''
    # Strip year suffix like MY26, MY2026 from modelo before matching
    modelo_base = re.sub(r'\s+MY\d{2,4}$', '', modelo_up).strip()
    # Try to remove modelo (with or without year suffix)
    for m in [modelo_up, modelo_base]:
        if m and m in text:
            text = text.replace(m, '').strip()
            break
    # Remove category prefix
    for prefix in CATEGORY_PREFIXES:
        if text.startswith(prefix + ' ') or text == prefix:
            text = text[len(prefix):].strip()
            break
    tokens = text.split()
    if not tokens:
        return '', ''
    last = tokens[-1]
    if SIZE_RE.match(last):
        talla = last
        color = ' '.join(tokens[:-1]).strip()
    else:
        talla = ''
        color = ' '.join(tokens).strip()
    return color, talla


def _clean_producto(descripcion: str, color: str, talla: str) -> str:
    """Remove color/talla tokens from the end of a product description."""
    name = descripcion.strip()
    if talla:
        name = re.sub(r'\s+' + re.escape(talla) + r'\s*$', '', name, flags=re.IGNORECASE).strip()
    if color:
        name = re.sub(r'\s+' + re.escape(color) + r'\s*$', '', name, flags=re.IGNORECASE).strip()
    return name


def _get_client_id(clave_cliente: str):
    conn = obtener_conexion()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id FROM clientes WHERE clave = %s LIMIT 1", (clave_cliente,))
        row = cur.fetchone()
        return row['id'] if row else None
    finally:
        cur.close()
        conn.close()


def _get_authorized_products(clave_cliente: str, id_cliente: int) -> list:
    """
    Returns list of dicts with keys: sku, producto, marca, modelo, color, talla.
    Uses proyecciones_cliente → proyecciones_ventas as primary source.
    Falls back to full proyecciones_ventas catalog if no rows found.
    """
    conn = obtener_conexion()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT DISTINCT
                pv.clave_factura              AS sku,
                pv.descripcion               AS producto,
                pv.modelo                    AS modelo,
                pv.spec                      AS spec,
                COALESCE(oc.marca, m.marca, '') AS marca
            FROM proyecciones_ventas pv
            JOIN proyecciones_cliente pc ON pc.id_proyeccion = pv.id
            LEFT JOIN odoo_catalogo oc ON oc.referencia_interna = pv.clave_odoo
            LEFT JOIN (
                SELECT referencia_interna, MAX(marca) AS marca
                FROM monitor
                WHERE marca IS NOT NULL AND marca != ''
                GROUP BY referencia_interna
            ) m ON m.referencia_interna = pv.clave_factura
            WHERE pc.id_cliente = %s
            ORDER BY pv.clave_factura
        """, (id_cliente,))
        rows = cur.fetchall()

        if not rows:
            cur.execute("""
                SELECT DISTINCT
                    pv.clave_factura              AS sku,
                    pv.descripcion               AS producto,
                    pv.modelo                    AS modelo,
                    pv.spec                      AS spec,
                    COALESCE(oc.marca, m.marca, '') AS marca
                FROM proyecciones_ventas pv
                LEFT JOIN odoo_catalogo oc ON oc.referencia_interna = pv.clave_odoo
                LEFT JOIN (
                    SELECT referencia_interna, MAX(marca) AS marca
                    FROM monitor
                    WHERE marca IS NOT NULL AND marca != ''
                    GROUP BY referencia_interna
                ) m ON m.referencia_interna = pv.clave_factura
                ORDER BY pv.clave_factura
            """)
            rows = cur.fetchall()

        result = []
        for r in rows:
            color, talla = _parse_color_talla(r['producto'] or '', r['modelo'] or '')
            result.append({
                'sku':      r['sku'] or '',
                'producto': _clean_producto(r['producto'] or '', color, talla),
                'marca':    r['marca'] or '',
                'modelo':   r['modelo'] or '',
                'color':    color,
                'talla':    talla,
            })
        return result
    finally:
        cur.close()
        conn.close()


def _validate_periodo(periodo: str) -> bool:
    """Validates format YYYY-YYYY with second year = first + 1."""
    m = re.match(r'^(\d{4})-(\d{4})$', periodo or '')
    if not m:
        return False
    a1, a2 = int(m.group(1)), int(m.group(2))
    return a2 == a1 + 1


# ─────────────────────────────────────────────────────
# Routes
# ─────────────────────────────────────────────────────

@forecast_bp.route('/forecast/template', methods=['GET'])
def descargar_template():
    """
    GET /forecast/template?clave=<clave_cliente>&periodo=<periodo>
    Devuelve un xlsx con los productos del whitelist (solo estos productos de Odoo).
    Columnas A-H bloqueadas. G1 = selector de nivel de distribuidor (desbloqueado).
    Cols I-T (meses) editables. Cols V-Y ocultas con los 4 precios por nivel.
    """
    if not OPENPYXL_OK:
        return jsonify({'error': 'openpyxl no instalado en el servidor'}), 500

    # Update whitelist with allowed SKUs
    _update_whitelist_skus()

    clave   = request.args.get('clave',   '').strip()
    periodo = request.args.get('periodo', '').strip()

    if not clave:
        return jsonify({'error': 'Falta parámetro clave'}), 400
    if not _validate_periodo(periodo):
        return jsonify({'error': 'Formato de periodo inválido (use YYYY-YYYY)'}), 400

    id_cliente = _get_client_id(clave)
    if id_cliente is None:
        return jsonify({'error': f'Cliente "{clave}" no encontrado'}), 404

    # Fuente de productos: solo Megamo del whitelist (la plantilla Scott se gestiona aparte)
    products = [p for p in _get_whitelist_products()
                if (p.get('marca') or '').upper() == 'MEGAMO']

    # Precios: Odoo pricelists → SKU_CATALOG → lst_price del catálogo → factores estándar Megamo
    skus        = [p['sku'] for p in products]
    prices      = _get_odoo_prices_for_skus(skus) if skus else {}
    # Mapa sku → lst_price del catálogo sincronizado (ya en products)
    catalog_lsp = {p['sku']: p.get('lst_price', 0.0) for p in products}

    # Factores de descuento estándar Megamo (derivados de SKU_CATALOG)
    _MEGAMO_TIER_FACTORS = {
        'Partner Elite Plus!': 0.695,
        'Partner Elite':       0.715,
        'Partner':             0.740,
        'Distribuidor':        0.760,
    }

    for sku in skus:
        cat_entry  = SKU_CATALOG.get(sku, {})
        cat_prices = cat_entry.get('prices', {})
        odoo_entry = prices.get(sku, {})

        # 1° Odoo pricelists (ya en odoo_entry), 2° SKU_CATALOG, 3° lst_price del catálogo
        for key in ['list_price'] + TIER_NAMES:
            if not odoo_entry.get(key):
                odoo_entry[key] = cat_prices.get(key, 0.0)
        if not odoo_entry.get('list_price'):
            odoo_entry['list_price'] = catalog_lsp.get(sku, 0.0)

        # 4° Fallback: si tenemos precio público pero faltan tiers, calcular con factores Megamo
        lp = odoo_entry.get('list_price', 0.0)
        if lp > 0:
            for tier, factor in _MEGAMO_TIER_FACTORS.items():
                if not odoo_entry.get(tier):
                    odoo_entry[tier] = round(lp * factor, 2)

        # 5° Último recurso: back-calcular PVP desde precio Distribuidor (Megamo: 76% del PVP)
        if not odoo_entry.get('list_price'):
            dist_price = float(odoo_entry.get('Distribuidor') or 0)
            if dist_price > 0:
                odoo_entry['list_price'] = round(dist_price / 0.760, 2)

        prices[sku] = odoo_entry

    # ── Índices de columnas (1-based) ──────────────────────────────────────────
    # A-F (1-6): CAMPOS_INFO
    # G (7):     Precio Público
    # H (8):     Precio [nivel distribuidor] — fórmula dinámica
    # I-T (9-20): meses May-Abr
    # U (21):    TOTAL
    # V-Y (22-25): precios por nivel (ocultas)
    PRICE_PUB_COL   = 7
    PRICE_DIST_COL  = 8
    MONTH_START     = 9
    TOTAL_COL       = 21   # U — total unidades
    TOTAL_PRICE_COL = 22   # V — total precio (H × U)
    TIER_COLS = {
        'Partner Elite Plus!': 23,   # W oculta
        'Partner Elite':       24,   # X oculta
        'Partner':             25,   # Y oculta
        'Distribuidor':        26,   # Z oculta
    }
    VISIBLE_COLS = TOTAL_PRICE_COL  # columnas visibles: 1-22

    # ── Estilos ────────────────────────────────────────────────────────────────
    ORANGE      = 'FFEB5E28'
    DARK_BG     = 'FF252422'
    HEADER_BG   = 'FF1A1918'
    SELECTOR_BG = 'FF2C2A28'
    PRICE_BG    = 'FF1B3A2B'

    info_font       = Font(bold=True, color='FFFFFFFF', size=10)
    price_hdr_font  = Font(bold=True, color='FF66FFB2', size=10)
    month_hdr_font  = Font(bold=True, color='FFFFFFFF', size=10)
    editable_font   = Font(color='FF111111', size=10)
    price_data_font = Font(color='FF333333', size=9)

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right',  vertical='center')
    thin   = Side(style='thin', color='FF666666')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Forecast {periodo}'

    # ── Fila 1: Selector de nivel de distribuidor ──────────────────────────────
    ws.row_dimensions[1].height = 32
    ws.merge_cells('A1:F1')
    lbl = ws['A1']
    lbl.value     = 'TIPO DE DISTRIBUIDOR  ▶'
    lbl.font      = Font(bold=True, color='FFFFCC00', size=11)
    lbl.fill      = PatternFill('solid', fgColor=SELECTOR_BG)
    lbl.alignment = right

    ws.merge_cells('G1:H1')
    sel = ws['G1']
    sel.value      = 'Distribuidor'
    sel.font       = Font(bold=True, color='FFEB5E28', size=12)
    sel.fill       = PatternFill('solid', fgColor=SELECTOR_BG)
    sel.alignment  = center
    sel.protection = Protection(locked=False)   # única celda editable de A-H

    dv = DataValidation(
        type='list',
        formula1='"Partner Elite Plus!,Partner Elite,Partner,Distribuidor"',
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add(ws['G1'])

    for ci in range(MONTH_START, VISIBLE_COLS + 1):
        ws.cell(row=1, column=ci).fill = PatternFill('solid', fgColor=SELECTOR_BG)

    # ── Fila 2: Título ─────────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 28
    ws.merge_cells(f'A2:{get_column_letter(VISIBLE_COLS)}2')
    tc = ws['A2']
    tc.value     = f'Forecast de Compra — Periodo Comercial {periodo}   |   Distribuidor: {clave}'
    tc.font      = Font(bold=True, color='FFEB5E28', size=12)
    tc.fill      = PatternFill('solid', fgColor=HEADER_BG)
    tc.alignment = center

    # ── Fila 3: Encabezados de columnas ───────────────────────────────────────
    BLOCKED_MONTH_LABELS = {'May', 'Jun'}   # columnas bloqueadas en toda la hoja
    ws.row_dimensions[3].height = 22
    ALL_HEADERS = CAMPOS_INFO + ['Precio Público', 'Precio'] + MESES_LABELS + ['TOTAL', 'Total $']
    for ci, h in enumerate(ALL_HEADERS, start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.alignment = center
        cell.border    = border
        if h in CAMPOS_INFO:
            cell.fill = PatternFill('solid', fgColor=DARK_BG)
            cell.font = info_font
        elif h in ('Precio Público', 'Precio'):
            cell.fill = PatternFill('solid', fgColor=PRICE_BG)
            cell.font = price_hdr_font
        elif h in ('TOTAL', 'Total $'):
            cell.fill = PatternFill('solid', fgColor=ORANGE)
            cell.font = Font(bold=True, color='FF000000', size=10)
        elif h in BLOCKED_MONTH_LABELS:
            # Mayo y Junio: header gris oscuro — no disponibles para pedido
            cell.fill = PatternFill('solid', fgColor='FF2E2E2E')
            cell.font = Font(bold=True, color='FF666666', size=10)
            cell.value = f'{h}\n⛔'
        else:
            cell.fill = PatternFill('solid', fgColor=ORANGE)
            cell.font = month_hdr_font

    # ── Fila 4: Instrucciones ──────────────────────────────────────────────────
    ws.row_dimensions[4].height = 38
    ws.merge_cells(f'A4:{get_column_letter(VISIBLE_COLS)}4')
    note = ws['A4']
    _instr_font = InlineFont(i=False, color='FF333333', sz=9)
    _bold_font  = InlineFont(b=True,  color='FF111111', sz=9)
    _warn_font  = InlineFont(b=True,  color='FFCC4400', sz=9)
    note.value = CellRichText(
        TextBlock(_bold_font,  '① NIVEL DE PRECIO: '),
        TextBlock(_instr_font, 'Seleccione su tipo en la celda naranja G1 (Distribuidor / Partner / Partner Elite / Partner Elite Plus!) '
                               '— la columna H se actualiza automáticamente.     '),
        TextBlock(_bold_font,  '② DISPONIBILIDAD: '),
        TextBlock(_instr_font, 'Las columnas May y Jun están '),
        TextBlock(_warn_font,  'BLOQUEADAS ⛔'),
        TextBlock(_instr_font, ' — solo puede capturar a partir de '),
        TextBlock(_bold_font,  'Julio. '),
        TextBlock(_instr_font, 'Celdas oscuras dentro de los meses disponibles indican que el modelo llega más tarde.     '),
        TextBlock(_bold_font,  '③ CAPTURA: '),
        TextBlock(_instr_font, 'Ingrese las CANTIDADES por mes. Los totales (unidades y monto) se calculan solos.     '),
        TextBlock(_warn_font,  '⚡ Entre más rápido envíe sus proyecciones, mayor prioridad tendrá su pedido. '
                               'Envíe este archivo a su Ejecutivo de Ventas.'),
    )
    note.fill      = PatternFill('solid', fgColor='FFFFF8F0')
    note.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ── Filas de datos (fila 5 en adelante) ───────────────────────────────────
    tier_pep_col = get_column_letter(TIER_COLS['Partner Elite Plus!'])
    tier_pe_col  = get_column_letter(TIER_COLS['Partner Elite'])
    tier_p_col   = get_column_letter(TIER_COLS['Partner'])
    tier_d_col   = get_column_letter(TIER_COLS['Distribuidor'])
    h_col        = get_column_letter(PRICE_DIST_COL)
    u_col        = get_column_letter(TOTAL_COL)
    first_m      = get_column_letter(MONTH_START)
    last_m       = get_column_letter(MONTH_START + len(MESES) - 1)
    first_data_row = 5

    for row_idx, p in enumerate(products, start=5):
        sku         = p['sku']
        prod_prices = prices.get(sku, {})

        # A-F: información del producto (bloqueada con la protección de hoja)
        for ci, val in enumerate(
            [sku, p['producto'], p['marca'], p['modelo'], p['color'], p['talla']], start=1
        ):
            c = ws.cell(row=row_idx, column=ci, value=val)
            c.font      = editable_font
            c.fill      = PatternFill('solid', fgColor='FFFAFAFA')
            c.alignment = left if ci == 2 else center
            c.border    = border

        # G: Precio Público con IVA (bloqueado)
        g = ws.cell(row=row_idx, column=PRICE_PUB_COL)
        g.value         = round(prod_prices.get('list_price', 0.0) * IVA_FACTOR, 2)
        g.font          = price_data_font
        g.fill          = PatternFill('solid', fgColor='FFE8F5E9')
        g.alignment     = center
        g.border        = border
        g.number_format = '"$"#,##0.00'

        # H: Precio por nivel con IVA — fórmula que lee G1 (selector) y columnas ocultas W-Z
        h = ws.cell(row=row_idx, column=PRICE_DIST_COL)
        h.value = (
            f'=IF($G$1="Partner Elite Plus!",{tier_pep_col}{row_idx},'
            f'IF($G$1="Partner Elite",{tier_pe_col}{row_idx},'
            f'IF($G$1="Partner",{tier_p_col}{row_idx},{tier_d_col}{row_idx})))'
        )
        h.font          = price_data_font
        h.fill          = PatternFill('solid', fgColor='FFF3E5F5')
        h.alignment     = center
        h.border        = border
        h.number_format = '"$"#,##0.00'

        # I-T: meses — Mayo y Junio siempre bloqueados; demás meses según SKU_CATALOG
        _MESES_BLOQ_TEMPLATE = {'mayo', 'junio'}
        cat_avail = SKU_CATALOG.get(sku, {}).get('avail', {})
        for mi in range(len(MESES)):
            mes_name = MESES[mi]
            # Mayo y junio bloqueados para todos; resto según disponibilidad del modelo
            is_avail = False if mes_name in _MESES_BLOQ_TEMPLATE else cat_avail.get(mes_name, True)
            c = ws.cell(row=row_idx, column=MONTH_START + mi)
            c.alignment     = center
            c.border        = border
            c.number_format = '0'
            if is_avail:
                c.value      = 0
                c.font       = editable_font
                c.fill       = PatternFill('solid', fgColor='FFFEFEFE')
                c.protection = Protection(locked=False)
            else:
                c.value = None
                c.font  = Font(color='FF888888', size=9)
                c.fill  = PatternFill('solid', fgColor='FF3A3A3A')

        # U: TOTAL unidades (fórmula, bloqueado)
        tc2 = ws.cell(row=row_idx, column=TOTAL_COL)
        tc2.value         = f'=SUM({first_m}{row_idx}:{last_m}{row_idx})'
        tc2.font          = Font(bold=True, color='FF000000', size=10)
        tc2.fill          = PatternFill('solid', fgColor='FFFFF0D0')
        tc2.alignment     = center
        tc2.border        = border
        tc2.number_format = '0'

        # V: TOTAL precio = H × U (fórmula, bloqueado)
        tp = ws.cell(row=row_idx, column=TOTAL_PRICE_COL)
        tp.value         = f'={h_col}{row_idx}*{u_col}{row_idx}'
        tp.font          = Font(bold=True, color='FF000000', size=10)
        tp.fill          = PatternFill('solid', fgColor='FFE8F0FF')
        tp.alignment     = center
        tp.border        = border
        tp.number_format = '"$"#,##0.00'

        # W-Z: precios por nivel con IVA (ocultas, referenciadas por la fórmula de H)
        for tier, col_idx in TIER_COLS.items():
            pc = ws.cell(row=row_idx, column=col_idx)
            pc.value         = round(prod_prices.get(tier, 0.0) * IVA_FACTOR, 2)
            pc.number_format = '"$"#,##0.00'

    # ── Fila de TOTALES ────────────────────────────────────────────────────────────
    last_data_row = first_data_row + len(products) - 1
    total_row     = last_data_row + 1

    ws.row_dimensions[total_row].height = 24

    # A-H: Label "TOTALES"
    ws.merge_cells(f'A{total_row}:H{total_row}')
    label = ws[f'A{total_row}']
    label.value     = 'TOTALES'
    label.font      = Font(bold=True, color='FFFFFFFF', size=11)
    label.fill      = PatternFill('solid', fgColor=ORANGE)
    label.alignment = center
    label.border    = border

    # I-T: suma de cada mes
    for mi in range(len(MESES)):
        col_letter = get_column_letter(MONTH_START + mi)
        c = ws.cell(row=total_row, column=MONTH_START + mi)
        c.value         = f'=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})'
        c.font          = Font(bold=True, color='FF000000', size=10)
        c.fill          = PatternFill('solid', fgColor=ORANGE)
        c.alignment     = center
        c.border        = border
        c.number_format = '0'

    # U: suma de total unidades
    tu = ws.cell(row=total_row, column=TOTAL_COL)
    tu.value         = f'=SUM({u_col}{first_data_row}:{u_col}{last_data_row})'
    tu.font          = Font(bold=True, color='FFFFFFFF', size=11)
    tu.fill          = PatternFill('solid', fgColor=ORANGE)
    tu.alignment     = center
    tu.border        = border
    tu.number_format = '0'

    # V: suma de total precio
    v_col_letter = get_column_letter(TOTAL_PRICE_COL)
    tp_total = ws.cell(row=total_row, column=TOTAL_PRICE_COL)
    tp_total.value         = f'=SUM({v_col_letter}{first_data_row}:{v_col_letter}{last_data_row})'
    tp_total.font          = Font(bold=True, color='FFFFFFFF', size=11)
    tp_total.fill          = PatternFill('solid', fgColor=ORANGE)
    tp_total.alignment     = center
    tp_total.border        = border
    tp_total.number_format = '"$"#,##0.00'

    # ── Fila de PRECIO POR MES (azul) — precio × cantidad por cada mes, sin repetir total final ──
    price_row = total_row + 1
    ws.row_dimensions[price_row].height = 24

    ws.merge_cells(f'A{price_row}:H{price_row}')
    label2 = ws[f'A{price_row}']
    label2.value     = 'TOTAL PRECIO POR MES'
    label2.font      = Font(bold=True, color='FFFFFFFF', size=11)
    label2.fill      = PatternFill('solid', fgColor='FF1B5E9C')
    label2.alignment = center
    label2.border    = border

    for mi in range(len(MESES)):
        col_letter = get_column_letter(MONTH_START + mi)
        c = ws.cell(row=price_row, column=MONTH_START + mi)
        c.value         = f'=SUMPRODUCT(${h_col}${first_data_row}:${h_col}${last_data_row},{col_letter}${first_data_row}:{col_letter}${last_data_row})'
        c.font          = Font(bold=True, color='FFFFFFFF', size=10)
        c.fill          = PatternFill('solid', fgColor='FF1B5E9C')
        c.alignment     = center
        c.border        = border
        c.number_format = '"$"#,##0.00'

    # U vacío (el total de unidades ya está en la fila naranja)
    cu = ws.cell(row=price_row, column=TOTAL_COL)
    cu.fill      = PatternFill('solid', fgColor='FF1B5E9C')
    cu.border    = border

    # V vacío — el total anual en precio ya está en la fila naranja, no se repite
    cv = ws.cell(row=price_row, column=TOTAL_PRICE_COL)
    cv.fill      = PatternFill('solid', fgColor='FF1B5E9C')
    cv.border    = border

    # ── Ocultar columnas de precios por nivel (W-Z) ─────────────────────────────────
    for col_idx in range(23, 27):
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    # ── Anchos de columna ──────────────────────────────────────────────────────
    col_widths = [18, 42, 16, 22, 14, 8, 14, 14] + [13] * 12 + [9, 18]
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Congelar primeras 4 filas ──────────────────────────────────────────────
    ws.freeze_panes = 'A5'

    # ── Protección de hoja: A-H y U bloqueadas; G1 y columnas I-T desbloqueadas ─
    ws.protection.sheet    = True
    ws.protection.password = 'masterkey'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'Forecast_{clave}_{periodo}.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@forecast_bp.route('/forecast/template-global', methods=['GET'])
def descargar_template_global():
    """
    GET /forecast/template-global
    Plantilla global sin cliente específico — el distribuidor ingresa su clave en B1.
    Todos los 92 SKUs del whitelist, mismo layout que template() pero portable.
    """
    if not OPENPYXL_OK:
        return jsonify({'error': 'openpyxl no instalado en el servidor'}), 500

    _update_whitelist_skus()

    # Usar período comercial actual como default (e.g., "2026-2027")
    from datetime import datetime
    current_year = datetime.now().year
    periodo = f"{current_year}-{current_year + 1}"

    products    = [p for p in _get_whitelist_products()
                   if (p.get('marca') or '').upper() == 'MEGAMO']
    skus        = [p['sku'] for p in products]
    prices      = _get_odoo_prices_for_skus(skus) if skus else {}
    catalog_lsp = {p['sku']: p.get('lst_price', 0.0) for p in products}

    _MEG_FACTORS = {
        'Partner Elite Plus!': 0.695,
        'Partner Elite':       0.715,
        'Partner':             0.740,
        'Distribuidor':        0.760,
    }

    # Fallback completo: Odoo → SKU_CATALOG → lst_price → factores Megamo → back-calc desde Dist.
    for sku in skus:
        cat_entry  = SKU_CATALOG.get(sku, {})
        cat_prices = cat_entry.get('prices', {})
        odoo_entry = prices.get(sku, {})
        for key in ['list_price'] + TIER_NAMES:
            if not odoo_entry.get(key):
                odoo_entry[key] = cat_prices.get(key, 0.0)
        if not odoo_entry.get('list_price'):
            odoo_entry['list_price'] = catalog_lsp.get(sku, 0.0)
        lp = float(odoo_entry.get('list_price') or 0)
        if lp > 0:
            for tier, factor in _MEG_FACTORS.items():
                if not odoo_entry.get(tier):
                    odoo_entry[tier] = round(lp * factor, 2)
        if not odoo_entry.get('list_price'):
            dist_price = float(odoo_entry.get('Distribuidor') or 0)
            if dist_price > 0:
                odoo_entry['list_price'] = round(dist_price / 0.760, 2)
        prices[sku] = odoo_entry

    # ── Índices de columnas (igual que template())
    PRICE_PUB_COL   = 7
    PRICE_DIST_COL  = 8
    MONTH_START     = 9
    TOTAL_COL       = 21
    TOTAL_PRICE_COL = 22
    TIER_COLS = {
        'Partner Elite Plus!': 23,
        'Partner Elite':       24,
        'Partner':             25,
        'Distribuidor':        26,
    }
    VISIBLE_COLS = TOTAL_PRICE_COL

    # ── Estilos (igual)
    ORANGE      = 'FFEB5E28'
    DARK_BG     = 'FF252422'
    HEADER_BG   = 'FF1A1918'
    SELECTOR_BG = 'FF2C2A28'
    PRICE_BG    = 'FF1B3A2B'

    info_font       = Font(bold=True, color='FFFFFFFF', size=10)
    price_hdr_font  = Font(bold=True, color='FF66FFB2', size=10)
    month_hdr_font  = Font(bold=True, color='FFFFFFFF', size=10)
    editable_font   = Font(color='FF111111', size=10)
    price_data_font = Font(color='FF333333', size=9)

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right',  vertical='center')
    thin   = Side(style='thin', color='FF666666')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Forecast Global'

    # ── Fila 1: Campo para CLAVE DISTRIBUIDOR ─────────────────────────────────────
    ws.row_dimensions[1].height = 28
    ws.merge_cells('A1:C1')
    lbl = ws['A1']
    lbl.value     = 'CLAVE / NOMBRE DISTRIBUIDOR'
    lbl.font      = Font(bold=True, color='FFFFFFFF', size=10)
    lbl.fill      = PatternFill('solid', fgColor=SELECTOR_BG)
    lbl.alignment = right

    ws.merge_cells('D1:F1')
    inp = ws['D1']
    inp.value      = ''  # Campo editable para que ingrese su clave
    inp.font       = Font(bold=True, color='FFEB5E28', size=11)
    inp.fill       = PatternFill('solid', fgColor=SELECTOR_BG)
    inp.alignment  = center
    inp.protection = Protection(locked=False)  # EDITABLE

    ws.merge_cells('G1:H1')
    sel = ws['G1']
    sel.value      = 'Distribuidor'
    sel.font       = Font(bold=True, color='FFEB5E28', size=12)
    sel.fill       = PatternFill('solid', fgColor=SELECTOR_BG)
    sel.alignment  = center
    sel.protection = Protection(locked=False)

    dv = DataValidation(
        type='list',
        formula1='"Partner Elite Plus!,Partner Elite,Partner,Distribuidor"',
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add(ws['G1'])

    for ci in range(MONTH_START, VISIBLE_COLS + 1):
        ws.cell(row=1, column=ci).fill = PatternFill('solid', fgColor=SELECTOR_BG)

    # ── Fila 2: Título ─────────────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 28
    ws.merge_cells(f'A2:{get_column_letter(VISIBLE_COLS)}2')
    tc = ws['A2']
    tc.value     = f'Plantilla de Forecast — Periodo Comercial {periodo}'
    tc.font      = Font(bold=True, color='FFEB5E28', size=12)
    tc.fill      = PatternFill('solid', fgColor=HEADER_BG)
    tc.alignment = center

    # ── Fila 3: Encabezados de columnas ─────────────────────────────────────────────
    _BLOCKED_LBL = {'May', 'Jun'}
    ws.row_dimensions[3].height = 22
    ALL_HEADERS = CAMPOS_INFO + ['Precio Público', 'Precio'] + MESES_LABELS + ['TOTAL', 'Total $']
    for ci, h in enumerate(ALL_HEADERS, start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.alignment = center
        cell.border    = border
        if h in CAMPOS_INFO:
            cell.fill = PatternFill('solid', fgColor=DARK_BG)
            cell.font = info_font
        elif h in ('Precio Público', 'Precio'):
            cell.fill = PatternFill('solid', fgColor=PRICE_BG)
            cell.font = price_hdr_font
        elif h in ('TOTAL', 'Total $'):
            cell.fill = PatternFill('solid', fgColor=ORANGE)
            cell.font = Font(bold=True, color='FF000000', size=10)
        elif h in _BLOCKED_LBL:
            cell.fill  = PatternFill('solid', fgColor='FF2E2E2E')
            cell.font  = Font(bold=True, color='FF666666', size=10)
            cell.value = f'{h}\n⛔'
        else:
            cell.fill = PatternFill('solid', fgColor=ORANGE)
            cell.font = month_hdr_font

    # ── Fila 4: Instrucciones ──────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 38
    ws.merge_cells(f'A4:{get_column_letter(VISIBLE_COLS)}4')
    note = ws['A4']
    _instr_font2 = InlineFont(i=False, color='FF333333', sz=9)
    _bold_font2  = InlineFont(b=True,  color='FF111111', sz=9)
    _warn_font2  = InlineFont(b=True,  color='FFCC4400', sz=9)
    note.value = CellRichText(
        TextBlock(_bold_font2,  '① CLAVE: '),
        TextBlock(_instr_font2, 'Escriba su nombre o clave en el campo gris (celda D1).     '),
        TextBlock(_bold_font2,  '② NIVEL DE PRECIO: '),
        TextBlock(_instr_font2, 'Seleccione su tipo en la celda naranja G1 (Distribuidor / Partner / Partner Elite / Partner Elite Plus!) '
                                '— la columna H se actualiza automáticamente.     '),
        TextBlock(_bold_font2,  '③ DISPONIBILIDAD: '),
        TextBlock(_instr_font2, 'Las columnas May y Jun están '),
        TextBlock(_warn_font2,  'BLOQUEADAS ⛔'),
        TextBlock(_instr_font2, ' — solo puede capturar a partir de '),
        TextBlock(_bold_font2,  'Julio. '),
        TextBlock(_instr_font2, 'Celdas oscuras indican que el modelo llega más tarde.     '),
        TextBlock(_bold_font2,  '④ CAPTURA: '),
        TextBlock(_instr_font2, 'Ingrese cantidades por mes. Totales se calculan solos.     '),
        TextBlock(_warn_font2,  '⚡ Entre más rápido envíe sus proyecciones, mayor prioridad tendrá su pedido.'),
    )
    note.fill      = PatternFill('solid', fgColor='FFFFF8F0')
    note.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ── Filas de datos (fila 5 en adelante) ────────────────────────────────────────
    tier_pep_col = get_column_letter(TIER_COLS['Partner Elite Plus!'])
    tier_pe_col  = get_column_letter(TIER_COLS['Partner Elite'])
    tier_p_col   = get_column_letter(TIER_COLS['Partner'])
    tier_d_col   = get_column_letter(TIER_COLS['Distribuidor'])
    h_col        = get_column_letter(PRICE_DIST_COL)
    u_col        = get_column_letter(TOTAL_COL)
    first_m      = get_column_letter(MONTH_START)
    last_m       = get_column_letter(MONTH_START + len(MESES) - 1)
    first_data_row = 5

    for row_idx, p in enumerate(products, start=5):
        sku         = p['sku']
        prod_prices = prices.get(sku, {})

        for ci, val in enumerate(
            [sku, p['producto'], p['marca'], p['modelo'], p['color'], p['talla']], start=1
        ):
            c = ws.cell(row=row_idx, column=ci, value=val)
            c.font      = editable_font
            c.fill      = PatternFill('solid', fgColor='FFFAFAFA')
            c.alignment = left if ci == 2 else center
            c.border    = border

        g = ws.cell(row=row_idx, column=PRICE_PUB_COL)
        g.value         = round(prod_prices.get('list_price', 0.0) * IVA_FACTOR, 2)
        g.font          = price_data_font
        g.fill          = PatternFill('solid', fgColor='FFE8F5E9')
        g.alignment     = center
        g.border        = border
        g.number_format = '"$"#,##0.00'

        h = ws.cell(row=row_idx, column=PRICE_DIST_COL)
        h.value = (
            f'=IF($G$1="Partner Elite Plus!",{tier_pep_col}{row_idx},'
            f'IF($G$1="Partner Elite",{tier_pe_col}{row_idx},'
            f'IF($G$1="Partner",{tier_p_col}{row_idx},{tier_d_col}{row_idx})))'
        )
        h.font          = price_data_font
        h.fill          = PatternFill('solid', fgColor='FFF3E5F5')
        h.alignment     = center
        h.border        = border
        h.number_format = '"$"#,##0.00'

        _MESES_BLOQ_G = {'mayo', 'junio'}
        cat_avail = SKU_CATALOG.get(sku, {}).get('avail', {})
        for mi in range(len(MESES)):
            mes_name = MESES[mi]
            is_avail = False if mes_name in _MESES_BLOQ_G else cat_avail.get(mes_name, True)
            c = ws.cell(row=row_idx, column=MONTH_START + mi)
            c.alignment     = center
            c.border        = border
            c.number_format = '0'
            if is_avail:
                c.value      = 0
                c.font       = editable_font
                c.fill       = PatternFill('solid', fgColor='FFFEFEFE')
                c.protection = Protection(locked=False)
            else:
                c.value = None
                c.font  = Font(color='FF888888', size=9)
                c.fill  = PatternFill('solid', fgColor='FF3A3A3A')

        tc2 = ws.cell(row=row_idx, column=TOTAL_COL)
        tc2.value         = f'=SUM({first_m}{row_idx}:{last_m}{row_idx})'
        tc2.font          = Font(bold=True, color='FF000000', size=10)
        tc2.fill          = PatternFill('solid', fgColor='FFFFF0D0')
        tc2.alignment     = center
        tc2.border        = border
        tc2.number_format = '0'

        tp = ws.cell(row=row_idx, column=TOTAL_PRICE_COL)
        tp.value         = f'={h_col}{row_idx}*{u_col}{row_idx}'
        tp.font          = Font(bold=True, color='FF000000', size=10)
        tp.fill          = PatternFill('solid', fgColor='FFE8F0FF')
        tp.alignment     = center
        tp.border        = border
        tp.number_format = '"$"#,##0.00'

        for tier, col_idx in TIER_COLS.items():
            pc = ws.cell(row=row_idx, column=col_idx)
            pc.value         = round(prod_prices.get(tier, 0.0) * IVA_FACTOR, 2)
            pc.number_format = '"$"#,##0.00'

    # ── Fila de TOTALES
    last_data_row = first_data_row + len(products) - 1
    total_row     = last_data_row + 1

    ws.row_dimensions[total_row].height = 24

    ws.merge_cells(f'A{total_row}:H{total_row}')
    label = ws[f'A{total_row}']
    label.value     = 'TOTALES'
    label.font      = Font(bold=True, color='FFFFFFFF', size=11)
    label.fill      = PatternFill('solid', fgColor=ORANGE)
    label.alignment = center
    label.border    = border

    for mi in range(len(MESES)):
        col_letter = get_column_letter(MONTH_START + mi)
        c = ws.cell(row=total_row, column=MONTH_START + mi)
        c.value         = f'=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})'
        c.font          = Font(bold=True, color='FF000000', size=10)
        c.fill          = PatternFill('solid', fgColor=ORANGE)
        c.alignment     = center
        c.border        = border
        c.number_format = '0'

    tu = ws.cell(row=total_row, column=TOTAL_COL)
    tu.value         = f'=SUM({u_col}{first_data_row}:{u_col}{last_data_row})'
    tu.font          = Font(bold=True, color='FFFFFFFF', size=11)
    tu.fill          = PatternFill('solid', fgColor=ORANGE)
    tu.alignment     = center
    tu.border        = border
    tu.number_format = '0'

    v_col_letter = get_column_letter(TOTAL_PRICE_COL)
    tp_total = ws.cell(row=total_row, column=TOTAL_PRICE_COL)
    tp_total.value         = f'=SUM({v_col_letter}{first_data_row}:{v_col_letter}{last_data_row})'
    tp_total.font          = Font(bold=True, color='FFFFFFFF', size=11)
    tp_total.fill          = PatternFill('solid', fgColor=ORANGE)
    tp_total.alignment     = center
    tp_total.border        = border
    tp_total.number_format = '"$"#,##0.00'

    # ── Fila de PRECIO POR MES (azul) — precio × cantidad por cada mes, sin repetir total final ──
    price_row = total_row + 1
    ws.row_dimensions[price_row].height = 24

    ws.merge_cells(f'A{price_row}:H{price_row}')
    label2 = ws[f'A{price_row}']
    label2.value     = 'TOTAL PRECIO POR MES'
    label2.font      = Font(bold=True, color='FFFFFFFF', size=11)
    label2.fill      = PatternFill('solid', fgColor='FF1B5E9C')
    label2.alignment = center
    label2.border    = border

    for mi in range(len(MESES)):
        col_letter = get_column_letter(MONTH_START + mi)
        c = ws.cell(row=price_row, column=MONTH_START + mi)
        c.value         = f'=SUMPRODUCT(${h_col}${first_data_row}:${h_col}${last_data_row},{col_letter}${first_data_row}:{col_letter}${last_data_row})'
        c.font          = Font(bold=True, color='FFFFFFFF', size=10)
        c.fill          = PatternFill('solid', fgColor='FF1B5E9C')
        c.alignment     = center
        c.border        = border
        c.number_format = '"$"#,##0.00'

    cu = ws.cell(row=price_row, column=TOTAL_COL)
    cu.fill   = PatternFill('solid', fgColor='FF1B5E9C')
    cu.border = border

    cv = ws.cell(row=price_row, column=TOTAL_PRICE_COL)
    cv.fill   = PatternFill('solid', fgColor='FF1B5E9C')
    cv.border = border

    # ── Ocultar columnas de precios por nivel (W-Z)
    for col_idx in range(23, 27):
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    # ── Anchos de columna
    col_widths = [18, 42, 16, 22, 14, 8, 14, 14] + [13] * 12 + [9, 18]
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes        = 'A5'
    ws.protection.sheet    = True
    ws.protection.password = 'masterkey'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'Forecast_Template_Global_{periodo}.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )



@forecast_bp.route('/forecast/template-global-scott', methods=['GET'])
def descargar_template_global_scott():
    """
    GET /forecast/template-global-scott
    Plantilla global solo con las bicicletas Scott MY27 nuevas (52 SKUs),
    con disponibilidad por mes según fechas de llegada.
    """
    if not OPENPYXL_OK:
        return jsonify({'error': 'openpyxl no instalado en el servidor'}), 500

    from datetime import datetime
    current_year = datetime.now().year
    periodo = f"{current_year}-{current_year + 1}"

    # Todos los SKUs Scott del catálogo
    SCOTT_SKUS_LIST = [
        '427536-8086006','427536-8086008','427536-8086010',
        '427539-8350006','427539-8350008','427539-8350010',
        '427537-8532006','427537-8532008','427537-8532010',
        '427540-8527006','427540-8527008','427540-8527010','427540-8527012',
        '427563-8535008','427563-8535010',
        '427590-8522002','427590-8522004','427590-8522006','427590-8522008','427590-8522010',
        '427590-3020002','427590-3020004','427590-3020006','427590-3020008','427590-3020010',
        '427591-8086002','427591-8086004','427591-8086006','427591-8086008','427591-8086010',
        '427588-8575004','427588-8575006','427588-8575008','427588-8575010',
        '427589-8350004','427589-8350006','427589-8350008','427589-8350010',
        '427541-0002006','427541-0002008','427541-0002010','427541-0002012',
        '427598-8575004','427598-8575006','427598-8575008',
        '427995-8585006','427995-8585008','427995-8585010',
        '427986-8565222','427986-8588222','427987-1494222','427987-8606222',
        # Scott adicionales MY26
        '286383-704','286383-706',
        '290310-704','290310-706','290310-908',
        '425790-3761222','425790-8269222',
        '425791-3028222','425791-8268222',
        '425792-2308222','425792-4173222',
        '425793-8265222',
        '425794-3774222',
    ]

    # Consultar Odoo para obtener nombre, marca, talla por variante
    from utils.odoo_utils import get_odoo_models, ODOO_DB, ODOO_PASSWORD
    uid, models_obj, err = get_odoo_models()
    products = []
    if uid and not err:
        try:
            _ODOO_CTX = {'lang': 'es_MX'}  # nombres en español (traducción activa en Odoo)
            prods = models_obj.execute_kw(ODOO_DB, uid, ODOO_PASSWORD,
                'product.product', 'search_read',
                [[['default_code', 'in', SCOTT_SKUS_LIST]]],
                {'fields': ['id', 'default_code', 'name', 'list_price',
                            'product_tmpl_id',
                            'product_template_attribute_value_ids'],
                 'limit': 300, 'context': _ODOO_CTX}
            )
            # Obtener el nombre canónico del template en español
            tmpl_ids = list({p['product_tmpl_id'][0] for p in prods if p.get('product_tmpl_id')})
            tmpl_name_map = {}
            if tmpl_ids:
                tmpls = models_obj.execute_kw(ODOO_DB, uid, ODOO_PASSWORD,
                    'product.template', 'search_read',
                    [[['id', 'in', tmpl_ids]]],
                    {'fields': ['id', 'name'], 'limit': 300, 'context': _ODOO_CTX}
                )
                tmpl_name_map = {t['id']: t['name'] for t in tmpls}

            # Contar cuántos atributos tiene cada template (más = más completo = correcto)
            tmpl_attr_count: dict = {}
            for p in prods:
                tmpl_id = (p.get('product_tmpl_id') or [None])[0]
                if tmpl_id:
                    tmpl_attr_count[tmpl_id] = tmpl_attr_count.get(tmpl_id, 0) + \
                        len(p.get('product_template_attribute_value_ids', []))

            # Para cada SKU, elegir el template con más atributos (producto actualizado en Odoo)
            sku_best_tmpl: dict = {}
            for p in prods:
                dc = p['default_code']
                tmpl_id = (p.get('product_tmpl_id') or [None])[0]
                if not tmpl_id:
                    continue
                cnt = tmpl_attr_count.get(tmpl_id, 0)
                cur = sku_best_tmpl.get(dc, (-1, 0))
                if cnt > cur[0] or (cnt == cur[0] and tmpl_id > cur[1]):
                    sku_best_tmpl[dc] = (cnt, tmpl_id)

            # Sobreescribir name con el del template más completo para ese SKU
            for p in prods:
                dc = p['default_code']
                best = sku_best_tmpl.get(dc)
                if best:
                    best_tmpl_id = best[1]
                    if best_tmpl_id in tmpl_name_map:
                        p['name'] = tmpl_name_map[best_tmpl_id]

            # Obtener atributos (Talla_Bici / COLOR / TALLA)
            avids = list({v for p in prods for v in p.get('product_template_attribute_value_ids', [])})
            attr_vals = models_obj.execute_kw(ODOO_DB, uid, ODOO_PASSWORD,
                'product.template.attribute.value', 'search_read',
                [[['id', 'in', avids]]],
                {'fields': ['id', 'name', 'attribute_id'], 'limit': 500}
            ) if avids else []
            av_map = {a['id']: a for a in attr_vals}
            adids = list({a['attribute_id'][0] for a in attr_vals})
            ad_map = {a['id']: a['name'] for a in models_obj.execute_kw(
                ODOO_DB, uid, ODOO_PASSWORD, 'product.attribute', 'search_read',
                [[['id', 'in', adids]]], {'fields': ['id', 'name'], 'limit': 50}
            )} if adids else {}

            import re as _re
            _COLOR_WORDS = {
                'NEGRO','NEGRA','BLANCO','BLANCA','ROJO','ROJA','AZUL','VERDE',
                'NARANJA','AMARILLO','AMARILLA','GRIS','DORADO','DORADA',
                'PLATEADO','PLATEADA','PETROLEO','PETRÓLEO','MORADO','MORADA',
                'ROSA','CAFE','BEIGE','MARRON','BRONCE','ORO','PLATA','COBRE',
                'TURQUESA','SALMON','ARENA','CREMA','LILA','CELESTE',
                'BCO','NGO','NEGRO/BLANCO','BLANCO/NEGRO',
            }

            def _parse_scott(name):
                name2 = _re.sub(r'^\d+\s+', '', name)
                name2 = _re.sub(r'^BICICLETA\s+', '', name2)
                # Quitar sufijos en paréntesis como (M), (MORADO), etc.
                name2 = _re.sub(r'\s*\([^)]+\)\s*$', '', name2).strip()
                parts = name2.split()
                marca = parts[0] if parts else 'SCOTT'
                mod, col, after_year = [], [], False
                for w in parts[1:]:
                    if w in ('MY27', 'MY26', 'MY25', 'MY28', 'MY24'):
                        after_year = True
                        continue
                    (col if after_year else mod).append(w)
                return marca, ' '.join(mod), ' '.join(col)

            # Agrupar todos los registros por default_code (Odoo puede tener duplicados)
            raw_prod_map: dict = {}
            attrs_ids_by_sku: dict = {}
            for p in prods:
                dc = p['default_code']
                raw_prod_map.setdefault(dc, []).append(p)
                attrs_ids_by_sku.setdefault(dc, set()).update(
                    p.get('product_template_attribute_value_ids', [])
                )

            # Para cada SKU preferir el registro cuyo nombre empiece con el prefijo numérico
            prod_map = {}
            for dc, records in raw_prod_map.items():
                prefix = dc[:6]
                good = [r for r in records if r.get('name', '').strip().startswith(prefix)]
                # Entre candidatos: mayor id = registro más reciente en Odoo (nombre actualizado)
                prod_map[dc] = max(good or records, key=lambda r: r.get('id', 0))

            for sku in SCOTT_SKUS_LIST:
                p = prod_map.get(sku)
                if not p:
                    continue
                # Construir attrs combinando IDs de TODOS los registros duplicados del mismo SKU
                attrs = {}
                attrs_lower = {}
                for vid in attrs_ids_by_sku.get(sku, []):
                    av = av_map.get(vid)
                    if av:
                        key = ad_map.get(av['attribute_id'][0], '?')
                        attrs[key] = av['name']
                        attrs_lower[key.lower().replace(' ', '_').replace('/', '_')] = av['name']
                raw_name = p.get('name', '')
                # Quitar "(copia)" y otros sufijos en paréntesis que Odoo agrega
                raw_name = _re.sub(r'\s*\(copia\)\s*', ' ', raw_name, flags=_re.IGNORECASE).strip()
                raw_name = _re.sub(r'\s*\([^)]+\)\s*$', '', raw_name).strip()
                # Normalizar prefijo numérico: siempre el correcto del SKU
                sku_prefix = sku[:6]
                raw_name = _re.sub(r'^\d{6}\b', sku_prefix, raw_name)
                # Si el nombre no empieza con el número, agregarlo para consistencia entre variantes
                if not raw_name.startswith(sku_prefix):
                    raw_name = f"{sku_prefix} {raw_name.lstrip()}"
                marca, modelo, color_from_name = _parse_scott(raw_name)
                # Atributo Odoo tiene prioridad sobre abreviaturas en el nombre
                color = attrs_lower.get('color', '') or color_from_name
                # Si el atributo de color está definido, limpiar cualquier
                # palabra de color que haya quedado al final del modelo
                if color:
                    mod_parts = modelo.split()
                    if mod_parts and mod_parts[-1].upper() in _COLOR_WORDS:
                        modelo = ' '.join(mod_parts[:-1])
                # Talla: buscar TALLA / Talla_Bici / talla; si no hay → Unitalla
                talla = (attrs_lower.get('talla_bici') or attrs_lower.get('talla') or '') or 'Unitalla'
                products.append({
                    'sku':     sku,
                    'producto': raw_name,
                    'marca':   marca,
                    'modelo':  modelo,
                    'color':   color,
                    'talla':   talla,
                })
        except Exception as e:
            logging.warning('[scott-template] Odoo error: %s', e)

    # Fallback: stubs para SKUs no encontrados en Odoo
    found = {p['sku'] for p in products}
    for sku in SCOTT_SKUS_LIST:
        if sku not in found:
            products.append({'sku': sku, 'producto': sku, 'marca': 'SCOTT',
                             'modelo': '', 'color': '', 'talla': ''})

    # Ordenar por SKU
    products.sort(key=lambda p: p['sku'])

    # ── Índices de columnas
    PRICE_PUB_COL   = 7
    PRICE_DIST_COL  = 8
    MONTH_START     = 9
    TOTAL_COL       = 21
    TOTAL_PRICE_COL = 22
    TIER_COLS = {
        'Partner Elite Plus!': 23,
        'Partner Elite':       24,
        'Partner':             25,
        'Distribuidor':        26,
    }
    VISIBLE_COLS = TOTAL_PRICE_COL

    ORANGE      = 'FFEB5E28'
    DARK_BG     = 'FF252422'
    HEADER_BG   = 'FF1A1918'
    SELECTOR_BG = 'FF2C2A28'
    PRICE_BG    = 'FF1B3A2B'

    info_font       = Font(bold=True, color='FFFFFFFF', size=10)
    price_hdr_font  = Font(bold=True, color='FF66FFB2', size=10)
    month_hdr_font  = Font(bold=True, color='FFFFFFFF', size=10)
    editable_font   = Font(color='FF111111', size=10)
    price_data_font = Font(color='FF333333', size=9)

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right',  vertical='center')
    thin   = Side(style='thin', color='FF666666')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Forecast Scott'

    # Fila 1: Clave distribuidor
    ws.row_dimensions[1].height = 28
    ws.merge_cells('A1:C1')
    lbl = ws['A1']
    lbl.value = 'CLAVE / NOMBRE DISTRIBUIDOR'
    lbl.font  = Font(bold=True, color='FFFFFFFF', size=10)
    lbl.fill  = PatternFill('solid', fgColor=SELECTOR_BG)
    lbl.alignment = right

    ws.merge_cells('D1:F1')
    inp = ws['D1']
    inp.value = ''
    inp.font  = Font(bold=True, color='FFEB5E28', size=11)
    inp.fill  = PatternFill('solid', fgColor=SELECTOR_BG)
    inp.alignment = center
    inp.protection = Protection(locked=False)

    ws.merge_cells('G1:H1')
    sel = ws['G1']
    sel.value = 'Distribuidor'
    sel.font  = Font(bold=True, color='FFEB5E28', size=12)
    sel.fill  = PatternFill('solid', fgColor=SELECTOR_BG)
    sel.alignment = center
    sel.protection = Protection(locked=False)

    dv = DataValidation(type='list',
        formula1='"Partner Elite Plus!,Partner Elite,Partner,Distribuidor"',
        allow_blank=False, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(ws['G1'])

    for ci in range(MONTH_START, VISIBLE_COLS + 1):
        ws.cell(row=1, column=ci).fill = PatternFill('solid', fgColor=SELECTOR_BG)

    # Fila 2: Título
    ws.row_dimensions[2].height = 28
    ws.merge_cells(f'A2:{get_column_letter(VISIBLE_COLS)}2')
    tc = ws['A2']
    tc.value = f'Plantilla de Forecast Scott — Periodo Comercial {periodo}'
    tc.font  = Font(bold=True, color='FFEB5E28', size=12)
    tc.fill  = PatternFill('solid', fgColor=HEADER_BG)
    tc.alignment = center

    # Fila 3: Encabezados
    ws.row_dimensions[3].height = 22
    ALL_HEADERS = CAMPOS_INFO + ['Precio Público', 'Precio'] + MESES_LABELS + ['TOTAL', 'Total $']
    for ci, h in enumerate(ALL_HEADERS, start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.alignment = center
        cell.border    = border
        if h in CAMPOS_INFO:
            cell.fill = PatternFill('solid', fgColor=DARK_BG); cell.font = info_font
        elif h in ('Precio Público', 'Precio'):
            cell.fill = PatternFill('solid', fgColor=PRICE_BG); cell.font = price_hdr_font
        elif h in ('TOTAL', 'Total $'):
            cell.fill = PatternFill('solid', fgColor=ORANGE)
            cell.font = Font(bold=True, color='FF000000', size=10)
        else:
            cell.fill = PatternFill('solid', fgColor=ORANGE); cell.font = month_hdr_font

    # Fila 4: Instrucciones
    ws.row_dimensions[4].height = 26
    ws.merge_cells(f'A4:{get_column_letter(VISIBLE_COLS)}4')
    note = ws['A4']
    note.value = (
        f'📌 Forecast Scott MY27 — Periodo {periodo}  |  '
        'Ingrese su CLAVE en D1 y seleccione su NIVEL en G1.  '
        'Las celdas oscuras indican meses en que el modelo aún no está disponible.'
    )
    note.font      = Font(italic=True, color='FF444444', size=9)
    note.fill      = PatternFill('solid', fgColor='FFFFF8F0')
    note.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Filas de datos
    tier_pep_col = get_column_letter(TIER_COLS['Partner Elite Plus!'])
    tier_pe_col  = get_column_letter(TIER_COLS['Partner Elite'])
    tier_p_col   = get_column_letter(TIER_COLS['Partner'])
    tier_d_col   = get_column_letter(TIER_COLS['Distribuidor'])
    h_col        = get_column_letter(PRICE_DIST_COL)
    u_col        = get_column_letter(TOTAL_COL)
    first_m      = get_column_letter(MONTH_START)
    last_m       = get_column_letter(MONTH_START + len(MESES) - 1)
    first_data_row = 5

    for row_idx, p in enumerate(products, start=5):
        sku        = p['sku']
        cat_entry  = SKU_CATALOG.get(sku, {})
        pr         = cat_entry.get('prices', {})
        cat_avail  = cat_entry.get('avail', {})

        for ci, val in enumerate(
            [sku, p['producto'], p['marca'], p['modelo'], p['color'], p['talla']], start=1
        ):
            c = ws.cell(row=row_idx, column=ci, value=val)
            c.font      = editable_font
            c.fill      = PatternFill('solid', fgColor='FFFAFAFA')
            c.alignment = left if ci == 2 else center
            c.border    = border

        g = ws.cell(row=row_idx, column=PRICE_PUB_COL)
        g.value         = round(pr.get('list_price', 0.0) * IVA_FACTOR, 2)
        g.font          = price_data_font
        g.fill          = PatternFill('solid', fgColor='FFE8F5E9')
        g.alignment     = center; g.border = border
        g.number_format = '"$"#,##0.00'

        h = ws.cell(row=row_idx, column=PRICE_DIST_COL)
        h.value = (
            f'=IF($G$1="Partner Elite Plus!",{tier_pep_col}{row_idx},'
            f'IF($G$1="Partner Elite",{tier_pe_col}{row_idx},'
            f'IF($G$1="Partner",{tier_p_col}{row_idx},{tier_d_col}{row_idx})))'
        )
        h.font = price_data_font; h.fill = PatternFill('solid', fgColor='FFF3E5F5')
        h.alignment = center; h.border = border; h.number_format = '"$"#,##0.00'

        for mi in range(len(MESES)):
            mes_name = MESES[mi]
            is_avail = cat_avail.get(mes_name, True)
            c = ws.cell(row=row_idx, column=MONTH_START + mi)
            c.alignment = center; c.border = border; c.number_format = '0'
            if is_avail:
                c.value = 0; c.font = editable_font
                c.fill  = PatternFill('solid', fgColor='FFFEFEFE')
                c.protection = Protection(locked=False)
            else:
                c.value = None; c.font = Font(color='FF888888', size=9)
                c.fill  = PatternFill('solid', fgColor='FF3A3A3A')

        tc2 = ws.cell(row=row_idx, column=TOTAL_COL)
        tc2.value = f'=SUM({first_m}{row_idx}:{last_m}{row_idx})'
        tc2.font  = Font(bold=True, color='FF000000', size=10)
        tc2.fill  = PatternFill('solid', fgColor='FFFFF0D0')
        tc2.alignment = center; tc2.border = border; tc2.number_format = '0'

        tprice = ws.cell(row=row_idx, column=TOTAL_PRICE_COL)
        tprice.value = f'={h_col}{row_idx}*{u_col}{row_idx}'
        tprice.font  = Font(bold=True, color='FF000000', size=10)
        tprice.fill  = PatternFill('solid', fgColor='FFE8F0FF')
        tprice.alignment = center; tprice.border = border
        tprice.number_format = '"$"#,##0.00'

        for tier, col_idx in TIER_COLS.items():
            pc = ws.cell(row=row_idx, column=col_idx)
            pc.value         = round(pr.get(tier, 0.0) * IVA_FACTOR, 2)
            pc.number_format = '"$"#,##0.00'

    # Fila TOTALES
    last_data_row = first_data_row + len(products) - 1
    total_row     = last_data_row + 1
    ws.row_dimensions[total_row].height = 24
    ws.merge_cells(f'A{total_row}:H{total_row}')
    label = ws[f'A{total_row}']
    label.value = 'TOTALES'; label.font = Font(bold=True, color='FFFFFFFF', size=11)
    label.fill  = PatternFill('solid', fgColor=ORANGE)
    label.alignment = center; label.border = border

    for mi in range(len(MESES)):
        col_letter = get_column_letter(MONTH_START + mi)
        c = ws.cell(row=total_row, column=MONTH_START + mi)
        c.value         = f'=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})'
        c.font          = Font(bold=True, color='FF000000', size=10)
        c.fill          = PatternFill('solid', fgColor=ORANGE)
        c.alignment     = center; c.border = border; c.number_format = '0'

    tu = ws.cell(row=total_row, column=TOTAL_COL)
    tu.value = f'=SUM({u_col}{first_data_row}:{u_col}{last_data_row})'
    tu.font  = Font(bold=True, color='FFFFFFFF', size=11)
    tu.fill  = PatternFill('solid', fgColor=ORANGE)
    tu.alignment = center; tu.border = border; tu.number_format = '0'

    v_col = get_column_letter(TOTAL_PRICE_COL)
    tp_total = ws.cell(row=total_row, column=TOTAL_PRICE_COL)
    tp_total.value = f'=SUM({v_col}{first_data_row}:{v_col}{last_data_row})'
    tp_total.font  = Font(bold=True, color='FFFFFFFF', size=11)
    tp_total.fill  = PatternFill('solid', fgColor=ORANGE)
    tp_total.alignment = center; tp_total.border = border
    tp_total.number_format = '"$"#,##0.00'

    # Ocultar columnas de precios por nivel (W-Z)
    for col_idx in range(23, 27):
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    col_widths = [18, 42, 12, 24, 22, 8, 14, 14] + [11] * 12 + [9, 18]
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes     = 'A5'
    ws.protection.sheet = True
    ws.protection.password = 'masterkey'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'Forecast_Template_Scott_{periodo}.xlsx'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@forecast_bp.route('/forecast/template-blank', methods=['GET'])
def descargar_template_blank():
    """
    GET /forecast/template-blank
    Plantilla de forecast vacía — solo encabezados y meses, sin productos precargados.
    El distribuidor llena manualmente SKU, Producto, Marca, Modelo, Color, Talla y cantidades.
    """
    if not OPENPYXL_OK:
        return jsonify({'error': 'openpyxl no instalado en el servidor'}), 500

    from datetime import datetime
    current_year = datetime.now().year
    periodo = f"{current_year}-{current_year + 1}"

    PRICE_PUB_COL   = 7
    PRICE_DIST_COL  = 8
    MONTH_START     = 9
    TOTAL_COL       = 21
    TOTAL_PRICE_COL = 22
    VISIBLE_COLS    = TOTAL_PRICE_COL
    NUM_BLANK_ROWS  = 25

    ORANGE      = 'FFEB5E28'
    DARK_BG     = 'FF252422'
    HEADER_BG   = 'FF1A1918'
    SELECTOR_BG = 'FF2C2A28'
    PRICE_BG    = 'FF1B3A2B'

    info_font       = Font(bold=True, color='FFFFFFFF', size=10)
    price_hdr_font  = Font(bold=True, color='FF66FFB2', size=10)
    month_hdr_font  = Font(bold=True, color='FFFFFFFF', size=10)
    editable_font   = Font(color='FF111111', size=10)

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right',  vertical='center')
    thin   = Side(style='thin', color='FF666666')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Forecast'

    # Fila 1: Campo clave distribuidor
    ws.row_dimensions[1].height = 28
    ws.merge_cells('A1:C1')
    lbl = ws['A1']
    lbl.value     = 'CLAVE / NOMBRE DISTRIBUIDOR'
    lbl.font      = Font(bold=True, color='FFFFFFFF', size=10)
    lbl.fill      = PatternFill('solid', fgColor=SELECTOR_BG)
    lbl.alignment = right

    ws.merge_cells('D1:F1')
    inp = ws['D1']
    inp.value      = ''
    inp.font       = Font(bold=True, color='FFEB5E28', size=11)
    inp.fill       = PatternFill('solid', fgColor=SELECTOR_BG)
    inp.alignment  = center
    inp.protection = Protection(locked=False)

    ws.merge_cells('G1:H1')
    sel = ws['G1']
    sel.value      = 'Distribuidor'
    sel.font       = Font(bold=True, color='FFEB5E28', size=12)
    sel.fill       = PatternFill('solid', fgColor=SELECTOR_BG)
    sel.alignment  = center
    sel.protection = Protection(locked=False)

    dv = DataValidation(
        type='list',
        formula1='"Partner Elite Plus!,Partner Elite,Partner,Distribuidor"',
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add(ws['G1'])

    for ci in range(MONTH_START, VISIBLE_COLS + 1):
        ws.cell(row=1, column=ci).fill = PatternFill('solid', fgColor=SELECTOR_BG)

    # Fila 2: Título
    ws.row_dimensions[2].height = 28
    ws.merge_cells(f'A2:{get_column_letter(VISIBLE_COLS)}2')
    tc = ws['A2']
    tc.value     = f'Plantilla de Forecast — Periodo Comercial {periodo}'
    tc.font      = Font(bold=True, color='FFEB5E28', size=12)
    tc.fill      = PatternFill('solid', fgColor=HEADER_BG)
    tc.alignment = center

    # Fila 3: Encabezados
    ws.row_dimensions[3].height = 22
    ALL_HEADERS = CAMPOS_INFO + ['Precio Público', 'Precio'] + MESES_LABELS + ['TOTAL', 'Total $']
    for ci, h in enumerate(ALL_HEADERS, start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.alignment = center
        cell.border    = border
        if h in CAMPOS_INFO:
            cell.fill = PatternFill('solid', fgColor=DARK_BG)
            cell.font = info_font
        elif h in ('Precio Público', 'Precio'):
            cell.fill = PatternFill('solid', fgColor=PRICE_BG)
            cell.font = price_hdr_font
        elif h in ('TOTAL', 'Total $'):
            cell.fill = PatternFill('solid', fgColor=ORANGE)
            cell.font = Font(bold=True, color='FF000000', size=10)
        else:
            cell.fill = PatternFill('solid', fgColor=ORANGE)
            cell.font = month_hdr_font

    # Fila 4: Leyenda
    ws.row_dimensions[4].height = 22
    ws.merge_cells(f'A4:{get_column_letter(VISIBLE_COLS)}4')
    note = ws['A4']
    note.value     = f'Proyecciones de compra — Periodo Comercial {periodo}   |   Distribuidor: ______________________________   Nivel: ______________________________'
    note.font      = Font(italic=True, color='FF444444', size=9)
    note.fill      = PatternFill('solid', fgColor='FFFFF8F0')
    note.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    # Filas vacías editables (5 en adelante)
    first_data_row = 5
    last_data_row  = first_data_row + NUM_BLANK_ROWS - 1
    h_col          = get_column_letter(PRICE_DIST_COL)
    u_col          = get_column_letter(TOTAL_COL)
    first_m        = get_column_letter(MONTH_START)
    last_m         = get_column_letter(MONTH_START + len(MESES) - 1)

    for row_idx in range(first_data_row, last_data_row + 1):
        ws.row_dimensions[row_idx].height = 22
        # Columnas info (A-F)
        for ci in range(1, 7):
            c = ws.cell(row=row_idx, column=ci)
            c.font      = editable_font
            c.fill      = PatternFill('solid', fgColor='FFFAFAFA')
            c.alignment = left if ci == 2 else center
            c.border    = border
            c.protection = Protection(locked=False)

        # Precio Público (G) — editable
        g = ws.cell(row=row_idx, column=PRICE_PUB_COL)
        g.font          = Font(color='FF333333', size=9)
        g.fill          = PatternFill('solid', fgColor='FFE8F5E9')
        g.alignment     = center
        g.border        = border
        g.number_format = '"$"#,##0.00'
        g.protection    = Protection(locked=False)

        # Precio distribuidor (H) — editable
        h = ws.cell(row=row_idx, column=PRICE_DIST_COL)
        h.font          = Font(color='FF333333', size=9)
        h.fill          = PatternFill('solid', fgColor='FFF3E5F5')
        h.alignment     = center
        h.border        = border
        h.number_format = '"$"#,##0.00'
        h.protection    = Protection(locked=False)

        # Meses (I-T) — editables con 0
        for mi in range(len(MESES)):
            c = ws.cell(row=row_idx, column=MONTH_START + mi)
            c.value         = None
            c.font          = editable_font
            c.fill          = PatternFill('solid', fgColor='FFFEFEFE')
            c.alignment     = center
            c.border        = border
            c.number_format = '0'
            c.protection    = Protection(locked=False)

        # TOTAL unidades (U)
        tc2 = ws.cell(row=row_idx, column=TOTAL_COL)
        tc2.value         = f'=SUM({first_m}{row_idx}:{last_m}{row_idx})'
        tc2.font          = Font(bold=True, color='FF000000', size=10)
        tc2.fill          = PatternFill('solid', fgColor='FFFFF0D0')
        tc2.alignment     = center
        tc2.border        = border
        tc2.number_format = '0'

        # Total $ (V)
        tp = ws.cell(row=row_idx, column=TOTAL_PRICE_COL)
        tp.value         = f'={h_col}{row_idx}*{u_col}{row_idx}'
        tp.font          = Font(bold=True, color='FF000000', size=10)
        tp.fill          = PatternFill('solid', fgColor='FFE8F0FF')
        tp.alignment     = center
        tp.border        = border
        tp.number_format = '"$"#,##0.00'

    # Fila de TOTALES
    total_row = last_data_row + 1
    ws.row_dimensions[total_row].height = 24
    ws.merge_cells(f'A{total_row}:H{total_row}')
    label = ws[f'A{total_row}']
    label.value     = 'TOTALES'
    label.font      = Font(bold=True, color='FFFFFFFF', size=11)
    label.fill      = PatternFill('solid', fgColor=ORANGE)
    label.alignment = center
    label.border    = border

    for mi in range(len(MESES)):
        col_letter = get_column_letter(MONTH_START + mi)
        c = ws.cell(row=total_row, column=MONTH_START + mi)
        c.value         = f'=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})'
        c.font          = Font(bold=True, color='FF000000', size=10)
        c.fill          = PatternFill('solid', fgColor=ORANGE)
        c.alignment     = center
        c.border        = border
        c.number_format = '0'

    tu = ws.cell(row=total_row, column=TOTAL_COL)
    tu.value         = f'=SUM({u_col}{first_data_row}:{u_col}{last_data_row})'
    tu.font          = Font(bold=True, color='FFFFFFFF', size=11)
    tu.fill          = PatternFill('solid', fgColor=ORANGE)
    tu.alignment     = center
    tu.border        = border
    tu.number_format = '0'

    v_col_letter = get_column_letter(TOTAL_PRICE_COL)
    tp_total = ws.cell(row=total_row, column=TOTAL_PRICE_COL)
    tp_total.value         = f'=SUM({v_col_letter}{first_data_row}:{v_col_letter}{last_data_row})'
    tp_total.font          = Font(bold=True, color='FFFFFFFF', size=11)
    tp_total.fill          = PatternFill('solid', fgColor=ORANGE)
    tp_total.alignment     = center
    tp_total.border        = border
    tp_total.number_format = '"$"#,##0.00'

    # Anchos de columna
    col_widths = [18, 42, 16, 22, 14, 8, 14, 14] + [13] * 12 + [9, 18]
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = 'A5'

    # Configuración de impresión — una sola hoja horizontal
    last_col_letter = get_column_letter(VISIBLE_COLS)
    ws.print_area   = f'A1:{last_col_letter}{total_row}'
    ws.page_setup.orientation   = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize     = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage     = True
    ws.page_setup.fitToWidth    = 1
    ws.page_setup.fitToHeight   = 1
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = '1:4'   # Repetir encabezados si llega a más de una hoja

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'Forecast_Plantilla_Vacia_{periodo}.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@forecast_bp.route('/forecast/importar', methods=['POST'])
def importar_forecast():
    """
    POST /forecast/importar  (multipart/form-data)
    Fields: clave_cliente (opcional si el archivo global tiene clave en D1),
            periodo (opcional — se lee de A2 si está vacío), file (xlsx).
    Validates and upserts rows into forecast_proyecciones.
    """
    if not OPENPYXL_OK:
        return jsonify({'error': 'openpyxl no instalado en el servidor'}), 500

    clave   = request.form.get('clave_cliente', '').strip()
    periodo = request.form.get('periodo', '').strip()
    archivo = request.files.get('file')

    if not archivo:
        return jsonify({'error': 'Falta el archivo Excel'}), 400

    ext = archivo.filename.rsplit('.', 1)[-1].lower() if archivo.filename else ''
    if ext not in ('xlsx', 'xls'):
        return jsonify({'error': 'El archivo debe ser Excel (.xlsx o .xls)'}), 400

    # Cargar el workbook una vez para poder leer clave/periodo del propio archivo si no vienen en el form
    try:
        content = archivo.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'No se pudo leer el archivo Excel: {str(e)}'}), 400

    # Leer clave desde D1 (plantilla global) si no viene en el form
    if not clave:
        d1_val = ws.cell(row=1, column=4).value  # D1 = campo de clave en plantilla global
        if d1_val:
            clave = str(d1_val).strip()

    # Leer periodo desde la fila de título (A2) si no viene en el form
    if not periodo:
        titulo = str(ws.cell(row=2, column=1).value or '')
        m_periodo = re.search(r'(\d{4}-\d{4})', titulo)
        if m_periodo:
            periodo = m_periodo.group(1)
        else:
            # Default: periodo comercial actual (Mayo–Abril)
            now = datetime.now()
            inicio = now.year if now.month >= 4 else now.year - 1
            periodo = f'{inicio}-{inicio + 1}'

    if not clave:
        return jsonify({'error': 'No se encontró la clave del distribuidor. '
                                 'Completa la celda de clave en la plantilla o envíala como parámetro.'}), 400
    if not _validate_periodo(periodo):
        return jsonify({'error': 'Formato de periodo inválido (use YYYY-YYYY)'}), 400

    id_cliente = _get_client_id(clave)
    if id_cliente is None:
        return jsonify({'error': f'Cliente "{clave}" no encontrado en el sistema'}), 404

    # Load valid SKUs — Excel catalog is primary; fall back to Odoo only when Excel is empty
    valid_skus = get_valid_skus()

    # wb/ws ya cargados arriba para leer clave/periodo — reutilizamos directamente

    # Find header row (row with 'SKU' in first column)
    header_row = None
    for r_idx in range(1, 10):
        val = ws.cell(row=r_idx, column=1).value
        if val and str(val).strip().upper() == 'SKU':
            header_row = r_idx
            break

    if header_row is None:
        return jsonify({'error': 'Estructura inválida: no se encontró fila de encabezado con "SKU"'}), 400

    # Map column names → indices
    col_map = {}
    for ci in range(1, ws.max_column + 1):
        h = ws.cell(row=header_row, column=ci).value
        if h:
            col_map[str(h).strip()] = ci

    required_headers = set(CAMPOS_INFO) | set(MESES_LABELS)
    missing = required_headers - set(col_map.keys())
    if missing:
        return jsonify({'error': f'Columnas faltantes en el archivo: {", ".join(sorted(missing))}'}), 400

    errors = []
    rows_to_save = []

    for r_idx in range(header_row + 1, ws.max_row + 1):
        sku = ws.cell(row=r_idx, column=col_map['SKU']).value
        if sku is None or str(sku).strip() == '':
            continue  # skip empty rows

        sku = str(sku).strip()

        # Validate SKU exists in catalog
        if sku not in valid_skus:
            errors.append(f'Fila {r_idx}: SKU "{sku}" no existe en el catálogo de productos')
            continue

        producto  = str(ws.cell(row=r_idx, column=col_map['Producto']).value or '').strip().upper()
        marca     = str(ws.cell(row=r_idx, column=col_map['Marca']).value or '').strip().upper() or 'N/A'
        modelo    = str(ws.cell(row=r_idx, column=col_map['Modelo']).value or '').strip().upper()
        color     = str(ws.cell(row=r_idx, column=col_map['Color']).value or '').strip().upper() or 'N/A'
        talla     = str(ws.cell(row=r_idx, column=col_map['Talla']).value or '').strip().upper() or 'N/A'

        month_values = {}
        month_error = False
        for mes_label, mes_col_name in zip(MESES, MESES_LABELS):
            raw = ws.cell(row=r_idx, column=col_map[mes_col_name]).value
            if raw is None or str(raw).strip() == '':
                raw = 0
            try:
                qty = int(float(str(raw)))
                if qty < 0:
                    errors.append(f'Fila {r_idx}, SKU {sku}: cantidad negativa en {mes_col_name}')
                    month_error = True
                    break
                month_values[mes_label] = qty
            except (ValueError, TypeError):
                errors.append(f'Fila {r_idx}, SKU {sku}: valor no numérico "{raw}" en {mes_col_name}')
                month_error = True
                break

        if month_error:
            continue

        rows_to_save.append({
            'sku':      sku,
            'producto': producto,
            'marca':    marca,
            'modelo':   modelo,
            'color':    color,
            'talla':    talla,
            **month_values,
        })

    if errors and not rows_to_save:
        return jsonify({'errores': errors, 'guardados': 0}), 422

    # Upsert rows
    saved = 0
    conn = obtener_conexion()
    cur = conn.cursor()
    try:
        for row in rows_to_save:
            cur.execute("""
                INSERT INTO forecast_proyecciones
                    (id_cliente, clave_cliente, periodo, sku, producto, marca, modelo,
                     color, talla, mayo, junio, julio, agosto, septiembre, octubre,
                     noviembre, diciembre, enero, febrero, marzo, abril)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE
                    producto   = VALUES(producto),
                    marca      = VALUES(marca),
                    modelo     = VALUES(modelo),
                    color      = VALUES(color),
                    talla      = VALUES(talla),
                    mayo       = VALUES(mayo),
                    junio      = VALUES(junio),
                    julio      = VALUES(julio),
                    agosto     = VALUES(agosto),
                    septiembre = VALUES(septiembre),
                    octubre    = VALUES(octubre),
                    noviembre  = VALUES(noviembre),
                    diciembre  = VALUES(diciembre),
                    enero      = VALUES(enero),
                    febrero    = VALUES(febrero),
                    marzo      = VALUES(marzo),
                    abril      = VALUES(abril),
                    actualizado_en = CURRENT_TIMESTAMP
            """, (
                id_cliente, clave, periodo,
                row['sku'], row['producto'], row['marca'], row['modelo'],
                row['color'], row['talla'],
                row.get('mayo', 0), row.get('junio', 0), row.get('julio', 0),
                row.get('agosto', 0), row.get('septiembre', 0), row.get('octubre', 0),
                row.get('noviembre', 0), row.get('diciembre', 0), row.get('enero', 0),
                row.get('febrero', 0), row.get('marzo', 0), row.get('abril', 0),
            ))
            saved += 1
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'error': f'Error al guardar: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()

    _forecast_cache.pop((clave, periodo), None)
    if _redis:
        try: _redis.delete(_rkey_forecast(clave, periodo))
        except Exception: pass

    result = {'guardados': saved, 'clave_cliente': clave, 'periodo': periodo}
    if errors:
        result['advertencias'] = errors
    return jsonify(result), 200


@forecast_bp.route('/forecast', methods=['GET'])
def listar_forecast():
    """
    GET /forecast?clave=<clave_cliente>&periodo=<periodo>
    Returns forecast rows for a client+period, only for whitelist products.
    Includes precio (client tier price) and nivel_precio per row.
    """
    clave  = request.args.get('clave', '').strip()
    periodo = request.args.get('periodo', '').strip()

    if not clave or not periodo:
        return jsonify({'error': 'Faltan parámetros: clave, periodo'}), 400

    _update_whitelist_skus()

    # ── L1 (memoria) ──────────────────────────────────────────────────────────
    _fc_key = (clave, periodo)
    _fc_hit  = _forecast_cache.get(_fc_key)
    if _fc_hit and (_time.time() - _fc_hit[0]) < _FORECAST_TTL:
        return jsonify(_fc_hit[1]), 200

    # ── L2 (Redis) ────────────────────────────────────────────────────────────
    _r_hit = _redis_get(_rkey_forecast(clave, periodo))
    if _r_hit is not None:
        _forecast_cache[_fc_key] = (_time.time(), _r_hit)
        return jsonify(_r_hit), 200

    # MySQL nivel → TIER_NAMES key
    NIVEL_TO_TIER = {
        'Partner Elite Plus!': 'Partner Elite Plus!',
        'Partner Elite Plus':  'Partner Elite Plus!',
        'Partner Elite':       'Partner Elite',
        'Partner':             'Partner',
        'Distribuidor':        'Distribuidor',
    }

    conn = obtener_conexion()
    cur = conn.cursor(dictionary=True)
    try:
        # Nivel del cliente para seleccionar precio correcto
        cur.execute("SELECT nivel FROM clientes WHERE clave = %s LIMIT 1", (clave,))
        cli = cur.fetchone()
        nivel = (cli or {}).get('nivel') or ''
        tier  = NIVEL_TO_TIER.get(nivel, 'Distribuidor')

        cur.execute("""
            SELECT
                f.id,
                f.sku,
                COALESCE(p.nombre_producto, ep.nombre, f.producto, f.sku) AS producto,
                CASE
                    WHEN COALESCE(p.marca, ep.marca) IS NOT NULL AND COALESCE(p.marca, ep.marca) NOT IN ('N/A', 'BICIS', 'BICICLETAS')
                         THEN COALESCE(p.marca, ep.marca)
                    WHEN f.marca IS NOT NULL AND f.marca NOT IN ('N/A', 'BICIS', 'BICICLETAS') THEN f.marca
                    WHEN UPPER(COALESCE(p.nombre_producto, ep.nombre, f.producto, '')) LIKE '%SCOTT%'   THEN 'SCOTT'
                    WHEN UPPER(COALESCE(p.nombre_producto, ep.nombre, f.producto, '')) LIKE '%MEGAMO%'  THEN 'MEGAMO'
                    WHEN UPPER(COALESCE(p.nombre_producto, ep.nombre, f.producto, '')) LIKE '%SYNCROS%' THEN 'SYNCROS'
                    ELSE COALESCE(p.marca, ep.marca, f.marca, 'N/A')
                END AS marca,
                CASE
                    WHEN COALESCE(p.categoria, '') LIKE '% / %'
                         THEN TRIM(SUBSTRING_INDEX(p.categoria, ' / ', -1))
                    WHEN COALESCE(p.categoria, '') NOT IN ('', 'N/A', 'BICIS', 'BICICLETAS')
                         THEN p.categoria
                    ELSE COALESCE(f.modelo, ep.modelo, '')
                END AS modelo,
                p.categoria AS categ_path,
                COALESCE(p.color,           ep.color,  f.color)           AS color,
                COALESCE(p.talla,           ep.talla,  f.talla)           AS talla,
                COALESCE(f.mayo, 0) AS mayo,
                COALESCE(f.junio, 0) AS junio,
                COALESCE(f.julio, 0) AS julio,
                COALESCE(f.agosto, 0) AS agosto,
                COALESCE(f.septiembre, 0) AS septiembre,
                COALESCE(f.octubre, 0) AS octubre,
                COALESCE(f.noviembre, 0) AS noviembre,
                COALESCE(f.diciembre, 0) AS diciembre,
                COALESCE(f.enero, 0) AS enero,
                COALESCE(f.febrero, 0) AS febrero,
                COALESCE(f.marzo, 0) AS marzo,
                COALESCE(f.abril, 0) AS abril,
                (COALESCE(f.mayo, 0) + COALESCE(f.junio, 0) + COALESCE(f.julio, 0) +
                 COALESCE(f.agosto, 0) + COALESCE(f.septiembre, 0) + COALESCE(f.octubre, 0) +
                 COALESCE(f.noviembre, 0) + COALESCE(f.diciembre, 0) + COALESCE(f.enero, 0) +
                 COALESCE(f.febrero, 0) + COALESCE(f.marzo, 0) + COALESCE(f.abril, 0)) AS total,
                CASE WHEN p.referencia_interna IS NOT NULL THEN 'whitelist' ELSE 'excel' END AS fuente,
                ep.precio_distribuidor,
                ep.precio_partner,
                ep.precio_partner_elite,
                ep.precio_partner_elite_plus,
                ep.precio_publico AS ep_precio_publico,
                f.actualizado_en
            FROM forecast_proyecciones f
            LEFT JOIN odoo_catalogo p          ON p.referencia_interna = f.sku
            LEFT JOIN forecast_excel_productos ep ON ep.sku = f.sku AND p.referencia_interna IS NULL
            WHERE f.clave_cliente = %s AND f.periodo = %s
            ORDER BY fuente, f.actualizado_en DESC, f.sku
        """, (clave, periodo))
        rows = cur.fetchall()

        all_skus = [r['sku'] for r in rows if r.get('sku')]

        # ── Resolver lista de precios del distribuidor en Odoo ───────────────────
        # Prioridad: (1) lista asignada individualmente en res.partner,
        #            (2) lista genérica del tier (para distribuidores sin registro en Odoo)
        partner_pricelist_id   = None
        partner_pricelist_name = ''
        try:
            from utils.odoo_utils import get_odoo_models, ODOO_DB, ODOO_PASSWORD, ODOO_COMPANY_ID
            uid_pl, models_pl, _ = get_odoo_models()
            if uid_pl:
                # 1) Lista asignada individualmente al socio
                partners_pl = models_pl.execute_kw(
                    ODOO_DB, uid_pl, ODOO_PASSWORD,
                    'res.partner', 'search_read',
                    [[('ref', '=', clave), ('company_id', '=', ODOO_COMPANY_ID)]],
                    {'fields': ['property_product_pricelist'], 'limit': 1}
                )
                if partners_pl:
                    pl = partners_pl[0].get('property_product_pricelist')
                    if pl and isinstance(pl, (list, tuple)) and len(pl) >= 2:
                        partner_pricelist_id   = pl[0]
                        partner_pricelist_name = str(pl[1])

                # 2) Si no tiene lista individual, usar la lista genérica del tier
                if not partner_pricelist_id:
                    odoo_pl_name = _TIER_TO_ODOO_PL.get(tier, tier.upper())
                    pl_fallback = models_pl.execute_kw(
                        ODOO_DB, uid_pl, ODOO_PASSWORD,
                        'product.pricelist', 'search_read',
                        [[['name', '=', odoo_pl_name]]],
                        {'fields': ['id', 'name'], 'order': 'id desc', 'limit': 1}
                    )
                    if pl_fallback:
                        partner_pricelist_id   = pl_fallback[0]['id']
                        partner_pricelist_name = pl_fallback[0]['name']
                        logging.info('[forecast] %s no en Odoo; usando lista generica "%s" (id=%s)',
                                     clave, partner_pricelist_name, partner_pricelist_id)
        except Exception as _e:
            logging.warning('[forecast] No se pudo obtener lista de precios de Odoo para %s: %s', clave, _e)

        # PRECIO PUBLICO (sin IVA) solo para bicis del whitelist — lista ID 4 en Odoo
        PRECIO_PUBLICO_PL_ID = 4
        whitelist_skus = [r['sku'] for r in rows if r.get('fuente') == 'whitelist' and r.get('sku')]
        pub_prices: dict = {}
        if whitelist_skus:
            try:
                pub_prices = _get_single_pricelist_prices(PRECIO_PUBLICO_PL_ID, whitelist_skus)
            except Exception as _ep:
                logging.warning('[forecast] No se pudo obtener precio publico: %s', _ep)

        _TIER_TO_EP_COL = {
            'Partner Elite Plus!': 'precio_partner_elite_plus',
            'Partner Elite':       'precio_partner_elite',
            'Partner':             'precio_partner',
            'Distribuidor':        'precio_distribuidor',
        }
        _MEGAMO_TIER_FACTORS = {
            'Partner Elite Plus!': 0.695,
            'Partner Elite':       0.715,
            'Partner':             0.740,
            'Distribuidor':        0.760,
        }

        if partner_pricelist_id and all_skus:
            raw_prices = _get_single_pricelist_prices(partner_pricelist_id, all_skus)
            for r in rows:
                sku = r.get('sku') or ''
                raw = raw_prices.get(sku, 0.0)
                if not raw:
                    cat_p = SKU_CATALOG.get(sku, {}).get('prices', {})
                    raw = cat_p.get(tier, cat_p.get('Distribuidor', 0.0))
                if not raw and r.get('fuente') == 'excel':
                    raw = float(r.get(_TIER_TO_EP_COL.get(tier, 'precio_distribuidor')) or 0)
                if not raw and r.get('marca', '').upper() == 'MEGAMO':
                    lp = pub_prices.get(sku, 0.0)
                    if lp:
                        raw = round(float(lp) * _MEGAMO_TIER_FACTORS.get(tier, 0.760), 2)
                r['precio']       = round(float(raw) * IVA_FACTOR, 2)
                r['nivel_precio'] = partner_pricelist_name or tier
                if pub_prices.get(sku):
                    r['precio_publico'] = round(float(pub_prices[sku]) * IVA_FACTOR, 2)
                elif r.get('fuente') == 'excel' and r.get('ep_precio_publico'):
                    r['precio_publico'] = round(float(r['ep_precio_publico']) * IVA_FACTOR, 2)
                else:
                    r['precio_publico'] = None
                if r.get('actualizado_en'):
                    r['actualizado_en'] = r['actualizado_en'].isoformat()
        else:
            for r in rows:
                sku = r.get('sku') or ''
                cat_p = SKU_CATALOG.get(sku, {}).get('prices', {})
                raw   = cat_p.get(tier, cat_p.get('Distribuidor', 0.0))
                if not raw and r.get('fuente') == 'excel':
                    raw = float(r.get(_TIER_TO_EP_COL.get(tier, 'precio_distribuidor')) or 0)
                if not raw and r.get('marca', '').upper() == 'MEGAMO':
                    lp = pub_prices.get(sku, 0.0)
                    if lp:
                        raw = round(float(lp) * _MEGAMO_TIER_FACTORS.get(tier, 0.760), 2)
                r['precio']       = round(float(raw) * IVA_FACTOR, 2)
                r['nivel_precio'] = tier
                if pub_prices.get(sku):
                    r['precio_publico'] = round(float(pub_prices[sku]) * IVA_FACTOR, 2)
                elif r.get('fuente') == 'excel' and r.get('ep_precio_publico'):
                    r['precio_publico'] = round(float(r['ep_precio_publico']) * IVA_FACTOR, 2)
                else:
                    r['precio_publico'] = None
                if r.get('actualizado_en'):
                    r['actualizado_en'] = r['actualizado_en'].isoformat()

        _forecast_cache[_fc_key] = (_time.time(), rows)
        _redis_set(_rkey_forecast(clave, periodo), rows, _FORECAST_R_TTL)
        return jsonify(rows), 200
    except Exception as e:
        logging.exception('[forecast] listar_forecast error: %s', e)
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


@forecast_bp.route('/forecast/precios-catalogo', methods=['GET'])
def precios_catalogo():
    """
    GET /forecast/precios-catalogo?clave=X
    Retorna {precios: {sku: {precio, precio_publico}}, nivel_precio: str}
    para TODOS los SKUs de la whitelist en un solo batch de Odoo.
    Llamado al abrir el tab de proyecciones para hacer lookups instantáneos al agregar productos.
    """
    clave = request.args.get('clave', '').strip()

    NIVEL_TO_TIER = {
        'Partner Elite Plus!': 'Partner Elite Plus!',
        'Partner Elite Plus':  'Partner Elite Plus!',
        'Partner Elite':       'Partner Elite',
        'Partner':             'Partner',
        'Distribuidor':        'Distribuidor',
    }

    tier          = 'Distribuidor'
    whitelist_skus = []
    try:
        conn = obtener_conexion()
        cur  = conn.cursor(dictionary=True)
        cur.execute("SELECT sku FROM forecast_sku_whitelist")
        whitelist_skus = [r['sku'] for r in cur.fetchall()]
        if clave:
            cur.execute("SELECT nivel FROM clientes WHERE clave = %s LIMIT 1", (clave,))
            cli   = cur.fetchone()
            nivel = (cli or {}).get('nivel') or ''
            tier  = NIVEL_TO_TIER.get(nivel, 'Distribuidor')
        cur.close()
        conn.close()
    except Exception as e:
        logging.warning('[precios-catalogo] MySQL error: %s', e)

    precios      = {}
    nivel_precio = tier

    if not whitelist_skus:
        return jsonify({'precios': {}, 'nivel_precio': nivel_precio}), 200

    # Precio público (lista 4) — un solo batch de Odoo para todos los SKUs
    try:
        pub_prices = _get_single_pricelist_prices(4, whitelist_skus)
        for sku, raw in pub_prices.items():
            if raw:
                precios.setdefault(sku, {})['precio_publico'] = round(raw * IVA_FACTOR, 2)
    except Exception as e:
        logging.warning('[precios-catalogo] precio_publico error: %s', e)

    # Precio distribuidor — un solo batch de Odoo para todos los SKUs
    try:
        from utils.odoo_utils import get_odoo_models, ODOO_DB, ODOO_PASSWORD, ODOO_COMPANY_ID
        uid_pl, models_pl, _ = get_odoo_models()
        partner_pricelist_id   = None
        partner_pricelist_name = ''
        if uid_pl and clave:
            partners_pl = models_pl.execute_kw(
                ODOO_DB, uid_pl, ODOO_PASSWORD,
                'res.partner', 'search_read',
                [[('ref', '=', clave), ('company_id', '=', ODOO_COMPANY_ID)]],
                {'fields': ['property_product_pricelist'], 'limit': 1}
            )
            if partners_pl:
                pl = partners_pl[0].get('property_product_pricelist')
                if pl and isinstance(pl, (list, tuple)) and len(pl) >= 2:
                    partner_pricelist_id   = pl[0]
                    partner_pricelist_name = str(pl[1])
            if not partner_pricelist_id:
                odoo_pl_name = _TIER_TO_ODOO_PL.get(tier, tier.upper())
                pl_fallback  = models_pl.execute_kw(
                    ODOO_DB, uid_pl, ODOO_PASSWORD,
                    'product.pricelist', 'search_read',
                    [[['name', '=', odoo_pl_name]]],
                    {'fields': ['id', 'name'], 'order': 'id desc', 'limit': 1}
                )
                if pl_fallback:
                    partner_pricelist_id   = pl_fallback[0]['id']
                    partner_pricelist_name = pl_fallback[0]['name']

        if partner_pricelist_id:
            dist_prices = _get_single_pricelist_prices(partner_pricelist_id, whitelist_skus)
            for sku, raw in dist_prices.items():
                if raw:
                    precios.setdefault(sku, {})['precio'] = round(raw * IVA_FACTOR, 2)
            nivel_precio = partner_pricelist_name or tier
        else:
            for sku in whitelist_skus:
                cat_p = SKU_CATALOG.get(sku, {}).get('prices', {})
                raw   = cat_p.get(tier, cat_p.get('Distribuidor', 0.0))
                if raw:
                    precios.setdefault(sku, {})['precio'] = round(raw * IVA_FACTOR, 2)
            nivel_precio = tier
    except Exception as e:
        logging.warning('[precios-catalogo] precio distribuidor error: %s', e)

    # Incluir precios de productos apparel (forecast_excel_productos) segun tier
    _ep_tier_col = {
        'Partner Elite Plus!': 'precio_partner_elite_plus',
        'Partner Elite':       'precio_partner_elite',
        'Partner':             'precio_partner',
        'Distribuidor':        'precio_distribuidor',
    }
    ep_col = _ep_tier_col.get(tier, 'precio_distribuidor')
    try:
        conn_ep = obtener_conexion()
        cur_ep  = conn_ep.cursor(dictionary=True)
        cur_ep.execute(f"""
            SELECT sku, {ep_col} AS precio_dist, precio_publico
            FROM forecast_excel_productos
            WHERE origen = 'excel'
              AND {ep_col} IS NOT NULL
        """)
        for ep_row in cur_ep.fetchall():
            ep_sku = ep_row['sku']
            if ep_sku not in precios:
                precios[ep_sku] = {}
            if ep_row.get('precio_dist'):
                precios[ep_sku]['precio'] = round(float(ep_row['precio_dist']) * IVA_FACTOR, 2)
            if ep_row.get('precio_publico'):
                precios[ep_sku]['precio_publico'] = round(float(ep_row['precio_publico']) * IVA_FACTOR, 2)
        cur_ep.close()
        conn_ep.close()
    except Exception as _eep:
        logging.warning('[precios-catalogo] Excel apparel prices error: %s', _eep)

    return jsonify({'precios': precios, 'nivel_precio': nivel_precio}), 200


@forecast_bp.route('/forecast/distribuidores-precios', methods=['GET'])
def distribuidores_precios():
    """
    GET /forecast/distribuidores-precios
    Devuelve todos los clientes con su nivel en MySQL y su lista de precios en Odoo.
    Permite detectar discrepancias entre ambas fuentes.
    """
    conn = obtener_conexion()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT id, clave, nombre_cliente, nivel, zona
            FROM clientes
            ORDER BY nombre_cliente
        """)
        clientes = cur.fetchall()
    except Exception as e:
        logging.exception('[distribuidores_precios] MySQL error: %s', e)
        return jsonify({'error': 'Error al obtener clientes'}), 500
    finally:
        cur.close()
        conn.close()

    if not clientes:
        return jsonify([]), 200

    claves = [c['clave'] for c in clientes if c.get('clave')]
    odoo_pricelists: dict = {}

    try:
        from utils.odoo_utils import get_odoo_models, ODOO_DB, ODOO_PASSWORD, ODOO_COMPANY_ID
        uid, models, err = get_odoo_models()
        if uid:
            partners = models.execute_kw(
                ODOO_DB, uid, ODOO_PASSWORD,
                'res.partner', 'search_read',
                [[('ref', 'in', claves), ('company_id', '=', ODOO_COMPANY_ID)]],
                {'fields': ['id', 'name', 'ref', 'property_product_pricelist']}
            )
            for p in partners:
                ref = (p.get('ref') or '').strip()
                pricelist = p.get('property_product_pricelist')
                if ref and pricelist:
                    odoo_pricelists[ref] = pricelist[1] if isinstance(pricelist, (list, tuple)) else str(pricelist)
        else:
            logging.warning('[distribuidores_precios] No Odoo connection: %s', err)
    except Exception as e:
        logging.exception('[distribuidores_precios] Odoo error: %s', e)

    result = []
    NIVEL_TO_TIER = {
        'Partner Elite Plus!': 'Partner Elite Plus!',
        'Partner Elite Plus':  'Partner Elite Plus!',
        'Partner Elite':       'Partner Elite',
        'Partner':             'Partner',
        'Distribuidor':        'Distribuidor',
    }
    for c in clientes:
        clave    = (c.get('clave') or '').strip()
        odoo_pl  = odoo_pricelists.get(clave)
        nivel    = c.get('nivel') or ''
        tier     = NIVEL_TO_TIER.get(nivel, None)
        coincide = (odoo_pl == nivel or odoo_pl == tier) if odoo_pl else None
        result.append({
            'id':                 c['id'],
            'clave':              clave,
            'nombre':             c.get('nombre_cliente'),
            'zona':               c.get('zona'),
            'nivel_mysql':        nivel,
            'lista_precios_odoo': odoo_pl,
            'coincide':           coincide,
        })

    return jsonify(result), 200


@forecast_bp.route('/forecast/periodos', methods=['GET'])
def listar_periodos():
    """
    GET /forecast/periodos?clave=<clave_cliente>
    Returns distinct periods available for a client.
    """
    clave = request.args.get('clave', '').strip()
    if not clave:
        return jsonify({'error': 'Falta parámetro clave'}), 400

    conn = obtener_conexion()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT DISTINCT periodo
            FROM forecast_proyecciones
            WHERE clave_cliente = %s
            ORDER BY periodo DESC
        """, (clave,))
        periodos = [r['periodo'] for r in cur.fetchall()]
        return jsonify(periodos), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


@forecast_bp.route('/forecast/avance', methods=['GET'])
def avance_forecast():
    """
    GET /forecast/avance?clave=<clave_cliente>&periodo=<periodo>

    Cross-reference: forecast quantities vs actual orders in the monitor table.
    Products are matched by normalizing SKU (stripping hyphens/spaces, uppercase),
    so clave_factura format (290189010) matches referencia_interna (290189-010).

    Returns a list of rows with:
      forecast_total, pedido_total, restante, pct_cubierto, estados (dict)
    """
    clave   = request.args.get('clave', '').strip()
    periodo = request.args.get('periodo', '').strip()
    refresh = request.args.get('refresh', '0') == '1'

    if not clave or not periodo:
        return jsonify({'error': 'Faltan parámetros: clave, periodo'}), 400
    if not _validate_periodo(periodo):
        return jsonify({'error': 'Formato de periodo inválido'}), 400

    if refresh:
        _avance_cache.pop((clave, periodo), None)
        try:
            import redis as _rl, os as _os
            _rl.Redis(host=_os.getenv('REDIS_HOST', 'localhost'),
                      port=int(_os.getenv('REDIS_PORT', 6379)), db=0).delete(_rkey_avance(clave, periodo))
        except Exception:
            pass

    m = re.match(r'^(\d{4})-(\d{4})$', periodo)
    year1, year2 = int(m.group(1)), int(m.group(2))
    # El periodo comercial empieza en mayo del year1 (columna "mayo" del forecast)
    # y termina en junio del year2. Usar julio era demasiado tardío y excluía
    # órdenes anticipadas colocadas en mayo-junio del mismo año (ej. S07027 del 2026-06-04).
    fecha_inicio = f'{year1}-05-01'
    fecha_fin    = f'{year2}-06-30'

    def _norm(s: str) -> str:
        """Remove hyphens/spaces, uppercase — for fuzzy SKU matching."""
        return re.sub(r'[\-\s]', '', str(s or '')).upper()

    conn = obtener_conexion()
    cur = conn.cursor(dictionary=True)
    try:
        # 1. Forecast rows for this client/period
        cur.execute("""
            SELECT id, sku, producto, marca, modelo, color, talla,
                   (mayo+junio+julio+agosto+septiembre+octubre+
                    noviembre+diciembre+enero+febrero+marzo+abril) AS forecast_total
            FROM forecast_proyecciones
            WHERE clave_cliente = %s AND periodo = %s
            ORDER BY marca, modelo, sku
        """, (clave, periodo))
        forecast_rows = cur.fetchall()

        if not forecast_rows:
            return jsonify([]), 200

        # 2. Build normalized SKU → canonical SKU map
        norm_to_sku: dict = {}
        for fr in forecast_rows:
            n = _norm(fr['sku'])
            if n:
                norm_to_sku[n] = fr['sku']

        # 3. Query Odoo sale.order.line — L1 (mem) → L2 (Redis) → Odoo
        orders_by_sku: dict = {}
        _cache_key = (clave, periodo)
        _cached = _avance_cache.get(_cache_key)
        if _cached and (_time.time() - _cached[0]) < _AVANCE_TTL:
            orders_by_sku = _cached[1]
            logging.debug('avance_forecast: caché L1 HIT para %s/%s', clave, periodo)
        elif (_r_avance := _redis_get(_rkey_avance(clave, periodo))) is not None:
            orders_by_sku = _r_avance
            _avance_cache[_cache_key] = (_time.time(), orders_by_sku)
            logging.debug('avance_forecast: caché L2 (Redis) HIT para %s/%s', clave, periodo)
        else:
          try:
            from utils.odoo_utils import get_odoo_models, ODOO_DB, ODOO_PASSWORD
            uid_oo, models_oo, err_oo = get_odoo_models()
            if uid_oo and models_oo:
                # Find partner(s) matching the client reference code
                partner_ids = models_oo.execute_kw(
                    ODOO_DB, uid_oo, ODOO_PASSWORD,
                    'res.partner', 'search',
                    [[['ref', '=', clave]]]
                )
                if partner_ids:
                    # Get confirmed sale orders in the commercial period
                    order_ids = models_oo.execute_kw(
                        ODOO_DB, uid_oo, ODOO_PASSWORD,
                        'sale.order', 'search',
                        [[['partner_id', 'in', partner_ids],
                          ['state', 'in', ['sale', 'done']],
                          ['date_order', '>=', fecha_inicio],
                          ['date_order', '<=', fecha_fin + ' 23:59:59']]]
                    )
                    if order_ids:
                        # Read order lines
                        sol = models_oo.execute_kw(
                            ODOO_DB, uid_oo, ODOO_PASSWORD,
                            'sale.order.line', 'search_read',
                            [[['order_id', 'in', order_ids],
                              ['state', 'not in', ['cancel']]]],
                            {'fields': ['product_id', 'product_uom_qty',
                                        'order_id'], 'limit': 0}
                        )
                        # Batch-load default_codes for all referenced products
                        prod_ids = list({l['product_id'][0]
                                         for l in sol if l.get('product_id')})
                        prods = models_oo.execute_kw(
                            ODOO_DB, uid_oo, ODOO_PASSWORD,
                            'product.product', 'search_read',
                            [[['id', 'in', prod_ids]]],
                            {'fields': ['id', 'default_code'], 'limit': 0}
                        )
                        prod_code_map = {
                            p['id']: (p.get('default_code') or '').strip() or f'ODOO:{p["id"]}'
                            for p in prods
                        }
                        # Aggregate quantities per forecast SKU
                        for l in sol:
                            pid = l['product_id'][0] if l.get('product_id') else None
                            dc  = prod_code_map.get(pid, '')
                            matched = norm_to_sku.get(_norm(dc))
                            if matched is None:
                                continue
                            qty = int(l.get('product_uom_qty') or 0)
                            if matched not in orders_by_sku:
                                orders_by_sku[matched] = {'pedido_total': 0,
                                                          'estados': {}}
                            orders_by_sku[matched]['pedido_total'] += qty
                            orders_by_sku[matched]['estados']['Orden Confirmada'] = (
                                orders_by_sku[matched]['estados'].get(
                                    'Orden Confirmada', 0) + qty
                            )
            else:
                logging.warning('avance_forecast: no se pudo conectar a Odoo – %s', err_oo)
            # Guardar en L1 + L2 (aunque esté vacío, para no repetir en fallo)
            _avance_cache[_cache_key] = (_time.time(), orders_by_sku)
            _redis_set(_rkey_avance(clave, periodo), orders_by_sku, _AVANCE_R_TTL)
          except Exception as _ex_oo:
            logging.exception('avance_forecast: error al consultar Odoo: %s', _ex_oo)

        # 4. Build result merging forecast + orders
        result = []
        for fr in forecast_rows:
            sku           = fr['sku']
            ord_data      = orders_by_sku.get(sku, {'pedido_total': 0, 'estados': {}})
            forecast_total = int(fr['forecast_total'] or 0)
            pedido_total   = ord_data['pedido_total']
            restante       = max(0, forecast_total - pedido_total)
            pct            = (round(pedido_total / forecast_total * 1000) / 10
                               if forecast_total > 0 else 0)
            result.append({
                'id':            fr['id'],
                'sku':           sku,
                'producto':      fr['producto'],
                'marca':         fr['marca'],
                'modelo':        fr['modelo'],
                'color':         fr['color'],
                'talla':         fr['talla'],
                'forecast_total': forecast_total,
                'pedido_total':   pedido_total,
                'restante':       restante,
                'pct_cubierto':   pct,
                'estados':        ord_data['estados'],
            })

        return jsonify(result), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


@forecast_bp.route('/precalentar-forecast', methods=['POST'])
def precalentar_forecast_route():
    """Dispara el precalentamiento Redis de /forecast y /forecast/avance para todos los clientes."""
    total = iniciar_precalentamiento_forecast()
    return jsonify({'status': 'iniciado', 'pares': total}), 202


@forecast_bp.route('/forecast/periodos/integral', methods=['GET'])
def listar_periodos_integral():
    """
    GET /forecast/periodos/integral?grupo_id=<id>
    Periodos con datos guardados para cualquier cliente del grupo.
    """
    grupo_id = request.args.get('grupo_id', '').strip()
    if not grupo_id:
        return jsonify({'error': 'Falta parámetro grupo_id'}), 400

    conn = obtener_conexion()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT DISTINCT fp.periodo
            FROM forecast_proyecciones fp
            INNER JOIN clientes c ON c.clave = fp.clave_cliente
            WHERE c.id_grupo = %s
            ORDER BY fp.periodo DESC
        """, (grupo_id,))
        periodos = [r['periodo'] for r in cur.fetchall()]
        return jsonify(periodos), 200
    except Exception as e:
        logging.exception('[forecast/periodos/integral] error: %s', e)
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


@forecast_bp.route('/forecast/integral', methods=['GET'])
def listar_forecast_integral():
    """
    GET /forecast/integral?grupo_id=<id>&periodo=<periodo>
    Proyecciones agregadas (suma) de todos los clientes del grupo, por SKU.
    """
    grupo_id = request.args.get('grupo_id', '').strip()
    periodo  = request.args.get('periodo', '').strip()
    if not grupo_id or not periodo:
        return jsonify({'error': 'Faltan parámetros: grupo_id, periodo'}), 400

    _update_whitelist_skus()

    conn = obtener_conexion()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT clave FROM clientes WHERE id_grupo = %s", (grupo_id,))
        claves = [r['clave'] for r in cur.fetchall()]
        if not claves:
            return jsonify([]), 200

        ph = ','.join(['%s'] * len(claves))
        cur.execute(f"""
            SELECT
                p.referencia_interna AS sku,
                p.nombre_producto    AS producto,
                p.marca,
                p.categoria          AS modelo,
                p.color,
                p.talla,
                COALESCE(SUM(f.mayo),       0) AS mayo,
                COALESCE(SUM(f.junio),      0) AS junio,
                COALESCE(SUM(f.julio),      0) AS julio,
                COALESCE(SUM(f.agosto),     0) AS agosto,
                COALESCE(SUM(f.septiembre), 0) AS septiembre,
                COALESCE(SUM(f.octubre),    0) AS octubre,
                COALESCE(SUM(f.noviembre),  0) AS noviembre,
                COALESCE(SUM(f.diciembre),  0) AS diciembre,
                COALESCE(SUM(f.enero),      0) AS enero,
                COALESCE(SUM(f.febrero),    0) AS febrero,
                COALESCE(SUM(f.marzo),      0) AS marzo,
                COALESCE(SUM(f.abril),      0) AS abril,
                (COALESCE(SUM(f.mayo),0)+COALESCE(SUM(f.junio),0)+COALESCE(SUM(f.julio),0)+
                 COALESCE(SUM(f.agosto),0)+COALESCE(SUM(f.septiembre),0)+COALESCE(SUM(f.octubre),0)+
                 COALESCE(SUM(f.noviembre),0)+COALESCE(SUM(f.diciembre),0)+COALESCE(SUM(f.enero),0)+
                 COALESCE(SUM(f.febrero),0)+COALESCE(SUM(f.marzo),0)+COALESCE(SUM(f.abril),0)) AS total
            FROM odoo_catalogo p
            LEFT JOIN forecast_proyecciones f
                ON  f.sku = p.referencia_interna
                AND f.clave_cliente IN ({ph})
                AND f.periodo = %s
            WHERE p.referencia_interna IN (SELECT sku FROM forecast_sku_whitelist)
            GROUP BY p.referencia_interna, p.nombre_producto, p.marca, p.categoria, p.color, p.talla
            ORDER BY p.marca, p.categoria, p.referencia_interna
        """, (*claves, periodo))
        rows = cur.fetchall()

        all_skus = [r['sku'] for r in rows if r.get('sku')]
        prices = _get_odoo_prices_for_skus(all_skus) if all_skus else {}
        for sku in all_skus:
            cat_entry  = SKU_CATALOG.get(sku, {})
            odoo_entry = prices.get(sku, {})
            for key in ['list_price'] + TIER_NAMES:
                if not odoo_entry.get(key):
                    odoo_entry[key] = cat_entry.get('prices', {}).get(key, 0.0)
            prices[sku] = odoo_entry

        for r in rows:
            p = prices.get(r.get('sku') or '', {})
            r['precio']       = round(p.get('Distribuidor', 0.0) * IVA_FACTOR, 2)
            r['nivel_precio'] = 'Grupo'

        return jsonify(rows), 200
    except Exception as e:
        logging.exception('[forecast/integral] error: %s', e)
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


@forecast_bp.route('/forecast/avance/integral', methods=['GET'])
def avance_forecast_integral():
    """
    GET /forecast/avance/integral?grupo_id=<id>&periodo=<periodo>
    Avance (forecast vs pedidos Odoo) agregado para todos los clientes del grupo.
    """
    grupo_id = request.args.get('grupo_id', '').strip()
    periodo  = request.args.get('periodo', '').strip()
    refresh  = request.args.get('refresh', '0') == '1'
    if not grupo_id or not periodo:
        return jsonify({'error': 'Faltan parámetros: grupo_id, periodo'}), 400
    if not _validate_periodo(periodo):
        return jsonify({'error': 'Formato de periodo inválido'}), 400

    if refresh:
        _avance_cache.pop(('integral', grupo_id, periodo), None)
        try:
            import redis as _rl, os as _os
            _rl.Redis(host=_os.getenv('REDIS_HOST', 'localhost'),
                      port=int(_os.getenv('REDIS_PORT', 6379)), db=0).delete(
                _rkey_avance(f'integral:{grupo_id}', periodo))
        except Exception:
            pass

    m = re.match(r'^(\d{4})-(\d{4})$', periodo)
    year1, year2 = int(m.group(1)), int(m.group(2))
    # Mismo ajuste que avance_forecast: usar mayo como inicio del periodo comercial
    fecha_inicio = f'{year1}-05-01'
    fecha_fin    = f'{year2}-06-30'

    def _norm(s: str) -> str:
        return re.sub(r'[\-\s]', '', str(s or '')).upper()

    conn = obtener_conexion()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT clave FROM clientes WHERE id_grupo = %s", (grupo_id,))
        claves = [r['clave'] for r in cur.fetchall()]
        if not claves:
            return jsonify([]), 200

        ph = ','.join(['%s'] * len(claves))
        cur.execute(f"""
            SELECT id, sku, producto, marca, modelo, color, talla,
                   (mayo+junio+julio+agosto+septiembre+octubre+
                    noviembre+diciembre+enero+febrero+marzo+abril) AS forecast_total
            FROM forecast_proyecciones
            WHERE clave_cliente IN ({ph}) AND periodo = %s
            ORDER BY marca, modelo, sku
        """, (*claves, periodo))
        rows_raw = cur.fetchall()

        if not rows_raw:
            return jsonify([]), 200

        # Agregamos el forecast total por SKU (un cliente puede tener el mismo SKU que otro)
        agg: dict = {}
        for fr in rows_raw:
            sku = fr['sku']
            if sku not in agg:
                agg[sku] = {**fr, 'forecast_total': 0}
            agg[sku]['forecast_total'] += int(fr['forecast_total'] or 0)
        forecast_rows = list(agg.values())

        norm_to_sku: dict = {}
        for fr in forecast_rows:
            n = _norm(fr['sku'])
            if n:
                norm_to_sku[n] = fr['sku']

        # Consulta Odoo para TODOS los partners del grupo en una sola pasada
        orders_by_sku: dict = {}
        _cache_key = ('integral', grupo_id, periodo)
        _cached = _avance_cache.get(_cache_key)
        if _cached and (_time.time() - _cached[0]) < _AVANCE_TTL:
            orders_by_sku = _cached[1]
        else:
            try:
                from utils.odoo_utils import get_odoo_models, ODOO_DB, ODOO_PASSWORD
                uid_oo, models_oo, err_oo = get_odoo_models()
                if uid_oo and models_oo:
                    # Buscar todos los partners del grupo por su ref
                    partner_ids = []
                    for clave in claves:
                        pids = models_oo.execute_kw(
                            ODOO_DB, uid_oo, ODOO_PASSWORD,
                            'res.partner', 'search',
                            [[['ref', '=', clave]]]
                        )
                        partner_ids.extend(pids)

                    if partner_ids:
                        order_ids = models_oo.execute_kw(
                            ODOO_DB, uid_oo, ODOO_PASSWORD,
                            'sale.order', 'search',
                            [[['partner_id', 'in', partner_ids],
                              ['state', 'in', ['sale', 'done']],
                              ['date_order', '>=', fecha_inicio],
                              ['date_order', '<=', fecha_fin + ' 23:59:59']]]
                        )
                        if order_ids:
                            sol = models_oo.execute_kw(
                                ODOO_DB, uid_oo, ODOO_PASSWORD,
                                'sale.order.line', 'search_read',
                                [[['order_id', 'in', order_ids],
                                  ['state', 'not in', ['cancel']]]],
                                {'fields': ['product_id', 'product_uom_qty'], 'limit': 0}
                            )
                            prod_ids = list({l['product_id'][0] for l in sol if l.get('product_id')})
                            prods = models_oo.execute_kw(
                                ODOO_DB, uid_oo, ODOO_PASSWORD,
                                'product.product', 'search_read',
                                [[['id', 'in', prod_ids]]],
                                {'fields': ['id', 'default_code'], 'limit': 0}
                            )
                            prod_code_map = {
                                p['id']: (p.get('default_code') or '').strip() or f'ODOO:{p["id"]}'
                                for p in prods
                            }
                            for l in sol:
                                pid     = l['product_id'][0] if l.get('product_id') else None
                                dc      = prod_code_map.get(pid, '')
                                matched = norm_to_sku.get(_norm(dc))
                                if matched is None:
                                    continue
                                qty = int(l.get('product_uom_qty') or 0)
                                if matched not in orders_by_sku:
                                    orders_by_sku[matched] = {'pedido_total': 0, 'estados': {}}
                                orders_by_sku[matched]['pedido_total'] += qty
                                orders_by_sku[matched]['estados']['Orden Confirmada'] = (
                                    orders_by_sku[matched]['estados'].get('Orden Confirmada', 0) + qty
                                )
                else:
                    logging.warning('avance_forecast_integral: no se pudo conectar a Odoo – %s', err_oo)
                _avance_cache[_cache_key] = (_time.time(), orders_by_sku)
            except Exception as _ex:
                logging.exception('avance_forecast_integral: error Odoo: %s', _ex)

        result = []
        for fr in forecast_rows:
            sku            = fr['sku']
            ord_data       = orders_by_sku.get(sku, {'pedido_total': 0, 'estados': {}})
            forecast_total = int(fr['forecast_total'] or 0)
            pedido_total   = ord_data['pedido_total']
            restante       = max(0, forecast_total - pedido_total)
            pct            = (round(pedido_total / forecast_total * 1000) / 10
                               if forecast_total > 0 else 0)
            result.append({
                'id':             fr['id'],
                'sku':            sku,
                'producto':       fr['producto'],
                'marca':          fr['marca'],
                'modelo':         fr['modelo'],
                'color':          fr['color'],
                'talla':          fr['talla'],
                'forecast_total': forecast_total,
                'pedido_total':   pedido_total,
                'restante':       restante,
                'pct_cubierto':   pct,
                'estados':        ord_data['estados'],
            })

        return jsonify(result), 200
    except Exception as e:
        logging.exception('[forecast/avance/integral] error: %s', e)
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


@forecast_bp.route('/forecast/sync-catalogo', methods=['POST'])
def sync_catalogo():
    """
    POST /forecast/sync-catalogo
    Manually trigger a re-sync of the Odoo product catalog into odoo_catalogo.
    Accepts optional JSON body: {"force": true} to re-sync even if already populated.
    """
    body  = request.get_json(silent=True) or {}
    force = bool(body.get('force', True))
    result = _trigger_catalogo_sync(force=force)
    return jsonify({'status': result}), 200


@forecast_bp.route('/forecast/sync-catalogo', methods=['GET'])
def sync_catalogo_status():
    """GET /forecast/sync-catalogo — returns catalog row count and sync status."""
    conn = obtener_conexion()
    cur  = conn.cursor()
    cur.execute('SELECT COUNT(*) FROM odoo_catalogo')
    cnt = cur.fetchone()[0]
    cur.close()
    conn.close()
    return jsonify({
        'total_productos': cnt,
        'syncing': _catalogo_syncing,
    }), 200


@forecast_bp.route('/forecast/debug-megamo-grupos', methods=['GET'])
def debug_megamo_grupos():
    """GET /forecast/debug-megamo-grupos — lista todos los grupos MEGAMO visibles en búsqueda."""
    conn = _safe_obtener_conexion()
    if conn is None:
        return jsonify({'error': 'DB unavailable'}), 503
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT DISTINCT oc.nombre_producto AS nombre
            FROM odoo_catalogo oc
            INNER JOIN forecast_sku_whitelist wl ON wl.sku = oc.referencia_interna
            WHERE oc.nombre_producto LIKE '%megamo%' OR oc.marca LIKE '%megamo%'
            ORDER BY oc.nombre_producto
        """)
        grupos = [r[0] for r in cur.fetchall()]

        cur.execute("""
            SELECT wl.sku
            FROM forecast_sku_whitelist wl
            WHERE wl.sku LIKE 'MH%'
              AND wl.sku NOT IN (SELECT referencia_interna FROM odoo_catalogo)
            ORDER BY wl.sku
        """)
        missing = [r[0] for r in cur.fetchall()]

        cur.execute("SELECT COUNT(*) FROM odoo_catalogo WHERE nombre_producto LIKE '%megamo%' OR marca LIKE '%megamo%'")
        total_megamo = cur.fetchone()[0]

        return jsonify({
            'grupos_visibles': len(grupos),
            'grupos': grupos,
            'skus_whitelist_sin_catalogo': missing,
            'total_megamo_en_catalogo': total_megamo,
        }), 200
    finally:
        cur.close()
        conn.close()


@forecast_bp.route('/forecast/<int:fid>', methods=['PUT'])
def actualizar_forecast(fid):
    """
    PUT /forecast/<id>
    Body: {mayo, junio, ..., abril}
    Updates monthly quantities for a single row.
    """
    data = request.get_json(force=True, silent=True) or {}

    updates = {}
    for mes in MESES:
        if mes in data:
            try:
                qty = int(data[mes])
                if qty < 0:
                    return jsonify({'error': f'Cantidad negativa para {mes}'}), 400
                updates[mes] = qty
            except (ValueError, TypeError):
                return jsonify({'error': f'Valor inválido para {mes}'}), 400

    if not updates:
        # Also allow full row update (adding new product line)
        new_data = {}
        for field in ['producto', 'marca', 'modelo', 'color', 'talla']:
            if field in data:
                new_data[field] = str(data[field])[:255]
        for mes in MESES:
            new_data[mes] = int(data.get(mes, 0))
        updates = new_data

    if not updates:
        return jsonify({'error': 'Sin campos para actualizar'}), 400

    set_clause = ', '.join(f'{k} = %s' for k in updates)
    values = list(updates.values()) + [fid]

    conn = obtener_conexion()
    cur = conn.cursor()
    try:
        cur.execute(f"UPDATE forecast_proyecciones SET {set_clause} WHERE id = %s", values)
        conn.commit()
        if cur.rowcount == 0:
            return jsonify({'error': 'Registro no encontrado'}), 404
        return jsonify({'ok': True}), 200
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


_SEARCH_PAGE       = 500   # fallback general — suficiente para mostrar todos sin paginación
_SEARCH_PAGE_WL    = 2000  # whitelist: todos los SKUs Megamo+Scott en una sola página

# Alias map for Spanish plural/variant forms that don't match DB names
_SEARCH_ALIAS = {
    'zapatillas': 'zapatos',
    'zapatilla':  'zapatos',
    'tenis':      'zapatos',
    'calzado':    'zapatos',
    'calzados':   'zapatos',
    'anforas':    'anfora',
    'anfora':     'anfora',
    'anforitas':  'anfora',
    'bidones':    'bidon',
    'llantas':    'llanta',
    'luces':      'luz',
    'frenos':     'freno',
    'pedales':    'pedal',
    'guantes':    'guante',
    'rodilleras': 'rodillera',
    'coderas':    'codera',
    'gafas':      'gafa',
    'lentes':     'lente',
    'mochilas':   'mochila',
    'bolsos':     'bolso',
    'gorras':     'gorra',
    'cascos':     'casco',
}


def _normalizar_query(q: str) -> str:
    """Return a normalized search term that better matches DB naming conventions."""
    lower = q.lower()
    if lower in _SEARCH_ALIAS:
        return _SEARCH_ALIAS[lower]
    # Strip common Spanish plural 's' for words longer than 5 chars
    if lower.endswith('s') and len(lower) > 5:
        return q[:-1]
    return q


def _tokens_busqueda(q: str) -> list:
    """Divide la query en tokens individuales y normaliza cada uno.
    Permite búsquedas como 'MEGAMO PULSE' → ['MEGAMO', 'PULS'] (AND entre tokens).
    """
    tokens = [t.strip() for t in q.split() if len(t.strip()) >= 2]
    return [_normalizar_query(t) for t in tokens] or [_normalizar_query(q)]


@forecast_bp.route('/forecast/catalogo-excel', methods=['POST'])
def cargar_catalogo_excel():
    """
    POST /forecast/catalogo-excel  (multipart/form-data)
    Field: file (xlsx/xls) con columnas: SKU, NOMBRE, [COLOR], [TALLA]
    Carga los productos en forecast_excel_productos para usarse como catálogo
    de validación y búsqueda en lugar de Odoo.
    """
    if not OPENPYXL_OK:
        return jsonify({'error': 'openpyxl no instalado en el servidor'}), 500

    archivo = request.files.get('file')
    if not archivo:
        return jsonify({'error': 'Falta el archivo (field: file)'}), 400

    ext = archivo.filename.rsplit('.', 1)[-1].lower() if archivo.filename else ''
    if ext not in ('xlsx', 'xls'):
        return jsonify({'error': 'El archivo debe ser Excel (.xlsx o .xls)'}), 400

    content = archivo.read()
    result = load_excel_products(content)

    if not result['success']:
        return jsonify({'error': result.get('message', 'Error al procesar el archivo')}), 422

    return jsonify({
        'cargados':                 result['cargados'],
        'total_filas_procesadas':   result['total_filas_procesadas'],
        'duplicados_actualizados':  result['duplicados_actualizados'],
        'advertencias':             result.get('errores', []),
    }), 200


@forecast_bp.route('/forecast/importar-csv-apparel', methods=['POST'])
def importar_csv_apparel():
    """
    POST /forecast/importar-csv-apparel  (multipart/form-data)
    Field: file (.csv) con columnas:
      MARCA, GENERIC, COLOUR CODE, SIZE CODE, VARIANT,
      GENERIC DESCRIPTION (ignorado), MODEL, COLOUR DESCRIPTION, SIZE DESCRIPTION

    Importa masivamente productos de apparel (ropa) que no existen en Odoo.
    Los SKUs quedan disponibles inmediatamente en buscar-producto y en forecast.
    """
    archivo = request.files.get('file')
    if not archivo:
        return jsonify({'error': 'Falta el archivo (field: file)'}), 400

    ext = (archivo.filename or '').rsplit('.', 1)[-1].lower()
    if ext != 'csv':
        return jsonify({'error': 'El archivo debe ser CSV (.csv)'}), 400

    content = archivo.read()

    # Intentar UTF-8 con BOM primero; si falla, latin-1
    for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
        result = load_csv_apparel_products(content, encoding=enc)
        if result.get('success') or result.get('total_filas_procesadas', 0) > 0:
            break
        if 'Columnas requeridas' in result.get('message', ''):
            break  # error de estructura, no de encoding

    if not result['success']:
        return jsonify({'error': result.get('message', 'Error al procesar el CSV')}), 422

    return jsonify({
        'cargados':                result['cargados'],
        'total_filas_procesadas':  result['total_filas_procesadas'],
        'duplicados_actualizados': result['duplicados_actualizados'],
        'advertencias':            result.get('errores', []),
    }), 200


@forecast_bp.route('/forecast/catalogo-excel', methods=['GET'])
def estado_catalogo_excel():
    """GET /forecast/catalogo-excel — total de productos cargados desde Excel."""
    conn = obtener_conexion()
    cur  = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM forecast_excel_productos WHERE origen = 'excel'")
    cnt = cur.fetchone()[0]
    cur.close()
    conn.close()
    return jsonify({'total_productos': cnt}), 200


@forecast_bp.route('/forecast/catalogo-excel/lista', methods=['GET'])
def listar_catalogo_excel():
    """
    GET /forecast/catalogo-excel/lista?q=<search>&limit=<int>&offset=<int>
    Paginates products from the Excel catalog.
    """
    q      = request.args.get('q', '').strip()
    try:
        limit  = min(int(request.args.get('limit', 50)), 500)
        offset = max(int(request.args.get('offset', 0)), 0)
    except (ValueError, TypeError):
        limit, offset = 50, 0

    result = list_excel_products(search=q, limit=limit, offset=offset)
    for p in result['productos']:
        if p.get('cargado_en'):
            p['cargado_en'] = p['cargado_en'].isoformat() if hasattr(p['cargado_en'], 'isoformat') else str(p['cargado_en'])
        if p.get('actualizado_en'):
            p['actualizado_en'] = p['actualizado_en'].isoformat() if hasattr(p['actualizado_en'], 'isoformat') else str(p['actualizado_en'])
    return jsonify(result), 200


@forecast_bp.route('/forecast/catalogo-excel', methods=['DELETE'])
def limpiar_catalogo_excel():
    """DELETE /forecast/catalogo-excel — elimina todos los productos del catálogo Excel."""
    result = clear_excel_catalog()
    if 'message' in result and 'Error' in result.get('message', ''):
        return jsonify({'error': result['message']}), 500
    return jsonify(result), 200


def _normalizar_talla(talla: str) -> str:
    """Strip wheel-size suffix ' 29"' from talla (e.g. 'S 29"' -> 'S').
    Keeps '27.5"' variants intact (e.g. 'S 27.5"', 'XS 27.5"')."""
    t = (talla or '').strip().upper()
    if t.endswith(' 29"'):
        t = t[:-4].strip()
    return t or 'N/A'


@forecast_bp.route('/forecast/buscar-producto', methods=['GET'])
def buscar_producto():
    """
    GET /forecast/buscar-producto?q=<query>&offset=<int>
    Searches forecast_excel_productos (Excel catalog, primary source).
    Always refreshes the SKU whitelist from the configured fixed list before searching.
    Falls back to odoo_catalogo when Excel catalog is empty,
    and to monitor table if neither is populated.
    Returns {results: [...], has_more: bool, offset: int} with up to 50 items per page.
    """
    _update_whitelist_skus()
    _ensure_scott_names()
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify({'results': [], 'has_more': False, 'offset': 0}), 200

    try:
        offset = max(0, int(request.args.get('offset', 0)))
    except (ValueError, TypeError):
        offset = 0

    tokens = _tokens_busqueda(q)
    # Para ORDER BY: el primer token como referencia de coincidencia exacta de SKU
    q_first = tokens[0]

    conn = obtener_conexion()
    cur = conn.cursor(dictionary=True)
    try:
        # ── Whitelist activa: buscar SOLO entre los SKUs configurados ──────────
        cur.execute("SELECT COUNT(*) as cnt FROM forecast_sku_whitelist")
        use_whitelist = cur.fetchone()['cnt'] > 0

        if use_whitelist:
            # Condición AND por cada token — incluye wl.sku para SKUs sin odoo_catalogo
            inner_conds  = []
            inner_params = []
            for tok in tokens:
                lk = f'%{tok.upper()}%'
                inner_conds.append(
                    "(UPPER(wl.sku) LIKE %s "
                    "OR UPPER(COALESCE(oc.nombre_producto,'')) LIKE %s "
                    "OR UPPER(COALESCE(oc.referencia_interna,'')) LIKE %s "
                    "OR UPPER(COALESCE(oc.marca,'')) LIKE %s "
                    "OR UPPER(COALESCE(oc.color,'')) LIKE %s "
                    "OR UPPER(COALESCE(oc.talla,'')) LIKE %s)"
                )
                inner_params.extend([lk, lk, lk, lk, lk, lk])

            inner_where = ' AND '.join(inner_conds)

            # LEFT JOIN: incluye SKUs en whitelist aunque no estén en odoo_catalogo
            cur.execute(f"""
                SELECT sku, nombre_src, categoria, marca, odoo_color, odoo_talla
                FROM (
                    SELECT
                        wl.sku                                                          AS sku,
                        COALESCE(MIN(oc.nombre_producto), wl.sku)                       AS nombre_src,
                        MIN(oc.categoria)                                               AS categoria,
                        COALESCE(MIN(oc.marca), '')                                     AS marca,
                        COALESCE(NULLIF(TRIM(oc.color), ''), 'N/A')                    AS odoo_color,
                        COALESCE(NULLIF(TRIM(oc.talla), ''), 'N/A')                    AS odoo_talla
                    FROM forecast_sku_whitelist wl
                    LEFT JOIN odoo_catalogo oc ON oc.referencia_interna = wl.sku
                    WHERE {inner_where}
                    GROUP BY wl.sku,
                             COALESCE(NULLIF(TRIM(oc.color), ''), 'N/A'),
                             COALESCE(NULLIF(TRIM(oc.talla), ''), 'N/A')
                ) AS deduped
                ORDER BY
                    CASE WHEN sku = %s THEN 0 ELSE 1 END,
                    nombre_src
                LIMIT %s OFFSET %s
            """, (*inner_params, q_first, _SEARCH_PAGE_WL + 1, offset))
            rows     = cur.fetchall()
            has_more = len(rows) > _SEARCH_PAGE_WL
            rows     = rows[:_SEARCH_PAGE_WL]

            _SEARCH_COLOR_WORDS = {
                'NEGRO','NEGRA','BLANCO','BLANCA','ROJO','ROJA','AZUL','VERDE',
                'NARANJA','AMARILLO','AMARILLA','GRIS','DORADO','DORADA',
                'PLATEADO','PLATEADA','PETROLEO','PETRÓLEO','MORADO','MORADA',
                'ROSA','CAFE','BEIGE','MARRON','BRONCE','ORO','PLATA','COBRE',
                'TURQUESA','SALMON','ARENA','CREMA','LILA','CELESTE',
            }

            results = []
            for r in rows:
                sku_r  = r['sku'] or ''
                cat    = r.get('categoria') or ''
                modelo = cat.split(' / ')[-1].strip().upper() if ' / ' in cat else ''
                color  = (r.get('odoo_color') or '').strip().upper() or 'N/A'
                talla  = _normalizar_talla(r.get('odoo_talla') or '')

                # Normalizar nombre: quitar sufijos en paréntesis y espacios extras
                raw_n  = (r.get('nombre_src') or sku_r)
                raw_n  = re.sub(r'\s*\([^)]+\)\s*$', '', raw_n).strip()
                nombre = ' '.join(raw_n.split()).upper()

                # Usar nombre correcto (es_MX) para SKUs Scott si está en cache
                if sku_r in _SCOTT_CORRECT_NAMES:
                    nombre = _SCOTT_CORRECT_NAMES[sku_r]

                # Usar color y talla de Odoo si odoo_catalogo no los tiene
                if color == 'N/A' and sku_r in _SCOTT_COLORS:
                    color = _SCOTT_COLORS[sku_r]
                if (not talla or talla == 'N/A') and sku_r in _SCOTT_TALLAS:
                    talla = _normalizar_talla(_SCOTT_TALLAS[sku_r])

                # Asegurar que el nombre empiece con el prefijo numérico del SKU
                sku_prefix = sku_r[:6]
                if sku_prefix.isdigit() and not nombre.startswith(sku_prefix):
                    nombre = f"{sku_prefix} {nombre}"

                # Marca: si odoo_catalogo dice algo incorrecto (BICIS, etc.) pero
                # el nombre contiene una marca conocida, usarla
                marca_r = (r.get('marca') or '').upper()
                if sku_r in SKU_CATALOG or 'SCOTT' in nombre:
                    marca_r = 'SCOTT'
                elif 'MEGAMO' in nombre:
                    marca_r = 'MEGAMO'
                elif 'SYNCROS' in nombre:
                    marca_r = 'SYNCROS'

                # Si el modelo está vacío, derivarlo del nombre quitando
                # "XXXXXX BICICLETA MARCA " del inicio
                if not modelo and nombre:
                    m_name = re.sub(r'^\d{6}\s+', '', nombre)
                    m_name = re.sub(r'^BICICLETA\s+', '', m_name)
                    m_name = re.sub(r'^(SCOTT|MEGAMO|SYNCROS)\s+', '', m_name)
                    # Quitar MY27/MY26 y lo que sigue
                    m_name = re.sub(r'\s+MY\d{2}.*$', '', m_name).strip()
                    # Quitar palabra de color al final
                    m_parts = m_name.split()
                    if m_parts and m_parts[-1] in _SEARCH_COLOR_WORDS:
                        m_parts = m_parts[:-1]
                    modelo = ' '.join(m_parts)

                results.append({
                    'sku':      sku_r,
                    'producto': nombre,
                    'marca':    marca_r or 'N/A',
                    'modelo':   modelo,
                    'color':    color,
                    'talla':    talla,
                    'fuente':   'whitelist',
                    'label':    f"{sku_r} — {nombre}",
                })

            # También buscar en el catálogo Excel (apparel, etc.) aunque la whitelist esté activa
            fb_conds2  = ' AND '.join(["(UPPER(sku) LIKE %s OR UPPER(nombre) LIKE %s OR UPPER(COALESCE(marca,'')) LIKE %s)"] * len(tokens))
            fb_params2 = []
            for tok in tokens:
                lk2 = f'%{tok.upper()}%'
                fb_params2.extend([lk2, lk2, lk2])
            cur.execute(f"""
                SELECT sku, nombre AS nombre_src, color AS odoo_color, talla AS odoo_talla, marca, modelo
                FROM forecast_excel_productos
                WHERE origen = 'excel'
                  AND {fb_conds2}
                ORDER BY CASE WHEN sku = %s THEN 0 ELSE 1 END, nombre
                LIMIT %s OFFSET %s
            """, (*fb_params2, q_first, _SEARCH_PAGE + 1, offset))
            excel_rows = cur.fetchall()
            has_more = has_more or len(excel_rows) > _SEARCH_PAGE
            excel_rows = excel_rows[:_SEARCH_PAGE]
            seen_skus = {r['sku'] for r in results}
            for r in excel_rows:
                sku_ex    = r.get('sku') or ''
                if sku_ex in seen_skus:
                    continue
                seen_skus.add(sku_ex)
                nombre_ex = ' '.join((r.get('nombre_src') or '').split()).upper()
                color_ex  = (r.get('odoo_color') or '').strip().upper() or 'N/A'
                talla_ex  = _normalizar_talla(r.get('odoo_talla') or '')
                marca_ex  = (r.get('marca')  or '').strip().upper() or 'N/A'
                modelo_ex = (r.get('modelo') or '').strip().upper()
                results.append({
                    'sku':      sku_ex,
                    'producto': nombre_ex,
                    'marca':    marca_ex,
                    'modelo':   modelo_ex,
                    'color':    color_ex,
                    'talla':    talla_ex,
                    'fuente':   'excel',
                    'label':    f"{sku_ex} — {nombre_ex}",
                })
            return jsonify({'results': results, 'has_more': has_more, 'offset': offset}), 200

        # ── Sin whitelist: cadena de fallback original ────────────────────────
        cur.execute("SELECT COUNT(*) as cnt FROM forecast_excel_productos WHERE origen = 'excel'")
        use_excel = cur.fetchone()['cnt'] > 0

        if use_excel:
            # Fallback Excel: también multi-token AND
            fb_conds  = ' AND '.join(["(UPPER(sku) LIKE %s OR UPPER(nombre) LIKE %s OR UPPER(COALESCE(marca,'')) LIKE %s)"] * len(tokens))
            fb_params = []
            for tok in tokens:
                lk = f'%{tok.upper()}%'
                fb_params.extend([lk, lk, lk])
            cur.execute(f"""
                SELECT sku, nombre AS nombre_src, color AS odoo_color, talla AS odoo_talla, marca, modelo
                FROM forecast_excel_productos
                WHERE origen = 'excel'
                  AND {fb_conds}
                ORDER BY
                    CASE WHEN sku = %s THEN 0 ELSE 1 END,
                    nombre
                LIMIT %s OFFSET %s
            """, (*fb_params, q_first, _SEARCH_PAGE + 1, offset))
            rows     = cur.fetchall()
            has_more = len(rows) > _SEARCH_PAGE
            rows     = rows[:_SEARCH_PAGE]

            results = []
            for r in rows:
                color  = (r.get('odoo_color') or '').strip().upper() or 'N/A'
                talla  = _normalizar_talla(r.get('odoo_talla') or '')
                nombre = (r.get('nombre_src') or '').strip().upper()
                marca  = (r.get('marca')  or '').strip().upper() or 'N/A'
                modelo = (r.get('modelo') or '').strip().upper()
                results.append({
                    'sku':      r['sku'] or '',
                    'producto': nombre,
                    'marca':    marca,
                    'modelo':   modelo,
                    'color':    color,
                    'talla':    talla,
                    'fuente':   'excel',
                    'label':    f"{r['sku']} — {nombre}",
                })
            return jsonify({'results': results, 'has_more': has_more, 'offset': offset}), 200

        # Fallback: odoo_catalogo completo
        cur.execute('SELECT COUNT(*) as cnt FROM odoo_catalogo')
        use_catalogo = cur.fetchone()['cnt'] > 0

        # Fallback catalogo/monitor: multi-token AND
        cat_conds  = []
        cat_params = []
        for tok in tokens:
            lk = f'%{tok.upper()}%'
            cat_conds.append(
                "(UPPER(oc.nombre_producto) LIKE %s OR UPPER(oc.referencia_interna) LIKE %s "
                "OR UPPER(oc.marca) LIKE %s OR UPPER(oc.categoria) LIKE %s "
                "OR UPPER(pv.descripcion) LIKE %s OR UPPER(pv.modelo) LIKE %s OR UPPER(pv.clave_factura) LIKE %s)"
            )
            cat_params.extend([lk, lk, lk, lk, lk, lk, lk])

        mon_conds  = []
        mon_params = []
        for tok in tokens:
            lk = f'%{tok.upper()}%'
            mon_conds.append(
                "(UPPER(m.nombre_producto) LIKE %s OR UPPER(m.referencia_interna) LIKE %s "
                "OR UPPER(m.marca) LIKE %s OR UPPER(pv.descripcion) LIKE %s "
                "OR UPPER(pv.modelo) LIKE %s OR UPPER(pv.clave_factura) LIKE %s)"
            )
            mon_params.extend([lk, lk, lk, lk, lk, lk])

        if use_catalogo:
            cur.execute(f"""
                SELECT
                    oc.referencia_interna                                       AS sku,
                    oc.nombre_producto                                          AS nombre_src,
                    oc.categoria                                                AS categoria,
                    oc.marca                                                    AS marca,
                    oc.color                                                    AS odoo_color,
                    oc.talla                                                    AS odoo_talla,
                    pv.descripcion                                              AS descripcion_pv,
                    pv.modelo                                                   AS modelo_pv
                FROM odoo_catalogo oc
                LEFT JOIN proyecciones_ventas pv ON pv.clave_odoo = oc.referencia_interna
                WHERE {' AND '.join(cat_conds)}
                ORDER BY
                    CASE WHEN oc.referencia_interna = %s THEN 0 ELSE 1 END,
                    oc.nombre_producto
                LIMIT %s OFFSET %s
            """, (*cat_params, q_first, _SEARCH_PAGE + 1, offset))
        else:
            # Last fallback: monitor table (only invoiced products)
            cur.execute(f"""
                SELECT
                    m.referencia_interna                                            AS sku,
                    ANY_VALUE(m.nombre_producto)                                    AS nombre_src,
                    ANY_VALUE(m.categoria_producto)                                 AS categoria,
                    ANY_VALUE(m.marca)                                              AS marca,
                    ANY_VALUE(pv.descripcion)                                       AS descripcion_pv,
                    ANY_VALUE(pv.modelo)                                            AS modelo_pv
                FROM monitor m
                LEFT JOIN proyecciones_ventas pv ON pv.clave_odoo = m.referencia_interna
                WHERE m.referencia_interna IS NOT NULL
                  AND m.referencia_interna != ''
                  AND {' AND '.join(mon_conds)}
                GROUP BY m.referencia_interna
                ORDER BY
                    CASE WHEN m.referencia_interna = %s THEN 0 ELSE 1 END,
                    ANY_VALUE(m.nombre_producto)
                LIMIT %s OFFSET %s
            """, (*mon_params, q_first, _SEARCH_PAGE + 1, offset))

        rows = cur.fetchall()
        has_more = len(rows) > _SEARCH_PAGE
        rows = rows[:_SEARCH_PAGE]

        results = []
        for r in rows:
            odoo_color = (r.get('odoo_color') or '').strip().upper()
            odoo_talla = _normalizar_talla(r.get('odoo_talla') or '')

            if not odoo_color and r.get('nombre_src'):
                _nombre_raw = re.sub(r'^\d{5,}\s+', '', (r['nombre_src'] or '').upper().strip())
                _paren = re.search(r'\(([^)]+)\)\s*$', _nombre_raw)
                if _paren:
                    odoo_color = _paren.group(1).strip()

            if odoo_color:
                color = odoo_color
                talla = odoo_talla
                if r.get('descripcion_pv'):
                    _, talla_pv = _parse_color_talla(r['descripcion_pv'], r.get('modelo_pv') or '')
                    if not talla:
                        talla = talla_pv
                    producto = _clean_producto(r['descripcion_pv'], color, talla_pv).upper()
                    modelo   = (r.get('modelo_pv') or '').upper()
                else:
                    raw = re.sub(r'^\d{5,}\s+', '', (r['nombre_src'] or '').upper().strip())
                    brand_up = (r.get('marca') or '').upper().strip()
                    if brand_up:
                        raw = re.sub(r'\b' + re.escape(brand_up) + r'\b\s*', '', raw).strip()
                    raw = re.sub(r'\s*\([^)]*\)\s*$', '', raw).strip()
                    categoria   = (r.get('categoria') or '')
                    modelo_hint = categoria.split(' / ')[-1].strip().upper() if ' / ' in categoria else ''
                    producto = _clean_producto(raw, color, talla)
                    modelo   = modelo_hint
            elif r.get('descripcion_pv'):
                color, talla = _parse_color_talla(r['descripcion_pv'], r.get('modelo_pv') or '')
                producto = _clean_producto(r['descripcion_pv'], color, talla).upper()
                modelo   = (r.get('modelo_pv') or '').upper()
            else:
                raw = re.sub(r'^\d{5,}\s+', '', (r['nombre_src'] or '').upper().strip())
                brand_up = (r.get('marca') or '').upper().strip()
                if brand_up:
                    raw = re.sub(r'\b' + re.escape(brand_up) + r'\b\s*', '', raw).strip()
                categoria   = (r.get('categoria') or '')
                modelo_hint = categoria.split(' / ')[-1].strip().upper() if ' / ' in categoria else ''
                paren_match = re.search(r'\(([^)]+)\)\s*$', raw)
                if paren_match:
                    color = paren_match.group(1).strip()
                    clean_raw = raw[:paren_match.start()].strip()
                    talla = ''
                    producto = _clean_producto(clean_raw, '', '')
                else:
                    color, talla = _parse_color_talla(raw, modelo_hint)
                    producto = _clean_producto(raw, color, talla)
                modelo   = modelo_hint

            results.append({
                'sku':      r['sku'] or '',
                'producto': producto,
                'marca':    (r.get('marca') or '').upper() or 'N/A',
                'modelo':   modelo,
                'color':    color.upper() or 'N/A',
                'talla':    talla.upper() or 'N/A',
                'label':    f"{r['sku']} — {producto}",
            })
        return jsonify({'results': results, 'has_more': has_more, 'offset': offset}), 200
    finally:
        cur.close()
        conn.close()


@forecast_bp.route('/forecast/<int:fid>', methods=['DELETE'])
def eliminar_forecast(fid):
    """DELETE /forecast/<id>"""
    conn = obtener_conexion()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT clave_cliente, periodo FROM forecast_proyecciones WHERE id = %s", (fid,))
        reg = cur.fetchone()
        cur.execute("DELETE FROM forecast_proyecciones WHERE id = %s", (fid,))
        conn.commit()
        if cur.rowcount == 0:
            return jsonify({'error': 'Registro no encontrado'}), 404
        if reg:
            _forecast_cache.pop((reg['clave_cliente'], reg['periodo']), None)
        return jsonify({'ok': True}), 200
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


@forecast_bp.route('/forecast/guardar', methods=['POST'])
def guardar_forecast():
    """
    POST /forecast/guardar
    Body: {clave_cliente, id_cliente, periodo, rows: [{sku, producto, marca, modelo,
           color, talla, mayo, ..., abril}]}
    Batch upsert – used from edit mode in frontend.
    """
    data = request.get_json(force=True, silent=True) or {}
    clave      = str(data.get('clave_cliente', '')).strip()
    id_cliente = data.get('id_cliente')
    periodo    = str(data.get('periodo', '')).strip()
    rows       = data.get('rows', [])

    # Si el frontend no envió id_cliente (vista usuario), derivarlo de la clave
    if not id_cliente and clave:
        id_cliente = _get_client_id(clave)

    if not clave or not id_cliente or not periodo:
        return jsonify({'error': 'Faltan campos: clave_cliente, id_cliente, periodo'}), 400
    if not _validate_periodo(periodo):
        return jsonify({'error': 'Formato de periodo inválido'}), 400
    if not isinstance(rows, list) or len(rows) == 0:
        return jsonify({'error': 'rows debe ser una lista no vacía'}), 400

    # Validate all SKUs exist in forecast_excel_productos (priority) + odoo_catalogo, monitor, proyecciones_ventas (fallback)
    # Aceptamos SKUs que existan en cualquiera de estas fuentes
    valid_skus = get_valid_skus()

    errors = []
    valid_rows = []
    for i, row in enumerate(rows):
        sku = str(row.get('sku', '')).strip()
        if not sku:
            errors.append(f'Fila {i+1}: SKU vacío')
            continue
        
        # Validar cantidades primero
        month_vals = {}
        month_valid = True
        for mes in MESES:
            try:
                qty = int(row.get(mes, 0))
                if qty < 0:
                    raise ValueError()
                month_vals[mes] = qty
            except (ValueError, TypeError):
                errors.append(f'Fila {i+1}, SKU {sku}: cantidad inválida para {mes}')
                month_valid = False
                break
        
        if not month_valid:
            continue
        
        # Validar SKU: aceptar si existe en cualquier tabla fuente O si tiene metadata de Odoo
        has_requiredMetadata = (
            row.get('producto', '').strip() and  # producto no vacío
            row.get('marca', '').strip() and       # marca no vacío
            row.get('modelo', '').strip() and      # modelo no vacío
            row.get('color', '').strip() and       # color no vacío
            row.get('talla', '').strip()           # talla no vacío
        )
        
        if sku not in valid_skus and not has_requiredMetadata:
            errors.append(f'Fila {i+1}: SKU "{sku}" no existe en el catálogo. Selecciona un producto válido desde el modal de búsqueda.')
            continue
        
        valid_rows.append({
            'sku':      sku,
            'producto': str(row.get('producto', '')).strip().upper()[:255],
            'marca':    (str(row.get('marca', '')).strip().upper() or 'N/A')[:100],
            'modelo':   str(row.get('modelo', '')).strip().upper()[:100],
            'color':    (str(row.get('color', '')).strip().upper() or 'N/A')[:100],
            'talla':    (str(row.get('talla', '')).strip().upper() or 'N/A')[:50],
            **month_vals,
        })

    if errors and not valid_rows:
        return jsonify({'errores': errors, 'guardados': 0}), 422

    saved = 0
    conn = obtener_conexion()
    cur = conn.cursor()
    try:
        for row in valid_rows:
            cur.execute("""
                INSERT INTO forecast_proyecciones
                    (id_cliente, clave_cliente, periodo, sku, producto, marca, modelo,
                     color, talla, mayo, junio, julio, agosto, septiembre, octubre,
                     noviembre, diciembre, enero, febrero, marzo, abril)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE
                    producto   = VALUES(producto),
                    marca      = VALUES(marca),
                    modelo     = VALUES(modelo),
                    color      = VALUES(color),
                    talla      = VALUES(talla),
                    mayo       = VALUES(mayo),
                    junio      = VALUES(junio),
                    julio      = VALUES(julio),
                    agosto     = VALUES(agosto),
                    septiembre = VALUES(septiembre),
                    octubre    = VALUES(octubre),
                    noviembre  = VALUES(noviembre),
                    diciembre  = VALUES(diciembre),
                    enero      = VALUES(enero),
                    febrero    = VALUES(febrero),
                    marzo      = VALUES(marzo),
                    abril      = VALUES(abril),
                    actualizado_en = CURRENT_TIMESTAMP
            """, (
                int(id_cliente), clave, periodo,
                row['sku'], row['producto'], row['marca'], row['modelo'],
                row['color'], row['talla'],
                row['mayo'], row['junio'], row['julio'], row['agosto'],
                row['septiembre'], row['octubre'], row['noviembre'], row['diciembre'],
                row['enero'], row['febrero'], row['marzo'], row['abril'],
            ))
            saved += 1
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'error': f'Error al guardar: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()

    # Invalidar caché para que la próxima carga refleje los cambios
    _forecast_cache.pop((clave, periodo), None)

    result = {'guardados': saved}
    if errors:
        result['advertencias'] = errors
    return jsonify(result), 200


# ─────────────────────────────────────────────────────
# SKU Whitelist management
# ─────────────────────────────────────────────────────

@forecast_bp.route('/forecast/sku-whitelist', methods=['GET'])
def listar_sku_whitelist():
    """GET /forecast/sku-whitelist — lista SKUs y productos del catálogo de proyecciones."""
    conn = _safe_obtener_conexion()
    skus = []
    if conn:
        cur = conn.cursor(dictionary=True)
        try:
            cur.execute("SELECT sku FROM forecast_sku_whitelist ORDER BY sku")
            skus = [r['sku'] for r in cur.fetchall()]
        finally:
            cur.close()
            conn.close()
    products = _get_whitelist_products() if skus else []
    return jsonify({'total': len(skus), 'skus': skus, 'productos': products}), 200


@forecast_bp.route('/forecast/sku-whitelist', methods=['POST'])
def set_sku_whitelist():
    """
    POST /forecast/sku-whitelist
    Body: {"skus": ["REF1", "REF2", ...], "replace": true}
    Reemplaza (o añade) SKUs al whitelist de proyecciones.
    """
    data    = request.get_json(force=True, silent=True) or {}
    skus    = data.get('skus', [])
    replace = bool(data.get('replace', True))

    if not isinstance(skus, list):
        return jsonify({'error': 'skus debe ser una lista'}), 400

    cleaned = [str(s).strip() for s in skus if str(s).strip()]
    if not cleaned:
        return jsonify({'error': 'Lista de SKUs vacía'}), 400

    conn = obtener_conexion()
    cur  = conn.cursor()
    try:
        if replace:
            cur.execute("DELETE FROM forecast_sku_whitelist")
        cur.executemany(
            "INSERT IGNORE INTO forecast_sku_whitelist (sku) VALUES (%s)",
            [(s,) for s in cleaned]
        )
        conn.commit()

        placeholders = ','.join(['%s'] * len(cleaned))
        cur.execute(
            f"SELECT COUNT(*) FROM odoo_catalogo WHERE referencia_interna IN ({placeholders})",
            cleaned
        )
        en_odoo   = cur.fetchone()[0]
        faltantes = len(cleaned) - en_odoo

        return jsonify({
            'cargados':         len(cleaned),
            'en_odoo_catalogo': en_odoo,
            'advertencia': (
                f'{faltantes} SKU(s) no encontrados en el catálogo Odoo local. '
                'Ejecute POST /forecast/sync-catalogo para sincronizar primero.'
            ) if faltantes > 0 else None,
        }), 200
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


@forecast_bp.route('/forecast/sku-whitelist', methods=['DELETE'])
def limpiar_sku_whitelist():
    """DELETE /forecast/sku-whitelist — vacía el whitelist de proyecciones."""
    conn = obtener_conexion()
    cur  = conn.cursor()
    try:
        cur.execute("DELETE FROM forecast_sku_whitelist")
        count = cur.rowcount
        conn.commit()
        return jsonify({'eliminados': count}), 200
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


@forecast_bp.route('/forecast/resumen-articulos', methods=['GET'])
def resumen_articulos():
    """
    GET /forecast/resumen-articulos

    Vista consolidada de los 92 artículos MY27:
    - Totales por mes sumando TODOS los distribuidores
    - Desglose de cuánto aportó cada distribuidor por mes
    - Total anual por artículo
    - Disponibilidad por mes
    - KPIs globales

    Query params:
      periodo (str) — ej. "2026-2027". Sin valor = todos los periodos.
    """
    periodo = request.args.get('periodo', '').strip()

    MESES = ['mayo', 'junio', 'julio', 'agosto', 'septiembre',
             'octubre', 'noviembre', 'diciembre', 'enero', 'febrero', 'marzo', 'abril']
    MESES_LABELS = ['May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic', 'Ene', 'Feb', 'Mar', 'Abr']

    conn = obtener_conexion()
    cur  = conn.cursor(dictionary=True)

    try:
        where  = "WHERE periodo = %s" if periodo else ""
        params = (periodo,) if periodo else ()

        # ── 1. Totales por SKU (suma de todos los distribuidores) ─────────────
        cur.execute(f"""
            SELECT
                sku,
                MAX(producto)   AS producto,
                MAX(marca)      AS marca,
                MAX(modelo)     AS modelo,
                MAX(color)      AS color,
                MAX(talla)      AS talla,
                SUM(mayo)       AS mayo,
                SUM(junio)      AS junio,
                SUM(julio)      AS julio,
                SUM(agosto)     AS agosto,
                SUM(septiembre) AS septiembre,
                SUM(octubre)    AS octubre,
                SUM(noviembre)  AS noviembre,
                SUM(diciembre)  AS diciembre,
                SUM(enero)      AS enero,
                SUM(febrero)    AS febrero,
                SUM(marzo)      AS marzo,
                SUM(abril)      AS abril,
                SUM(mayo+junio+julio+agosto+septiembre+
                    octubre+noviembre+diciembre+
                    enero+febrero+marzo+abril) AS total_anual,
                COUNT(DISTINCT clave_cliente)  AS num_distribuidores
            FROM forecast_proyecciones
            {where}
            GROUP BY sku
        """, params)
        totales_map = {r['sku']: r for r in cur.fetchall()}

        # ── 2. Desglose por SKU + distribuidor ────────────────────────────────
        cur.execute(f"""
            SELECT
                sku, clave_cliente,
                mayo, junio, julio, agosto, septiembre,
                octubre, noviembre, diciembre,
                enero, febrero, marzo, abril,
                (mayo+junio+julio+agosto+septiembre+
                 octubre+noviembre+diciembre+
                 enero+febrero+marzo+abril) AS total_dist
            FROM forecast_proyecciones
            {where}
            ORDER BY sku, clave_cliente
        """, params)
        filas_dist = cur.fetchall()

        # Agrupa desglose por SKU
        desglose_map: dict = {}
        for fd in filas_dist:
            s = fd['sku']
            if s not in desglose_map:
                desglose_map[s] = []
            desglose_map[s].append({
                'clave_cliente': fd['clave_cliente'],
                'total':         int(fd['total_dist'] or 0),
                'meses': {
                    mes: int(fd[mes] or 0) for mes in MESES
                }
            })

        # ── 3. Construir respuesta con los 92 artículos del catálogo ─────────
        articulos = []
        for sku in FORECAST_SKU_WHITELIST:
            cat_info  = SKU_CATALOG.get(sku, {})
            avail_map = cat_info.get('avail', {})
            precios   = cat_info.get('prices', {})
            t         = totales_map.get(sku)

            meses_data = {}
            for mes in MESES:
                meses_data[mes] = {
                    'cantidad':   int(t[mes] or 0) if t else 0,
                    'disponible': avail_map.get(mes, True),
                }

            total_anual = int(t['total_anual'] or 0) if t else 0

            articulos.append({
                'sku':               sku,
                'producto':          (t['producto'] or '') if t else '',
                'marca':             (t['marca']    or '') if t else '',
                'modelo':            (t['modelo']   or '') if t else '',
                'color':             (t['color']    or '') if t else '',
                'talla':             (t['talla']    or '') if t else '',
                'precio_dist':       float(precios.get('Distribuidor', 0)),
                'num_distribuidores': int(t['num_distribuidores'] or 0) if t else 0,
                'total_anual':       total_anual,
                'meses':             meses_data,
                'desglose':          desglose_map.get(sku, []),
            })

        # ── 4. Totales de columna (suma de los 92 artículos por mes) ─────────
        totales_mes = {
            mes: sum(a['meses'][mes]['cantidad'] for a in articulos)
            for mes in MESES
        }
        total_general = sum(totales_mes.values())

        # ── 5. KPIs globales ──────────────────────────────────────────────────
        distribuidores_activos = set()
        for a in articulos:
            for d in a['desglose']:
                if d['total'] > 0:
                    distribuidores_activos.add(d['clave_cliente'])

        return jsonify({
            'articulos':              articulos,
            'totales_mes':            totales_mes,
            'total_general':          total_general,
            'meses':                  MESES,
            'meses_labels':           MESES_LABELS,
            'periodo':                periodo or '2026-2027',
            'kpis': {
                'total_articulos':        len(articulos),
                'articulos_con_pedido':   sum(1 for a in articulos if a['total_anual'] > 0),
                'articulos_sin_pedido':   sum(1 for a in articulos if a['total_anual'] == 0),
                'total_unidades':         total_general,
                'distribuidores_activos': len(distribuidores_activos),
            },
        }), 200

    except Exception as e:
        logging.exception('[forecast] resumen-articulos error: %s', e)
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()
