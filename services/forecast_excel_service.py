"""
Servicio de gestión de catálogo de productos desde Excel.

Proporciona funcionalidad de:
- Carga de productos desde archivos Excel
- Validación de SKUs contra el catálogo Excel
- Administración (listar, eliminar, limpiar)
"""

from db_conexion import obtener_conexion
import logging
import io
import csv
import re

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ─────────────────────────────────────────────────────
# Initialization
# ─────────────────────────────────────────────────────

def ensure_excel_producto_table():
    """Create forecast_excel_productos table if it doesn't exist."""
    conn = obtener_conexion()
    cur = conn.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS forecast_excel_productos (
                sku VARCHAR(100) NOT NULL PRIMARY KEY,
                nombre VARCHAR(400) NOT NULL,
                color VARCHAR(150) DEFAULT NULL,
                talla VARCHAR(100) DEFAULT NULL,
                marca VARCHAR(255) DEFAULT NULL,
                modelo VARCHAR(255) DEFAULT NULL,
                categoria VARCHAR(255) DEFAULT NULL,
                precio_distribuidor DECIMAL(10,2) DEFAULT NULL,
                precio_partner DECIMAL(10,2) DEFAULT NULL,
                precio_partner_elite DECIMAL(10,2) DEFAULT NULL,
                precio_partner_elite_plus DECIMAL(10,2) DEFAULT NULL,
                precio_publico DECIMAL(10,2) DEFAULT NULL,
                origen ENUM('excel', 'odoo') DEFAULT 'excel',
                cargado_en DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                actualizado_en DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_origen (origen),
                FULLTEXT idx_ft_nombre (nombre)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)
        # Migraciones: agregar columnas si la tabla ya existía sin ellas
        for col_migrate in [
            "ALTER TABLE forecast_excel_productos ADD COLUMN marca VARCHAR(255) DEFAULT NULL",
            "ALTER TABLE forecast_excel_productos ADD COLUMN modelo VARCHAR(255) DEFAULT NULL",
            "ALTER TABLE forecast_excel_productos ADD COLUMN categoria VARCHAR(255) DEFAULT NULL",
            "ALTER TABLE forecast_excel_productos ADD COLUMN precio_distribuidor DECIMAL(10,2) DEFAULT NULL",
            "ALTER TABLE forecast_excel_productos ADD COLUMN precio_partner DECIMAL(10,2) DEFAULT NULL",
            "ALTER TABLE forecast_excel_productos ADD COLUMN precio_partner_elite DECIMAL(10,2) DEFAULT NULL",
            "ALTER TABLE forecast_excel_productos ADD COLUMN precio_partner_elite_plus DECIMAL(10,2) DEFAULT NULL",
            "ALTER TABLE forecast_excel_productos ADD COLUMN precio_publico DECIMAL(10,2) DEFAULT NULL",
        ]:
            try:
                cur.execute(col_migrate)
                conn.commit()
            except Exception:
                pass
        # Migración: inferir marca desde el nombre para registros sin marca
        try:
            cur.execute("""
                UPDATE forecast_excel_productos
                SET marca = CASE
                    WHEN UPPER(nombre) LIKE '%SCOTT%'   THEN 'SCOTT'
                    WHEN UPPER(nombre) LIKE '%MEGAMO%'  THEN 'MEGAMO'
                    WHEN UPPER(nombre) LIKE '%SYNCROS%' THEN 'SYNCROS'
                END
                WHERE (marca IS NULL OR marca = '' OR marca = 'N/A')
                  AND (UPPER(nombre) LIKE '%SCOTT%' OR UPPER(nombre) LIKE '%MEGAMO%' OR UPPER(nombre) LIKE '%SYNCROS%')
            """)
            conn.commit()
        except Exception:
            conn.rollback()
        conn.commit()
    finally:
        cur.close()
        conn.close()


# ─────────────────────────────────────────────────────
# Public Functions
# ─────────────────────────────────────────────────────

def get_product_from_sources(sku: str) -> dict or None:
    """
    Busca un producto por SKU en ambas fuentes (Excel primero, luego Odoo).

    Args:
        sku: código del producto

    Returns:
        dict con keys: sku, nombre, color, talla, origen
        o None si no existe en ninguna fuente
    """
    conn = obtener_conexion()
    cur = conn.cursor(dictionary=True)
    try:
        # Buscar primero en Excel (tiene prioridad)
        cur.execute("""
            SELECT sku, nombre, color, talla, origen
            FROM forecast_excel_productos
            WHERE sku = %s AND origen = 'excel'
        """, (sku,))
        row = cur.fetchone()
        if row:
            return row

        # Fallback a Odoo catalog
        cur.execute("""
            SELECT referencia_interna AS sku,
                   nombre_producto AS nombre,
                   color,
                   talla,
                   'odoo' AS origen
            FROM odoo_catalogo
            WHERE referencia_interna = %s
        """, (sku,))
        row = cur.fetchone()
        return row
    finally:
        cur.close()
        conn.close()


def load_excel_products(file_content: bytes) -> dict:
    """
    Parsea y carga productos desde un archivo Excel.

    Estructura esperada:
    - Encabezado con: SKU, NOMBRE, [COLOR], [TALLA]

    Args:
        file_content: contenido del archivo Excel en bytes

    Returns:
        dict con keys:
        - 'success': bool
        - 'cargados': int
        - 'total_filas_procesadas': int
        - 'duplicados_actualizados': int
        - 'errores': list
        - 'message': str (si error)
    """
    ensure_excel_producto_table()
    result = {
        'success': False,
        'cargados': 0,
        'total_filas_procesadas': 0,
        'duplicados_actualizados': 0,
        'errores': []
    }

    if not OPENPYXL_OK:
        result['message'] = 'openpyxl no está instalado'
        return result

    try:
        # Parse Excel
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        ws = wb.active
    except Exception as e:
        result['message'] = f'No se pudo leer el archivo Excel: {str(e)}'
        logging.error('[load_excel_products] Parse error: %s', e)
        return result

    # Find header row
    header_row = None
    for r_idx in range(1, 10):
        val = ws.cell(row=r_idx, column=1).value
        if val and str(val).strip().upper() == 'SKU':
            header_row = r_idx
            break

    if header_row is None:
        result['message'] = 'Estructura inválida: no se encontró encabezado con "SKU"'
        return result

    # Map columns
    col_map = {}
    for c_idx in range(1, 10):
        header = ws.cell(row=header_row, column=c_idx).value
        if not header:
            break
        header_clean = str(header).strip().upper()
        col_map[header_clean] = c_idx

    required_cols = {'SKU', 'NOMBRE'}
    if not required_cols.issubset(set(col_map.keys())):
        result['message'] = f'Faltan columnas requeridas. Encontradas: {list(col_map.keys())}'
        return result

    # Parse rows
    productos = []
    errores = []
    skus_cargados = set()

    for r_idx in range(header_row + 1, ws.max_row + 1):
        sku = ws.cell(row=r_idx, column=col_map['SKU']).value
        nombre = ws.cell(row=r_idx, column=col_map['NOMBRE']).value

        if not sku or not nombre:
            continue

        sku = str(sku).strip()
        nombre = str(nombre).strip().upper()
        color = (str(ws.cell(row=r_idx, column=col_map.get('COLOR', 0)).value or '')).strip().upper()
        talla = (str(ws.cell(row=r_idx, column=col_map.get('TALLA', 0)).value or '')).strip().upper()
        marca  = (str(ws.cell(row=r_idx, column=col_map.get('MARCA',  0)).value or '')).strip().upper()
        modelo = (str(ws.cell(row=r_idx, column=col_map.get('MODELO', 0)).value or '')).strip().upper()
        if not marca:
            nombre_upper = nombre.upper()
            if 'SCOTT' in nombre_upper:    marca = 'SCOTT'
            elif 'MEGAMO' in nombre_upper: marca = 'MEGAMO'
            elif 'SYNCROS' in nombre_upper: marca = 'SYNCROS'

        if not sku:
            errores.append(f'Fila {r_idx}: SKU vacío')
            continue

        if sku in skus_cargados:
            errores.append(f'Fila {r_idx}: SKU "{sku}" duplicado dentro del archivo')
            continue

        skus_cargados.add(sku)
        productos.append({
            'sku':    sku,
            'nombre': nombre,
            'color':  color  or None,
            'talla':  talla  or None,
            'marca':  marca  or None,
            'modelo': modelo or None,
        })

    if not productos:
        result['message'] = 'No se encontraron productos válidos en el archivo'
        result['errores'] = errores
        return result

    # Check for existing SKUs
    conn = obtener_conexion()
    cur = conn.cursor(dictionary=True)
    try:
        placeholders = ','.join(['%s'] * len(skus_cargados))
        cur.execute(
            f"SELECT sku FROM forecast_excel_productos WHERE sku IN ({placeholders})",
            list(skus_cargados)
        )
        duplicados = set(r['sku'] for r in cur.fetchall())
    finally:
        cur.close()
        conn.close()

    # Upsert into database
    conn = obtener_conexion()
    cur = conn.cursor()
    try:
        for p in productos:
            cur.execute("""
                INSERT INTO forecast_excel_productos
                    (sku, nombre, color, talla, marca, modelo, origen, cargado_en)
                VALUES (%s, %s, %s, %s, %s, %s, 'excel', CURRENT_TIMESTAMP)
                ON DUPLICATE KEY UPDATE
                    nombre = VALUES(nombre),
                    color  = VALUES(color),
                    talla  = VALUES(talla),
                    marca  = VALUES(marca),
                    modelo = VALUES(modelo),
                    actualizado_en = CURRENT_TIMESTAMP
            """, (p['sku'], p['nombre'], p['color'], p['talla'], p['marca'], p['modelo']))
        conn.commit()
        logging.info('[load_excel_products] Loaded %d products', len(productos))
    except Exception as e:
        conn.rollback()
        logging.exception('[load_excel_products] Insert error: %s', e)
        result['message'] = f'Error al guardar: {str(e)}'
        return result
    finally:
        cur.close()
        conn.close()

    result['success'] = True
    result['cargados'] = len(productos)
    result['total_filas_procesadas'] = len(skus_cargados)
    result['duplicados_actualizados'] = len(duplicados)
    if errores:
        result['errores'] = errores

    return result


def list_excel_products(search: str = '', limit: int = 100, offset: int = 0) -> dict:
    """
    Lista productos del catálogo Excel.

    Args:
        search: búsqueda por SKU exacto o NOMBRE (fulltext)
        limit: máximo resultados
        offset: paginación

    Returns:
        dict con: total, productos, limit, offset
    """
    ensure_excel_producto_table()
    limit = min(limit, 500)
    conn = obtener_conexion()
    cur = conn.cursor(dictionary=True)
    try:
        if search:
            like = f'%{search}%'
            cur.execute("""
                SELECT sku, nombre, color, talla, cargado_en, actualizado_en
                FROM forecast_excel_productos
                WHERE origen = 'excel' AND (
                    sku LIKE %s
                    OR nombre LIKE %s
                )
                ORDER BY cargado_en DESC
                LIMIT %s OFFSET %s
            """, (like, like, limit, offset))
        else:
            cur.execute("""
                SELECT sku, nombre, color, talla, cargado_en, actualizado_en
                FROM forecast_excel_productos
                WHERE origen = 'excel'
                ORDER BY cargado_en DESC
                LIMIT %s OFFSET %s
            """, (limit, offset))

        productos = cur.fetchall()

        # Total count
        if search:
            like = f'%{search}%'
            cur.execute("""
                SELECT COUNT(*) as cnt
                FROM forecast_excel_productos
                WHERE origen = 'excel' AND (
                    sku LIKE %s
                    OR nombre LIKE %s
                )
            """, (like, like))
        else:
            cur.execute("SELECT COUNT(*) as cnt FROM forecast_excel_productos WHERE origen = 'excel'")

        total = cur.fetchone()['cnt']

        return {
            'total': total,
            'limit': limit,
            'offset': offset,
            'productos': productos
        }
    finally:
        cur.close()
        conn.close()


def delete_excel_product(sku: str) -> dict:
    """
    Elimina un producto del catálogo Excel.

    Args:
        sku: código del producto

    Returns:
        dict con: eliminado (bool), sku, message (si error)
    """
    sku = str(sku).strip()
    if not sku:
        return {'eliminado': False, 'message': 'SKU requerido'}

    conn = obtener_conexion()
    cur = conn.cursor()
    try:
        cur.execute(
            "DELETE FROM forecast_excel_productos WHERE sku = %s AND origen = 'excel'",
            (sku,)
        )
        if cur.rowcount == 0:
            return {'eliminado': False, 'message': f'Producto "{sku}" no encontrado'}
        conn.commit()
        logging.info('[delete_excel_product] Deleted SKU: %s', sku)
        return {'eliminado': True, 'sku': sku}
    except Exception as e:
        logging.exception('[delete_excel_product] Error: %s', e)
        return {'eliminado': False, 'message': f'Error: {str(e)}'}
    finally:
        cur.close()
        conn.close()


def clear_excel_catalog() -> dict:
    """
    Elimina TODOS los productos del catálogo Excel.

    Returns:
        dict con: eliminados (int), message
    """
    conn = obtener_conexion()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM forecast_excel_productos WHERE origen = 'excel'")
        count = cur.rowcount
        conn.commit()
        logging.warning('[clear_excel_catalog] Cleared %d products', count)
        return {'eliminados': count, 'mensaje': 'Catálogo Excel vaciado'}
    except Exception as e:
        logging.exception('[clear_excel_catalog] Error: %s', e)
        return {'eliminados': 0, 'message': f'Error: {str(e)}'}
    finally:
        cur.close()
        conn.close()


def load_csv_apparel_products(file_content: bytes, encoding: str = 'utf-8-sig') -> dict:
    """
    Parsea e importa productos desde el CSV de apparel.

    Formato esperado (columnas obligatorias):
        SKU, DESCRIPCION, COLOR, TALLA,
        DISTRIBUIDOR_SIN_IVA, PARTNER_SIN_IVA, PARTNER_ELITE_SIN_IVA,
        PARTNER_ELITE_PLUS_SIN_IVA, PRECIO_PUBLICO_SIN_IVA

    Columnas ignoradas: CODIGO_PRECIO, PRECIO_PUBLICO_CON_IVA.
    Los precios se guardan SIN IVA; listar_forecast los multiplica x 1.16 al servir.
    """
    ensure_excel_producto_table()
    result = {
        'success': False,
        'cargados': 0,
        'total_filas_procesadas': 0,
        'duplicados_actualizados': 0,
        'errores': []
    }

    try:
        text = file_content.decode(encoding, errors='replace')
        reader = csv.DictReader(io.StringIO(text))
    except Exception as e:
        result['message'] = f'No se pudo leer el CSV: {str(e)}'
        logging.error('[load_csv_apparel] Decode/parse error: %s', e)
        return result

    required_cols = {
        'SKU', 'DESCRIPCION', 'COLOR', 'TALLA',
        'DISTRIBUIDOR_SIN_IVA', 'PARTNER_SIN_IVA',
        'PARTNER_ELITE_SIN_IVA', 'PARTNER_ELITE_PLUS_SIN_IVA',
        'PRECIO_PUBLICO_SIN_IVA',
    }

    def _norm_key(k):
        return k.replace('\xa0', ' ').strip().upper()

    def _safe_price(val):
        try:
            v = float(str(val).replace(',', '').strip())
            return round(v, 4) if v else None
        except (ValueError, TypeError):
            return None

    productos = []
    errores = []
    skus_vistos = set()

    for row_num, row in enumerate(reader, start=2):
        row_norm = {_norm_key(k): (v or '').strip() for k, v in row.items()}

        missing = required_cols - set(row_norm.keys())
        if row_num == 2 and missing:
            result['message'] = (
                f'Columnas requeridas faltantes: {missing}. '
                f'Encontradas: {list(row_norm.keys())}'
            )
            return result

        sku = row_norm.get('SKU', '').strip()
        if not sku:
            continue

        result['total_filas_procesadas'] += 1

        if sku in skus_vistos:
            errores.append(f'Fila {row_num}: SKU "{sku}" duplicado en el archivo')
            continue
        skus_vistos.add(sku)

        descripcion = row_norm.get('DESCRIPCION', '').strip()
        color       = row_norm.get('COLOR', '').strip() or None
        talla       = row_norm.get('TALLA', '').strip() or None

        # Extraer categoria: todo lo que viene despues del codigo numerico inicial
        m = re.match(r'^\d+\s+(.+)$', descripcion)
        categoria = m.group(1).strip() if m else descripcion

        marca = 'SYNCROS' if 'SYNCROS' in descripcion.upper() else 'SCOTT'

        productos.append({
            'sku':                       sku,
            'nombre':                    descripcion,
            'color':                     color,
            'talla':                     talla,
            'marca':                     marca,
            'modelo':                    categoria,
            'categoria':                 categoria,
            'precio_distribuidor':       _safe_price(row_norm.get('DISTRIBUIDOR_SIN_IVA')),
            'precio_partner':            _safe_price(row_norm.get('PARTNER_SIN_IVA')),
            'precio_partner_elite':      _safe_price(row_norm.get('PARTNER_ELITE_SIN_IVA')),
            'precio_partner_elite_plus': _safe_price(row_norm.get('PARTNER_ELITE_PLUS_SIN_IVA')),
            'precio_publico':            _safe_price(row_norm.get('PRECIO_PUBLICO_SIN_IVA')),
        })

    if not productos:
        result['message'] = 'No se encontraron filas validas en el CSV'
        result['errores'] = errores
        return result

    # Detectar duplicados existentes en BD
    conn = obtener_conexion()
    cur = conn.cursor(dictionary=True)
    try:
        placeholders = ','.join(['%s'] * len(skus_vistos))
        cur.execute(
            f"SELECT sku FROM forecast_excel_productos WHERE sku IN ({placeholders})",
            list(skus_vistos)
        )
        duplicados = {r['sku'] for r in cur.fetchall()}
    finally:
        cur.close()
        conn.close()

    # Upsert en lotes de 500
    BATCH = 500
    conn = obtener_conexion()
    cur = conn.cursor()
    try:
        for i in range(0, len(productos), BATCH):
            batch = productos[i:i + BATCH]
            cur.executemany("""
                INSERT INTO forecast_excel_productos
                    (sku, nombre, color, talla, marca, modelo, categoria,
                     precio_distribuidor, precio_partner, precio_partner_elite,
                     precio_partner_elite_plus, precio_publico,
                     origen, cargado_en)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'excel', CURRENT_TIMESTAMP)
                ON DUPLICATE KEY UPDATE
                    nombre                    = VALUES(nombre),
                    color                     = VALUES(color),
                    talla                     = VALUES(talla),
                    marca                     = VALUES(marca),
                    modelo                    = VALUES(modelo),
                    categoria                 = VALUES(categoria),
                    precio_distribuidor       = VALUES(precio_distribuidor),
                    precio_partner            = VALUES(precio_partner),
                    precio_partner_elite      = VALUES(precio_partner_elite),
                    precio_partner_elite_plus = VALUES(precio_partner_elite_plus),
                    precio_publico            = VALUES(precio_publico),
                    actualizado_en            = CURRENT_TIMESTAMP
            """, [(
                p['sku'], p['nombre'], p['color'], p['talla'],
                p['marca'], p['modelo'], p['categoria'],
                p['precio_distribuidor'], p['precio_partner'],
                p['precio_partner_elite'], p['precio_partner_elite_plus'],
                p['precio_publico'],
            ) for p in batch])
        conn.commit()
        logging.info('[load_csv_apparel] Imported %d apparel SKUs', len(productos))
    except Exception as e:
        conn.rollback()
        logging.exception('[load_csv_apparel] Insert error: %s', e)
        result['message'] = f'Error al guardar: {str(e)}'
        return result
    finally:
        cur.close()
        conn.close()

    result['success'] = True
    result['cargados'] = len(productos)
    result['duplicados_actualizados'] = len(duplicados)
    if errores:
        result['errores'] = errores
    return result


def get_valid_skus() -> set:
    """
    Returns the set of valid SKUs: whitelist + Excel catalog merged.
    Falls back to odoo_catalogo only when neither source has data.
    """
    conn = obtener_conexion()
    cur  = conn.cursor(dictionary=True)
    try:
        # 1. Whitelist (bicicletas Scott/Megamo del catálogo oficial)
        whitelist: set = set()
        try:
            cur.execute("SELECT sku FROM forecast_sku_whitelist")
            whitelist = set(r['sku'] for r in cur.fetchall())
        except Exception:
            pass  # tabla puede no existir aún en entornos recién migrados

        # 2. Catálogo Excel (apparel y otros productos cargados manualmente)
        excel_skus: set = set()
        try:
            cur.execute("SELECT DISTINCT sku FROM forecast_excel_productos WHERE origen = 'excel'")
            excel_skus = set(r['sku'] for r in cur.fetchall())
        except Exception:
            pass

        # Ambas fuentes son válidas simultáneamente
        combined = whitelist | excel_skus
        if combined:
            return combined

        # 3. Fallback: catálogo Odoo sincronizado (sin whitelist ni Excel)
        cur.execute("SELECT referencia_interna FROM odoo_catalogo")
        return set(r['referencia_interna'] for r in cur.fetchall())
    finally:
        cur.close()
        conn.close()
