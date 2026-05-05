from __future__ import annotations
import io
import logging
import time
import unicodedata
from utils.tiempo import ahora_str

try:
    import pandas as pd
    PANDAS_OK = True
except Exception as e:
    logging.warning(f"Pandas import error: {e}")
    pd = None  # type: ignore
    PANDAS_OK = False

import requests
from openpyxl.styles import Alignment, Font, PatternFill


def _norm(s: str) -> str:
    return unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii").upper().strip()


def _find_col(df, target: str):
    if not PANDAS_OK or pd is None:
        return None
    target_n = _norm(target)
    for col in df.columns:
        if _norm(col) == target_n:
            return col
    return None


SHEET_CSV_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1Uv2ygqBPRkGJqJlu98XLZRy6hSRqrmudttsDQA6jW3I"
    "/export?format=csv&sheet=SUGERIDO+GARANTIAS+(NO+ACTIVO)"
)
CACHE_TTL = 300

_cache: dict = {"df": None, "ts": 0.0}

MESES_ORDER = [
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
]


def _fetch_df():
    if not PANDAS_OK or pd is None:
        return None
    global _cache
    now = time.time()
    if _cache["df"] is not None and (now - _cache["ts"]) < CACHE_TTL:
        return _cache["df"]

    resp = requests.get(SHEET_CSV_URL, timeout=20, allow_redirects=True)
    resp.raise_for_status()

    # Las cabeceras reales están en la fila 5 (igual que la otra hoja)
    df = pd.read_csv(io.BytesIO(resp.content), dtype=str, header=5, encoding="utf-8")
    df.columns = [str(c).strip() for c in df.columns]
    df.dropna(how="all", inplace=True)

    _cache = {"df": df, "ts": now}
    logging.info("Garantías: CSV cargado — %d filas", len(df))
    return df


def _to_num(series):
    if not PANDAS_OK or pd is None:
        return None
    return pd.to_numeric(series, errors="coerce")


def invalidar_cache() -> None:
    _cache["ts"] = 0.0


def get_dashboard_data() -> dict:
    if not PANDAS_OK or pd is None:
        logging.warning("Pandas no disponible - retornando datos vacíos de garantías")
        return {
            "total": 0, "cerradas": 0, "pct_cierre": 0.0, "lat_promedio": 0.0,
            "por_estatus": {}, "lat_mensual": {}, "lat_cliente": {},
            "gar_cliente": {}, "desc_dano": {}, "ubic_dano": {},
            "por_mes": {}
        }

    df = _fetch_df().copy()

    def col(name: str) -> pd.Series:
        real = _find_col(df, name)
        if real:
            return df[real]
        return pd.Series([""] * len(df), dtype=str)

    # ── Estatus (col AW = "ESTATUS", valores: Cerrada / Abierta) ──────────
    col_estatus_real = _find_col(df, "ESTATUS") or _find_col(df, "Estatus")
    df["_estatus"] = (
        df[col_estatus_real].fillna("Sin estatus").str.strip()
        if col_estatus_real
        else pd.Series(["Sin estatus"] * len(df))
    )

    # ── Latencia de atención (col AF = "LATENCIA RECEPCION DOC. CORRECTOS") ─
    col_lat_at = _find_col(df, "LATENCIA RECEPCION DOC. CORRECTOS")
    df["_lat_atencion"] = _to_num(df[col_lat_at]) if col_lat_at else pd.Series(dtype=float)

    # ── Latencia total (col AZ = "LATENCIA TOTAL") para el KPI global ────
    col_lat_tot = _find_col(df, "LATENCIA TOTAL")
    df["_lat_total"] = (
        _to_num(df[col_lat_tot]) if col_lat_tot else df["_lat_atencion"]
    )

    df["_cliente"] = col("RAZON SOCIAL").fillna("Sin cliente").str.strip()
    df["_dano"]    = col("DESCRIPCION DEL DANO").fillna("Sin descripción").str.strip()
    df["_ubic"]    = col("UBICACION DEL DANO").fillna("Sin ubicación").str.strip()
    df["_mes"]     = col("MES").fillna("").str.strip().str.lower()

    total    = len(df)
    cerradas = int(df["_estatus"].str.lower().isin(["cerrada", "cerrado"]).sum())
    lat_val  = df["_lat_total"].mean()
    lat_prom = round(float(lat_val), 1) if pd.notna(lat_val) else 0.0

    # 1. Conteo por estatus (col AW: Cerrada / Abierta)
    por_estatus = df["_estatus"].value_counts().head(15).to_dict()

    # 2. Latencia de atención mensual (usa col AF)
    lat_mensual_raw = (
        df.groupby("_mes")["_lat_atencion"].mean().dropna().round(1).to_dict()
    )
    lat_mensual: dict = {}
    for m in MESES_ORDER:
        if m in lat_mensual_raw:
            lat_mensual[m.capitalize()] = lat_mensual_raw[m]
    for k, v in lat_mensual_raw.items():
        label = k.capitalize()
        if label not in lat_mensual:
            lat_mensual[label] = v

    # 3. Latencia de atención por cliente — top 30 (usa col AF)
    lat_cliente = (
        df.groupby("_cliente")["_lat_atencion"]
        .mean().dropna().round(1)
        .sort_values(ascending=False)
        .head(30).to_dict()
    )

    # 4. Garantías por cliente — top 30
    gar_cliente = df["_cliente"].value_counts().head(30).to_dict()

    # 5. Descripción del daño — top 25
    desc_dano = df["_dano"].value_counts().head(25).to_dict()

    # 6. Ubicación del daño — top 25
    ubic_dano = df["_ubic"].value_counts().head(25).to_dict()

    return {
        "kpis": {
            "total":             int(total),
            "cerradas":          cerradas,
            "en_proceso":        int(total - cerradas),
            "latencia_promedio": lat_prom,
        },
        "por_estatus":           por_estatus,
        "latencia_mensual":      lat_mensual,
        "latencia_por_cliente":  lat_cliente,
        "garantias_por_cliente": gar_cliente,
        "descripcion_dano":      desc_dano,
        "ubicacion_dano":        ubic_dano,
        "ultima_actualizacion":  ahora_str("%d/%m/%Y %H:%M"),
    }


def exportar_excel() -> bytes:
    if not PANDAS_OK or pd is None:
        logging.warning("Pandas no disponible - no se puede exportar Excel")
        return b""

    df = _fetch_df().copy()
    if df is None:
        return b""
    cols_excluir = [c for c in df.columns if "FOTOGRAFIA" in c.upper() or "EVIDENCIA" in c.upper()]
    df.drop(columns=cols_excluir, inplace=True, errors="ignore")

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Garantías")
        ws = writer.sheets["Garantías"]

        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill("solid", fgColor="1A3C5E")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 45)

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30

    return buf.getvalue()
