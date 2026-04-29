#!/usr/bin/env python3
"""
Script para actualizar instrucciones en el Excel de proyecciones
Agrega texto sobre prioridad y fecha límite (30 de abril del 2026)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# Ruta del archivo
excel_path = r"C:\Users\Usuario\Desktop\MONITOR\EB_BACK\2026.04.23 FORMATO PARA ORDENAR SCALE Y CONTRAIL MY27 PARA DISTRIBUIDOR (1).xlsx"

try:
    # Cargar el workbook
    print("Cargando archivo Excel...")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Buscar la fila donde ya hay instrucciones o un lugar apropiado
    # Generalmente está en las primeras filas o en una sección de instrucciones
    # Vamos a agregar en la fila 2, columna A

    # Crear el texto de instrucciones
    instrucciones = (
        "INSTRUCCIONES IMPORTANTES:\n"
        "• Entre más rápido envíen la proyección recibirán PRIORIDAD en sus solicitudes\n"
        "• Las proyecciones deben enviarse a su Ejecutivo de Ventas\n"
        "• FECHA LÍMITE: 30 de Abril del 2026"
    )

    # Buscar una celda disponible para las instrucciones (revisamos columnas vacías)
    # Insertaremos una fila nueva en la posición 1
    ws.insert_rows(1, 5)  # Insertamos 5 filas para el texto

    # Agregar el texto
    cell = ws['A1']
    cell.value = instrucciones

    # Formatear la celda
    cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Rojo
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Ajustar altura de filas para el texto
    for i in range(1, 6):
        ws.row_dimensions[i].height = 20

    # Guardar el archivo
    print("Guardando archivo actualizado...")
    wb.save(excel_path)

    print("[OK] Instrucciones agregadas correctamente al Excel")
    print(f"Archivo guardado en: {excel_path}")

except FileNotFoundError:
    print(f"[ERROR] Archivo no encontrado en {excel_path}")
except Exception as e:
    print(f"[ERROR] Error al actualizar el Excel: {str(e)}")
